from flask import Flask, render_template, request, redirect, url_for, session, flash, current_app
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix
from datetime import datetime, date, timedelta
import base64
import hashlib
import hmac
from flask import session 
from flask_login import current_user, login_required
import os
import re
import uuid
import time
import psycopg2
from cryptography.fernet import Fernet, InvalidToken
from prometheus_client import CONTENT_TYPE_LATEST, Counter, Gauge, Histogram, generate_latest
from dotenv import load_dotenv
import traceback
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import random
import string
import reportlab
from flask import send_file
from flask_cors import CORS
from flasgger import Swagger, swag_from
from flask_jwt_extended import (
    JWTManager,
    create_access_token,
    create_refresh_token,
    get_jwt,
    get_jwt_identity,
    jwt_required,
)
from flask_login import LoginManager, login_required, current_user, login_user, logout_user, UserMixin
from flask_login import login_user, logout_user
from webauthn import (
    generate_authentication_options,
    generate_registration_options,
    verify_authentication_response,
    verify_registration_response,
)
from webauthn.helpers import base64url_to_bytes, bytes_to_base64url, options_to_json
from webauthn.helpers.structs import (
    AuthenticatorSelectionCriteria,
    PublicKeyCredentialDescriptor,
    ResidentKeyRequirement,
    UserVerificationRequirement,
)

# ==================== CREAR APP (SOLO UNA VEZ) ====================
app = Flask(__name__)
# Render termina HTTPS delante de Gunicorn; respeta el esquema y host públicos.
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
load_dotenv()
app.secret_key = os.getenv('SECRET_KEY', 'upq_bolsa_trabajo_secret_key')
app.config['JWT_SECRET_KEY'] = os.getenv('JWT_SECRET_KEY', app.secret_key)
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = 3600
app.config['JWT_REFRESH_TOKEN_EXPIRES'] = 2592000
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)
jwt = JWTManager(app)


def webauthn_config():
    """Devuelve el dominio y origen autorizados para las passkeys."""
    host = request.host.split(':', 1)[0]
    return (
        os.getenv('WEBAUTHN_RP_ID', host).strip(),
        os.getenv('WEBAUTHN_ORIGIN', f'{request.scheme}://{request.host}').strip(),
    )

# ==================== CIFRADO DE DATOS SENSIBLES ====================
# En producción DATA_ENCRYPTION_KEY debe ser un secreto aleatorio independiente.
# Se deriva una llave Fernet válida para admitir los secretos generados por Render.
_configured_encryption_key = os.getenv('DATA_ENCRYPTION_KEY', '').strip()
_key_material = _configured_encryption_key or app.secret_key
_fernet_key = base64.urlsafe_b64encode(hashlib.sha256(_key_material.encode()).digest())
data_cipher = Fernet(_fernet_key)


def encrypt_sensitive(value):
    """Cifra texto con AES-128-CBC + HMAC mediante el formato autenticado Fernet."""
    if value is None:
        return None
    return data_cipher.encrypt(str(value).encode()).decode()


def decrypt_sensitive(value):
    """Descifra texto Fernet; acepta valores antiguos sin cifrar para migración gradual."""
    if value is None:
        return None
    try:
        return data_cipher.decrypt(str(value).encode()).decode()
    except (InvalidToken, ValueError):
        return str(value)


def secure_equals_encrypted(encrypted_value, candidate):
    return hmac.compare_digest(decrypt_sensitive(encrypted_value) or '', str(candidate))


def profile_photo_url(value):
    """Resuelve fotos persistentes en BD y nombres de archivo heredados."""
    if not value:
        return url_for('static', filename='images/default-profile.png')
    value = str(value)
    if value.startswith(('data:image/', 'https://', 'http://')):
        return value
    return url_for('static', filename=f'uploads/{value}')


app.jinja_env.globals['profile_photo_url'] = profile_photo_url


# ==================== MÉTRICAS PROMETHEUS ====================
HTTP_REQUESTS = Counter(
    'talentupq_http_requests_total', 'Peticiones HTTP recibidas', ['method', 'endpoint', 'status']
)
HTTP_LATENCY = Histogram(
    'talentupq_http_request_duration_seconds', 'Duración de peticiones HTTP', ['method', 'endpoint']
)
ACTIVE_REQUESTS = Gauge('talentupq_http_requests_active', 'Peticiones HTTP activas')
DB_HEALTH = Gauge('talentupq_database_available', 'Disponibilidad de PostgreSQL (1=disponible)')


@app.before_request
def start_request_metrics():
    request._metrics_started_at = time.monotonic()
    ACTIVE_REQUESTS.inc()


@app.after_request
def finish_request_metrics(response):
    endpoint = request.endpoint or 'not_found'
    HTTP_REQUESTS.labels(request.method, endpoint, response.status_code).inc()
    started_at = getattr(request, '_metrics_started_at', time.monotonic())
    HTTP_LATENCY.labels(request.method, endpoint).observe(time.monotonic() - started_at)
    ACTIVE_REQUESTS.dec()
    return response

# ==================== CONFIGURACIÓN FLASK-LOGIN ====================
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor inicia sesión para acceder a esta página.'
login_manager.login_message_category = 'warning'

@login_manager.user_loader
def load_user(user_id):
    """Carga un usuario por su ID para Flask-Login"""
    try:
        result = execute_query("SELECT * FROM Usuarios WHERE UsuarioID = ?", [user_id])
        if result:
            usuario = result[0]
            return User(
                id=usuario['UsuarioID'],
                email=usuario['Email'],
                tipo=usuario['TipoUsuario'],
                activo=usuario.get('Activo', 1)
            )
        return None
    except Exception as e:
        print(f"Error loading user: {e}")
        return None

# ==================== CLASE USER PARA FLASK-LOGIN ====================
class User(UserMixin):
    """Clase para manejar usuarios en Flask-Login"""
    def __init__(self, id, email, tipo, activo=True):
        self.id = id
        self.email = email
        self.tipo = tipo
        self.activo = activo
    
    def get_id(self):
        return str(self.id)
    
    @property
    def is_active(self):
        return self.activo == 1 or self.activo == True

# ==================== CORS ====================
CORS(app, origins=['http://localhost:3000', 'http://127.0.0.1:3000'])

# ==================== SWAGGER ====================
app.config['SWAGGER'] = {
    'title': 'TalentUPQ API',
    'version': '1.0.0',
    'description': 'API para la bolsa de trabajo TalentUPQ - Universidad Politécnica de Querétaro',
    'termsOfService': '/terms',
    'contact': {
        'name': 'TalentUPQ',
        'email': 'bolsa.trabajo@upq.edu.mx',
        'url': 'http://127.0.0.1:5000/'
    },
    'license': {
        'name': 'MIT',
        'url': 'https://opensource.org/licenses/MIT'
    },
    'specs_route': '/apidocs/',
    'uiversion': 3,
    'swagger_ui': True,
    'specs': [
        {
            'endpoint': 'apispec',
            'route': '/apispec.json',
            'rule_filter': lambda rule: True,
            'model_filter': lambda tag: True,
        }
    ],
    'static_url_path': '/flasgger_static',
    'swagger_ui_bundle_js': '//unpkg.com/swagger-ui-dist@3/swagger-ui-bundle.js',
    'swagger_ui_standalone_preset_js': '//unpkg.com/swagger-ui-dist@3/swagger-ui-standalone-preset.js',
    'swagger_ui_css': '//unpkg.com/swagger-ui-dist@3/swagger-ui.css',
}

swagger = Swagger(app)

# ==================== CONFIGURACIONES ADICIONALES ====================
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['ALLOWED_EXTENSIONS'] = {'pdf', 'png', 'jpg', 'jpeg'}

app.config['DB_HOST'] = os.getenv('DB_HOST', 'localhost')
app.config['DB_PORT'] = int(os.getenv('DB_PORT', '5432'))
app.config['DB_NAME'] = os.getenv('DB_NAME', 'bolsatrabajoupq')
app.config['DB_USER'] = os.getenv('DB_USER', os.getenv('USER', 'postgres'))
app.config['DB_PASSWORD'] = os.getenv('DB_PASSWORD', '')

app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')





def enviar_correo_bienvenida(email_usuario, nombre_usuario, tipo_usuario):
    """Envía un correo de bienvenida al usuario registrado"""
    try:
        
        if not app.config.get('MAIL_USERNAME') or app.config['MAIL_USERNAME'] == 'tu_email@gmail.com':
            print("❌ ERROR: Configuración de correo no establecida correctamente")
            print("⚠️  Configura tu GMAIL real y contraseña de aplicación")
            return False
        
        print(f"📧 Intentando enviar correo a: {email_usuario}")
        print(f"📧 Usando cuenta: {app.config['MAIL_USERNAME']}")
        
        msg = MIMEMultipart()
        msg['From'] = f'TalentUPQ <{app.config["MAIL_USERNAME"]}>'
        msg['To'] = email_usuario
        
        if tipo_usuario == 'candidato':
            msg['Subject'] = f'¡Bienvenido a TalentUPQ, {nombre_usuario.split()[0]}!'
        elif tipo_usuario == 'empresa':
            msg['Subject'] = f'¡Bienvenida a TalentUPQ, {nombre_usuario}!'
        else:
            msg['Subject'] = f'¡Bienvenido a TalentUPQ!'
        
        
        if tipo_usuario == 'candidato':
            cuerpo = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    body {{ font-family: 'Arial', sans-serif; line-height: 1.6; color: #333; max-width: 600px; margin: 0 auto; padding: 20px; background: #f8f9fa; }}
                    .header {{ background: linear-gradient(135deg, #3498db 0%, #2ecc71 100%); color: white; padding: 40px; text-align: center; border-radius: 10px; margin-bottom: 30px; }}
                    .content {{ background: white; padding: 30px; border-radius: 10px; box-shadow: 0 5px 15px rgba(0,0,0,0.1); }}
                    .footer {{ margin-top: 30px; padding-top: 20px; border-top: 2px solid #3498db; color: #5a6c7d; font-size: 13px; text-align: center; }}
                    .btn {{ display: inline-block; background: linear-gradient(135deg, #3498db, #2ecc71); color: white; padding: 12px 30px; text-decoration: none; border-radius: 25px; margin: 15px 0; font-weight: bold; }}
                    .steps {{ margin: 25px 0; }}
                    .step {{ background: #f1f8ff; padding: 15px; border-left: 4px solid #3498db; margin-bottom: 10px; border-radius: 5px; }}
                    .highlight {{ background: #e8f6ff; padding: 15px; border-radius: 8px; margin: 20px 0; border: 2px dashed #3498db; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>🎓 ¡Bienvenido a TalentUPQ!</h1>
                    <p>Tu puente hacia oportunidades profesionales</p>
                </div>
                
                <div class="content">
                    <h2>Hola {nombre_usuario},</h2>
                    <p>¡Gracias por registrarte en <strong>TalentUPQ</strong>, la bolsa de trabajo oficial de la Universidad Politécnica de Querétaro!</p>
                    
                    <p>Tu cuenta ha sido creada exitosamente:</p>
                    <ul>
                        <li><strong>Nombre:</strong> {nombre_usuario}</li>
                        <li><strong>Email:</strong> {email_usuario}</li>
                        <li><strong>Tipo de cuenta:</strong> Candidato</li>
                        <li><strong>Fecha de registro:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M')}</li>
                    </ul>
                    
                    <div class="highlight">
                        <p>🌟 <strong>¡Tu próximo paso importante!</strong></p>
                        <p>Completa tu perfil profesional para aumentar tus posibilidades de ser contratado:</p>
                    </div>
                    
                    <p><strong>¿Qué puedes hacer ahora?</strong></p>
                    <div class="steps">
                        <div class="step">
                            <strong>1. Completa tu perfil</strong><br>
                            Sube tu CV, foto profesional y agrega tus habilidades
                        </div>
                        <div class="step">
                            <strong>2. Agrega tu experiencia</strong><br>
                            Destaca tu trayectoria profesional y académica
                        </div>
                        <div class="step">
                            <strong>3. Busca vacantes</strong><br>
                            Encuentra oportunidades que se ajusten a tu perfil
                        </div>
                        <div class="step">
                            <strong>4. Postúlate</strong><br>
                            Aplica a las mejores empresas de la región
                        </div>
                    </div>
                    
                    <div style="text-align: center; margin: 30px 0;">
                        <a href="{url_for('candidato_perfil', _external=True)}" class="btn">
                            ✨ Completar Mi Perfil
                        </a>
                    </div>
                    
                    <p><strong>Beneficios exclusivos para candidatos UPQ:</strong></p>
                    <ul>
                        <li>✅ Acceso a vacantes exclusivas para egresados UPQ</li>
                        <li>✅ Conexión directa con empresas aliadas</li>
                        <li>✅ Asesoría profesional y consejería</li>
                        <li>✅ Eventos de networking y ferias de empleo</li>
                        <li>✅ Seguimiento personalizado de tus postulaciones</li>
                    </ul>
                </div>
                
                <div class="footer">
                    <p><strong>TalentUPQ - Bolsa de Trabajo UPQ</strong></p>
                    <p>📍 Universidad Politécnica de Querétaro</p>
                    <p>📞 (773) 108-7368 | ✉️ bolsa.trabajo@upq.edu.mx</p>
                    <p>© {datetime.now().year} TalentUPQ. Todos los derechos reservados.</p>
                </div>
            </body>
            </html>
            """
        elif tipo_usuario == 'empresa':
            cuerpo = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    body {{ font-family: 'Arial', sans-serif; line-height: 1.6; color: #333; max-width: 600px; margin: 0 auto; padding: 20px; background: #f8f9fa; }}
                    .header {{ background: linear-gradient(135deg, #9b59b6 0%, #3498db 100%); color: white; padding: 40px; text-align: center; border-radius: 10px; margin-bottom: 30px; }}
                    .content {{ background: white; padding: 30px; border-radius: 10px; box-shadow: 0 5px 15px rgba(0,0,0,0.1); }}
                    .footer {{ margin-top: 30px; padding-top: 20px; border-top: 2px solid #9b59b6; color: #5a6c7d; font-size: 13px; text-align: center; }}
                    .btn {{ display: inline-block; background: linear-gradient(135deg, #9b59b6, #3498db); color: white; padding: 12px 30px; text-decoration: none; border-radius: 25px; margin: 15px 0; font-weight: bold; }}
                    .benefits {{ margin: 25px 0; }}
                    .benefit {{ background: #f5f0fa; padding: 15px; border-left: 4px solid #9b59b6; margin-bottom: 10px; border-radius: 5px; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>🏢 ¡Bienvenida a TalentUPQ, {nombre_usuario}!</h1>
                    <p>Encuentra al talento que tu empresa necesita</p>
                </div>
                
                <div class="content">
                    <h2>Estimado equipo de {nombre_usuario},</h2>
                    <p>¡Gracias por unirse a <strong>TalentUPQ</strong>, la bolsa de trabajo oficial de la Universidad Politécnica de Querétaro!</p>
                    
                    <p>Su cuenta de empresa ha sido creada exitosamente:</p>
                    <ul>
                        <li><strong>Empresa:</strong> {nombre_usuario}</li>
                        <li><strong>Email:</strong> {email_usuario}</li>
                        <li><strong>Tipo de cuenta:</strong> Empresa</li>
                        <li><strong>Fecha de registro:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M')}</li>
                    </ul>
                    
                    <div style="background: #e8f4fc; padding: 20px; border-radius: 8px; margin: 20px 0; border: 2px solid #3498db;">
                        <p>💼 <strong>¡Comience a publicar vacantes!</strong></p>
                        <p>Acceda a nuestro talento calificado de egresados UPQ</p>
                    </div>
                    
                    <p><strong>Beneficios para empresas aliadas:</strong></p>
                    <div class="benefits">
                        <div class="benefit">
                            <strong>Acceso a talento especializado</strong><br>
                            Egresados de ingenierías y licenciaturas UPQ
                        </div>
                        <div class="benefit">
                            <strong>Proceso de selección optimizado</strong><br>
                            Filtros inteligentes y perfiles detallados
                        </div>
                        <div class="benefit">
                            <strong>Eventos exclusivos</strong><br>
                            Ferias de empleo y días de entrevistas
                        </div>
                        <div class="benefit">
                            <strong>Soporte personalizado</strong><br>
                            Asesoría en reclutamiento y selección
                        </div>
                    </div>
                    
                    <div style="text-align: center; margin: 30px 0;">
                        <a href="{url_for('empresa_dashboard', _external=True)}" class="btn">
                            📝 Publicar Mi Primera Vacante
                        </a>
                    </div>
                    
                    <p><strong>Próximos pasos recomendados:</strong></p>
                    <ol>
                        <li>Complete el perfil de su empresa</li>
                        <li>Suba su logo y descripción corporativa</li>
                        <li>Publique su primera vacante (aprobación requerida)</li>
                        <li>Revise candidatos y programe entrevistas</li>
                    </ol>
                </div>
                
                <div class="footer">
                    <p><strong>TalentUPQ - Bolsa de Trabajo UPQ</strong></p>
                    <p>📍 Universidad Politécnica de Querétaro</p>
                    <p>📞 (442) 192-1200 | ✉️ bolsa.trabajo@upq.edu.mx</p>
                    <p>© {datetime.now().year} TalentUPQ. Todos los derechos reservados.</p>
                </div>
            </body>
            </html>
            """
        
        msg.attach(MIMEText(cuerpo, 'html'))
        
        print(f"📤 Conectando a SMTP: {app.config['MAIL_SERVER']}:{app.config['MAIL_PORT']}")
        
        server = smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT'])
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])
        
        print("✅ Login exitoso")
        
        server.send_message(msg)
        server.quit()
        
        print(f"✅ Correo de bienvenida enviado exitosamente a: {email_usuario}")
        return True
        
    except smtplib.SMTPAuthenticationError as e:
        print(f"❌ ERROR DE AUTENTICACIÓN: {e}")
        print("🔑 Verifica:")
        print("   1. Tu correo GMAIL real")
        print("   2. Contraseña de APLICACIÓN (no la normal)")
        print("   3. Verifica que tengas activada la verificación en 2 pasos")
        return False
    except Exception as e:
        print(f"❌ ERROR enviando correo: {str(e)}")
        import traceback
        traceback.print_exc()
        return False
    

@app.route('/test_correo/<email>')
def test_correo(email):
    """Ruta para probar el envío de correos"""
    try:
        print(f"🧪 Probando envío de correo a: {email}")
        
        if not app.config.get('MAIL_USERNAME'):
            return jsonify({
                'success': False,
                'message': 'Configuración de correo no encontrada'
            })
        
        resultado = enviar_correo_bienvenida(email, "Usuario de Prueba", "candidato")
        
        if resultado:
            return jsonify({
                'success': True, 
                'message': '✅ Correo enviado correctamente',
                'to': email,
                'from': app.config['MAIL_USERNAME']
            })
        else:
            return jsonify({
                'success': False, 
                'message': '❌ Error enviando correo'
            })
    except Exception as e:
        return jsonify({
            'success': False, 
            'message': f'❌ Error: {str(e)}'
        })



def generar_codigo_recuperacion():
    """Genera un código de 6 dígitos para recuperación"""
    return ''.join(random.choices(string.digits, k=6))

def enviar_codigo_recuperacion(email, codigo):
    """Envía correo con código de verificación"""
    try:
        usuario = execute_query(
            "SELECT UsuarioID, Email FROM Usuarios WHERE Email = ?",
            (email,)
        )
        
        if not usuario:
            return False
        
        msg = MIMEMultipart()
        msg['From'] = f'TalentUPQ <{app.config["MAIL_USERNAME"]}>'
        msg['To'] = email
        msg['Subject'] = 'Código de recuperación - TalentUPQ'
        
        cuerpo = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: 'Arial', sans-serif; line-height: 1.6; color: #333; max-width: 600px; margin: 0 auto; padding: 20px; background: #f8f9fa; }}
                .header {{ background: linear-gradient(135deg, #3498db 0%, #2ecc71 100%); color: white; padding: 30px; text-align: center; border-radius: 10px; margin-bottom: 30px; }}
                .content {{ background: white; padding: 30px; border-radius: 10px; box-shadow: 0 5px 15px rgba(0,0,0,0.1); text-align: center; }}
                .code {{ background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%); padding: 25px; font-family: monospace; font-size: 32px; font-weight: bold; text-align: center; letter-spacing: 8px; border-radius: 12px; margin: 20px 0; color: #2c3e50; border: 2px dashed #3498db; }}
                .footer {{ margin-top: 30px; padding-top: 20px; border-top: 2px solid #3498db; color: #5a6c7d; font-size: 13px; text-align: center; }}
                .warning {{ background: #fff3cd; border: 1px solid #ffeaa7; padding: 15px; border-radius: 8px; margin: 20px 0; color: #856404; font-size: 14px; }}
                .btn {{ display: inline-block; background: #3498db; color: white; padding: 12px 30px; text-decoration: none; border-radius: 25px; margin: 20px 0; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>🔐 Código de Recuperación</h1>
                <p>TalentUPQ - Bolsa de Trabajo UPQ</p>
            </div>
            
            <div class="content">
                <h2>¡Hola!</h2>
                <p>Hemos recibido una solicitud para restablecer la contraseña de tu cuenta en <strong>TalentUPQ</strong>.</p>
                
                <div class="warning">
                    <p>⚠️ <strong>Importante:</strong> Si no solicitaste este cambio, ignora este correo. Tu contraseña permanecerá segura.</p>
                </div>
                
                <p>Tu código de verificación es:</p>
                
                <div class="code">
                    {codigo}
                </div>
                
                <p>Este código expirará en <strong>10 minutos</strong> por razones de seguridad.</p>
                <p>Ingresa este código en la página de recuperación para restablecer tu contraseña.</p>
            </div>
            
            <div class="footer">
                <p><strong>TalentUPQ - Bolsa de Trabajo UPQ</strong></p>
                <p>📍 Universidad Politécnica de Querétaro</p>
                <p>📞 (773) 108-7368 | ✉️ bolsa.trabajo@upq.edu.mx</p>
            </div>
        </body>
        </html>
        """
        
        msg.attach(MIMEText(cuerpo, 'html'))
        
        server = smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT'])
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])
        
        server.send_message(msg)
        server.quit()
        
        print(f"✅ Código de recuperación enviado a: {email}")
        return True
        
    except Exception as e:
        print(f"❌ Error enviando código de recuperación: {str(e)}")
        return False

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        
        if not email:
            flash('Por favor ingresa tu correo electrónico', 'error')
            return redirect(url_for('forgot_password'))
        
        # Verificar si el email existe
        usuario = execute_query(
            "SELECT UsuarioID, Email FROM Usuarios WHERE Email = ?",
            (email,)
        )
        
        if usuario:
            # Generar código de 6 dígitos
            codigo = generar_codigo_recuperacion()
            fecha_expiracion = datetime.now()
            
            # Guardar código en la base de datos
            execute_query(
                """UPDATE Usuarios SET 
                ResetToken = ?, 
                ResetTokenExpira = DATEADD(MINUTE, 10, GETDATE())
                WHERE Email = ?""",
                (encrypt_sensitive(codigo), email),
                fetch=False
            )
            
            # Enviar correo con el código
            if enviar_codigo_recuperacion(email, codigo):
                # Guardar email en sesión temporal
                session['reset_email'] = email
                flash('Se ha enviado un código de verificación a tu correo electrónico. Revisa tu bandeja de entrada.', 'success')
                return redirect(url_for('verify_code'))
            else:
                flash('Hubo un problema al enviar el código. Por favor intenta más tarde.', 'error')
        else:
            # Por seguridad, no revelar si el email existe o no
            flash('Si el correo existe en nuestro sistema, recibirás un código de verificación.', 'info')
            return redirect(url_for('forgot_password'))
    
    return render_template('forgot_password.html')

@app.route('/verify_code', methods=['GET', 'POST'])
def verify_code():
    # Verificar que hay un email en sesión
    if 'reset_email' not in session:
        flash('Por favor inicia el proceso de recuperación primero.', 'error')
        return redirect(url_for('forgot_password'))
    
    email = session['reset_email']
    
    if request.method == 'POST':
        codigo = request.form.get('codigo', '').strip()
        
        if not codigo:
            flash('Por favor ingresa el código de verificación', 'error')
            return redirect(url_for('verify_code'))
        
        # Verificar código
        usuario = execute_query(
            """SELECT UsuarioID, ResetToken, ResetTokenExpira
            FROM Usuarios WHERE Email = ?""",
            (email,)
        )
        
        if not usuario or not secure_equals_encrypted(usuario[0]['ResetToken'], codigo):
            flash('Código de verificación incorrecto.', 'error')
            return redirect(url_for('verify_code'))
        
        usuario = usuario[0]
        
        # Verificar si el código ha expirado (10 minutos)
        if usuario['ResetTokenExpira']:
            if datetime.now() > usuario['ResetTokenExpira']:
                flash('El código de verificación ha expirado. Por favor solicita uno nuevo.', 'error')
                return redirect(url_for('forgot_password'))
        
        # Guardar token en sesión para el siguiente paso
        session['reset_token'] = codigo
        return redirect(url_for('reset_password'))
    
    return render_template('verify_code.html', email=email)

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    # Verificar que hay un código válido en sesión
    if 'reset_email' not in session or 'reset_token' not in session:
        flash('Proceso de recuperación no válido. Por favor solicita un nuevo código.', 'error')
        return redirect(url_for('forgot_password'))
    
    email = session['reset_email']
    token = session['reset_token']
    
    # Verificar nuevamente que el código es válido
    usuario = execute_query(
        """SELECT UsuarioID, ResetToken, ResetTokenExpira
        FROM Usuarios WHERE Email = ?""",
        (email,)
    )
    
    if not usuario or not secure_equals_encrypted(usuario[0]['ResetToken'], token):
        flash('El código de verificación ya no es válido. Por favor solicita uno nuevo.', 'error')
        return redirect(url_for('forgot_password'))
    
    usuario = usuario[0]
    
    if request.method == 'POST':
        password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        # Validaciones
        errors = []
        
        if len(password) < 8:
            errors.append('La contraseña debe tener al menos 8 caracteres')
        
        if not any(c.isupper() for c in password):
            errors.append('La contraseña debe contener al menos una letra mayúscula')
        
        if not any(c.islower() for c in password):
            errors.append('La contraseña debe contener al menos una letra minúscula')
        
        if not any(c.isdigit() for c in password):
            errors.append('La contraseña debe contener al menos un número')
        
        if password != confirm_password:
            errors.append('Las contraseñas no coinciden')
        
        if errors:
            for error in errors:
                flash(error, 'error')
            return redirect(url_for('reset_password'))
        
        try:
            # Actualizar contraseña
            password_hash = generate_password_hash(password)
            execute_query(
                """UPDATE Usuarios SET 
                PasswordHash = ?, 
                ResetToken = NULL, 
                ResetTokenExpira = NULL 
                WHERE UsuarioID = ?""",
                (password_hash, usuario['UsuarioID']),
                fetch=False
            )
            
            # Limpiar sesión
            session.pop('reset_email', None)
            session.pop('reset_token', None)
            
            flash('Contraseña actualizada correctamente. Ahora puedes iniciar sesión con tu nueva contraseña.', 'success')
            return redirect(url_for('login'))
            
        except Exception as e:
            current_app.logger.error(f"Error al actualizar contraseña: {str(e)}")
            flash('Ocurrió un error al actualizar la contraseña. Por favor intenta nuevamente.', 'error')
            return redirect(url_for('reset_password'))
    
    return render_template('reset_password.html')



def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Por favor inicia sesión para acceder a esta página.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def role_required(role):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if session.get('tipo') != role:
                flash('No tienes permisos para acceder a esta página.', 'danger')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator


_COLUMN_NAMES = """
UsuarioID Email PasswordHash TipoUsuario FechaRegistro Activo ResetToken
ResetTokenExpira CandidatoID Nombre ApellidoPaterno ApellidoMaterno Telefono
EstadoCivil Sexo FechaNacimiento Nacionalidad RFC Direccion Reubicacion Viajar
LicenciaConducir ModalidadTrabajo PuestoActual PuestoSolicitado FotoPerfil CV
ResumenProfesional EmpresaID SitioWeb Descripcion Logo AdministradorID
PreparacionID Grado Cedula Estatus Institucion Pais FechaInicio FechaFin
ExperienciaID Empresa Domicilio Puesto FechaIngreso FechaSalida Funciones
SueldoInicial SueldoFinal MotivoSeparacion ReferenciaID Ocupacion AnosConocer
Documento HabilidadID CompetenciaID VacanteID GradoEstudios Resumen Plazas
PlazasDisponibles FechaPublicacion FechaAprobacion ComentariosAdmin Salario
TipoContrato Modalidad Ubicacion ExperienciaRequerida Beneficios FechaCierre
PostulacionID FechaPostulacion Comentarios NotificacionID Mensaje Tipo Fecha
Leida ConversacionID Activa MensajeID RemitenteID RemitenteTipo FechaEnvio
Leido FechaLectura EmpresaNombre EmpresaDescripcion CandidatoNombre
CandidatoApellido CandidatoFoto VacantePuesto UltimoMensaje UltimoMensajeFecha
NoLeidos Total Count TotalVacantes TotalUsuarios TotalEmpresas TotalCandidatos
TotalPostulaciones NumPostulaciones EstadoClase HabilidadesRequeridas
HabilidadesCoincidentes PorcentajeCoincidencia Anios Mes Aceptadas Rechazadas
Pendientes TotalVacantes Habilidad
""".split()
_CANONICAL_COLUMN = {name.lower(): name for name in _COLUMN_NAMES}


class CaseInsensitiveDict(dict):
    """Diccionario compatible con los nombres CamelCase usados por las vistas."""
    def __init__(self, values=(), **kwargs):
        normalized = {
            _CANONICAL_COLUMN.get(str(key).lower(), key): value
            for key, value in dict(values, **kwargs).items()
        }
        super().__init__(normalized)

    def _key(self, key):
        if isinstance(key, str):
            for existing in self.keys():
                if isinstance(existing, str) and existing.lower() == key.lower():
                    return existing
        return key

    def __getitem__(self, key):
        return super().__getitem__(self._key(key))

    def get(self, key, default=None):
        return super().get(self._key(key), default)

    def __contains__(self, key):
        return super().__contains__(self._key(key))


def _postgres_query(query):
    """Convierte el pequeño subconjunto T-SQL heredado y sus placeholders."""
    query = re.sub(
        r"FORMAT\(([^,]+),\s*'dd/MM/yyyy'\)",
        r"TO_CHAR(\1, 'DD/MM/YYYY')",
        query,
        flags=re.IGNORECASE,
    )
    query = re.sub(
        r"FORMAT\(([^,]+),\s*'yyyy-MM'\)",
        r"TO_CHAR(\1, 'YYYY-MM')",
        query,
        flags=re.IGNORECASE,
    )
    query = re.sub(
        r"DATEADD\(MINUTE,\s*10,\s*GETDATE\(\)\)",
        "CURRENT_TIMESTAMP + INTERVAL '10 minutes'",
        query,
        flags=re.IGNORECASE,
    )
    query = re.sub(
        r"DATEADD\(MONTH,\s*-12,\s*GETDATE\(\)\)",
        "CURRENT_TIMESTAMP - INTERVAL '12 months'",
        query,
        flags=re.IGNORECASE,
    )
    query = re.sub(r"GETDATE\(\)", "CURRENT_TIMESTAMP", query, flags=re.IGNORECASE)
    query = re.sub(r"ISNULL\(", "COALESCE(", query, flags=re.IGNORECASE)
    for column in ('Activo', 'Leida', 'Leido', 'Activa', 'Reubicacion', 'Viajar', 'LicenciaConducir'):
        query = re.sub(
            rf"\b({column})\s*=\s*([01])\b",
            lambda match: f"{match.group(1)} = {'TRUE' if match.group(2) == '1' else 'FALSE'}",
            query,
            flags=re.IGNORECASE,
        )
    # Las consultas de esta aplicación no contienen signos ? literales.
    return query.replace('?', '%s')


class PostgreSQLCursor:
    """Adaptador pequeño para conservar cursor.execute(sql, params) del código existente."""
    def __init__(self, cursor):
        self._cursor = cursor

    def execute(self, query, params=None):
        self._cursor.execute(_postgres_query(query), params)
        return self

    def __getattr__(self, name):
        return getattr(self._cursor, name)


class PostgreSQLConnection:
    def __init__(self, connection):
        self._connection = connection

    def cursor(self):
        return PostgreSQLCursor(self._connection.cursor())

    def __getattr__(self, name):
        return getattr(self._connection, name)


def get_db_connection():
    database_url = os.getenv('DATABASE_URL')
    if database_url:
        connection = psycopg2.connect(database_url)
    else:
        connection = psycopg2.connect(
            host=app.config['DB_HOST'],
            port=app.config['DB_PORT'],
            dbname=app.config['DB_NAME'],
            user=app.config['DB_USER'],
            password=app.config['DB_PASSWORD'],
        )
    return PostgreSQLConnection(connection)

def execute_query(query, params=None, fetch=True):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        if params:
            cursor.execute(query, params)
        else:
            cursor.execute(query)
        
        if fetch:
            if cursor.description:
                columns = [column[0] for column in cursor.description]
                results = [CaseInsensitiveDict(zip(columns, row)) for row in cursor.fetchall()]
                if query.strip().upper().startswith(('INSERT', 'UPDATE', 'DELETE')):
                    conn.commit()
                return results
            elif query.strip().upper().startswith(('INSERT', 'UPDATE', 'DELETE')):
                conn.commit()
                return cursor.rowcount
        else:
            conn.commit()
            return None
            
    except Exception as e:
        conn.rollback()
        current_app.logger.error(f"Database error: {str(e)}")
        raise e
    finally:
        cursor.close()
        conn.close()


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def calcular_edad(fecha_nacimiento):
    if not fecha_nacimiento:
        return 0
    hoy = date.today()
    return hoy.year - fecha_nacimiento.year - ((hoy.month, hoy.day) < (fecha_nacimiento.month, fecha_nacimiento.day))

def get_usuario_actual():
    if 'user_id' in session:
        result = execute_query(
            "SELECT * FROM Usuarios WHERE UsuarioID = ?", 
            (session['user_id'],)
        )
        return result[0] if result else None
    return None

def get_candidato_actual():
    usuario = get_usuario_actual()
    if usuario and usuario['TipoUsuario'] == 'candidato':
        result = execute_query(
            "SELECT * FROM Candidatos WHERE UsuarioID = ?", 
            (usuario['UsuarioID'],)
        )
        return result[0] if result else None
    return None

def get_empresa_actual():
    usuario = get_usuario_actual()
    if usuario and usuario['TipoUsuario'] == 'empresa':
        result = execute_query(
            "SELECT * FROM Empresas WHERE UsuarioID = ?", 
            (usuario['UsuarioID'],)
        )
        return result[0] if result else None
    return None

def get_admin_actual():
    usuario = get_usuario_actual()
    if usuario and usuario['TipoUsuario'] == 'admin':
        result = execute_query(
            "SELECT * FROM Administradores WHERE UsuarioID = ?", 
            (usuario['UsuarioID'],)
        )
        return result[0] if result else None
    return None


@app.route('/')
def index():

    vacantes = execute_query(
    "SELECT v.*, e.Nombre as EmpresaNombre FROM Vacantes v "
    "JOIN Empresas e ON v.EmpresaID = e.EmpresaID "
    "WHERE v.Estatus = 'aprobada' ORDER BY v.FechaPublicacion DESC LIMIT 3"
)


    return render_template('index.html', usuario=get_usuario_actual(), vacantes=vacantes)

##################################################################################################

@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        email = request.form['email'].strip()
        password = request.form['password'].strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        tipo = request.form['tipo']
        nombre = request.form.get('nombre', '').strip()
        apellido_paterno = request.form.get('apellido_paterno', '').strip()
        apellido_materno = request.form.get('apellido_materno', '').strip()
        telefono = request.form.get('telefono', '').strip()
        direccion = request.form.get('direccion', '').strip()
        

        errors = []

        if not email or '@' not in email or '.' not in email.split('@')[-1]:
            errors.append('Ingrese un correo electrónico válido')

        if len(password) < 8:
            errors.append('La contraseña debe tener al menos 8 caracteres')         
        if password != confirm_password:
            errors.append('Las contraseñas no coinciden')

        if tipo not in ['candidato', 'empresa', 'admin']:
            errors.append('Tipo de usuario no válido')

        if tipo == 'candidato':
            if not email.lower().endswith('@upq.edu.mx'):
                errors.append('Los candidatos deben registrarse con su correo institucional @upq.edu.mx')
            if not nombre:
                errors.append('El nombre es obligatorio')
            if not apellido_paterno:
                errors.append('El apellido paterno es obligatorio')
            if telefono and (not telefono.isdigit() or len(telefono) != 10):
                errors.append('El teléfono debe tener 10 dígitos')
                
        if tipo == 'empresa':
            if not nombre:
                errors.append('El nombre de la empresa es obligatorio')
            if telefono and (not telefono.isdigit() or len(telefono) != 10):
                errors.append('El teléfono debe tener 10 dígitos')

        if errors:
            for error in errors:
                flash(error, 'error')
            return redirect(url_for('registro'))
    

        existe = execute_query(
            "SELECT 1 FROM Usuarios WHERE Email = ?", 
            (email,)
        )
        
        if existe:
            flash('El correo electrónico ya está registrado.', 'error')
            return redirect(url_for('registro'))
        
        try:
  
            execute_query(
                "INSERT INTO Usuarios (Email, PasswordHash, TipoUsuario) VALUES (?, ?, ?)",
                (email, generate_password_hash(password), tipo),
                fetch=False
            )
            
        
            nuevo_usuario = execute_query(
                "SELECT UsuarioID FROM Usuarios WHERE Email = ?",
                (email,)
            )
            usuario_id = nuevo_usuario[0]['UsuarioID']
            
            
            if tipo == 'candidato':
                execute_query(
                    """INSERT INTO Candidatos 
                    (CandidatoID, UsuarioID, Nombre, ApellidoPaterno, ApellidoMaterno, Telefono, Direccion)
                    VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (usuario_id, usuario_id, nombre, apellido_paterno, apellido_materno, telefono, direccion),
                    fetch=False
                )
                
              
                nombre_correo = f"{nombre} {apellido_paterno} {apellido_materno}".strip()
                
            elif tipo == 'empresa':
                execute_query(
                    """INSERT INTO Empresas 
                    (EmpresaID, UsuarioID, Nombre, Telefono, Direccion)
                    VALUES (?, ?, ?, ?, ?)""",
                    (usuario_id, usuario_id, nombre, telefono, direccion),
                    fetch=False
                )
                
                
                nombre_correo = nombre
                
            elif tipo == 'admin':
                execute_query(
                    "INSERT INTO Administradores (AdministradorID, UsuarioID) VALUES (?, ?)",
                    (usuario_id, usuario_id),
                    fetch=False
                )
                
           
                nombre_correo = email.split('@')[0]
            

            try:
                if tipo in ['candidato', 'empresa']:  
                    enviar_correo_bienvenida(email, nombre_correo, tipo)
                    print(f"📧 Correo de bienvenida enviado a: {email}")
                else:
                    print(f"ℹ️  No se envía correo de bienvenida para tipo de usuario: {tipo}")
            except Exception as e:
    
                print(f"⚠️  Error enviando correo de bienvenida: {str(e)}")
         
            
            flash('Registro exitoso. Por favor inicia sesión.', 'success')
            return redirect(url_for('login'))
        
        except Exception as e:
            current_app.logger.error(f"Error en registro: {str(e)}")
            flash('Ocurrió un error durante el registro. Por favor intente nuevamente.', 'error')
            return redirect(url_for('registro'))
    
    return render_template('registro.html')



@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        if request.form.get('is_admin') == 'true':
            email = request.form.get('email', '').strip().lower()
            password = request.form.get('password', '')
            admins = execute_query(
                """SELECT UsuarioID, Email, PasswordHash, TipoUsuario, Activo
                   FROM Usuarios WHERE LOWER(Email) = ? AND TipoUsuario = 'admin'""",
                (email,)
            )
            if admins and admins[0]['Activo'] and check_password_hash(admins[0]['PasswordHash'], password):
                user_obj = User(
                    id=admins[0]['UsuarioID'],
                    email=admins[0]['Email'],
                    tipo='admin',
                    activo=admins[0]['Activo'],
                )
                login_user(user_obj, remember=True, duration=timedelta(days=30))
                session['email'] = admins[0]['Email']
                session['tipo'] = 'admin'
                session['user_id'] = admins[0]['UsuarioID']
                session.permanent = True
                return redirect(url_for('admin_dashboard'))
            else:
                flash('Credenciales administrativas incorrectas', 'error')
                return redirect(url_for('login'))
        
        email = request.form['email']
        password = request.form['password']
        
        usuario = execute_query(
            "SELECT * FROM Usuarios WHERE Email = ?", 
            (email,)
        )
        
        if usuario and check_password_hash(usuario[0]['PasswordHash'], password):
            # ===== USAR FLASK-LOGIN =====
            from flask_login import login_user
            
            # Crear objeto User para Flask-Login
            user_obj = User(
                id=usuario[0]['UsuarioID'],
                email=usuario[0]['Email'],
                tipo=usuario[0]['TipoUsuario'],
                activo=usuario[0].get('Activo', 1)
            )
            
            # Autenticar con Flask-Login
            remember = request.form.get('remember') == 'on'
            login_user(user_obj, remember=remember, duration=timedelta(days=30))
            
            # También mantener la sesión para compatibilidad con código existente
            session['email'] = email
            session['tipo'] = usuario[0]['TipoUsuario']
            session['user_id'] = usuario[0]['UsuarioID']
            session.permanent = remember
            
            flash('Inicio de sesión exitoso.', 'success')
            
            if usuario[0]['TipoUsuario'] == 'admin':
                return redirect(url_for('admin_dashboard'))
            elif usuario[0]['TipoUsuario'] == 'empresa':
                return redirect(url_for('empresa_dashboard'))
            else:
                return redirect(url_for('candidato_dashboard'))
        else:
            flash('Correo electrónico o contraseña incorrectos.', 'error')
    
    return render_template('login.html')


@app.route('/logout')
def logout():
    from flask_login import logout_user
    logout_user()  # Cerrar sesión de Flask-Login
    session.clear()
    flash('Has cerrado sesión correctamente.', 'info')
    return redirect(url_for('index'))

@app.route('/candidato')
@login_required
@role_required('candidato')
def candidato_dashboard():
    candidato = get_candidato_actual()
    if not candidato:
        flash('Perfil de candidato no encontrado', 'error')
        return redirect(url_for('login'))
    

    candidato_id = candidato['CandidatoID']
    

    edad = calcular_edad(candidato.get('FechaNacimiento'))
    

    preparacion_academica = execute_query(
        "SELECT * FROM PreparacionAcademica WHERE CandidatoID = ?",
        (candidato_id,)
    )
    

    experiencia_laboral = execute_query(
        "SELECT * FROM ExperienciaLaboral WHERE CandidatoID = ? ORDER BY FechaIngreso DESC LIMIT 1",
        (candidato_id,)
    )
    

    habilidades = execute_query(
        "SELECT h.Nombre FROM CandidatoHabilidades ch "
        "JOIN Habilidades h ON ch.HabilidadID = h.HabilidadID "
        "WHERE ch.CandidatoID = ?",
        (candidato_id,)
    )
    

    referencias = execute_query(
        "SELECT * FROM Referencias WHERE CandidatoID = ?",
        (candidato_id,)
    )
    

    postulaciones = execute_query(
        "SELECT p.*, v.Puesto, e.Nombre as EmpresaNombre "
        "FROM Postulaciones p "
        "JOIN Vacantes v ON p.VacanteID = v.VacanteID "
        "JOIN Empresas e ON v.EmpresaID = e.EmpresaID "
        "WHERE p.CandidatoID = ? "
        "ORDER BY p.FechaPostulacion DESC LIMIT 3",
        (candidato_id,)
    )
    
   
    completed = 0
    if candidato.get('Nombre'): completed += 20
    if preparacion_academica: completed += 20
    if experiencia_laboral: completed += 20
    if habilidades: completed += 20
    if referencias: completed += 20
    
    return render_template('candidato/dashboard.html', 
                         candidato=candidato,
                         edad=edad,
                         completed=completed,
                         preparacion_academica=preparacion_academica,
                         experiencia_laboral=experiencia_laboral,
                         habilidades=habilidades,
                         referencias=referencias,
                         postulaciones=postulaciones)

@app.route('/candidato/perfil', methods=['GET', 'POST'])
@login_required
@role_required('candidato')
def candidato_perfil():
    candidato = get_candidato_actual()
    if not candidato:
        flash('Perfil de candidato no encontrado', 'error')
        return redirect(url_for('login'))

    if request.method == 'POST':
        try:
   
            errors = []
            

            if not request.form['nombre'].strip() or not request.form['apellido_paterno'].strip():
                errors.append('Nombre y apellido paterno son obligatorios')
                
  
            telefono = request.form['telefono'].strip()
            if not telefono.isdigit() or len(telefono) != 10:
                errors.append('El teléfono debe tener 10 dígitos')
                

            rfc = request.form['rfc'].strip().upper()
            if len(rfc) < 12 or len(rfc) > 13:
                errors.append('El RFC debe tener entre 12 y 13 caracteres')
                

            fecha_nacimiento_str = request.form['fecha_nacimiento']
            if fecha_nacimiento_str:
                try:
                    fecha_nacimiento = datetime.strptime(fecha_nacimiento_str, '%Y-%m-%d').date()
                    hoy = date.today()
                    edad = hoy.year - fecha_nacimiento.year - ((hoy.month, hoy.day) < (fecha_nacimiento.month, fecha_nacimiento.day))
                    if edad < 18:
                        errors.append('Debes ser mayor de edad (18 años)')
                except ValueError:
                    errors.append('Formato de fecha inválido (AAAA-MM-DD)')
            else:
                fecha_nacimiento = None
                
 
            if 'foto_perfil' in request.files:
                file = request.files['foto_perfil']
                if file and file.filename:
                    if not allowed_file(file.filename):
                        errors.append('Formato de imagen no permitido')
                    if file.content_length > 2 * 1024 * 1024:  # 2MB
                        errors.append('La imagen no debe exceder 2MB')
                        
            if 'cv' in request.files:
                file = request.files['cv']
                if file and file.filename:
                    if not file.filename.lower().endswith('.pdf'):
                        errors.append('El CV debe ser un archivo PDF')
                    if file.content_length > 5 * 1024 * 1024:  # 5MB
                        errors.append('El CV no debe exceder 5MB')
            
            if errors:
                for error in errors:
                    flash(error, 'error')
                return redirect(url_for('candidato_perfil'))


            update_data = {
                'Nombre': request.form['nombre'],
                'ApellidoPaterno': request.form['apellido_paterno'],
                'ApellidoMaterno': request.form['apellido_materno'],
                'Telefono': telefono,
                'EstadoCivil': request.form['estado_civil'],
                'Sexo': request.form['sexo'],
                'FechaNacimiento': fecha_nacimiento,
                'Nacionalidad': request.form['nacionalidad'],
                'RFC': rfc,
                'Direccion': request.form['direccion'],
                'Reubicacion': 'reubicacion' in request.form,
                'Viajar': 'viajar' in request.form,
                'LicenciaConducir': 'licencia_conducir' in request.form,
                'ModalidadTrabajo': request.form['modalidad'],
                'PuestoActual': request.form['puesto_actual'],
                'PuestoSolicitado': request.form['puesto_solicitado'],
                'ResumenProfesional': request.form['resumen']
            }


            if 'foto_perfil' in request.files:
                file = request.files['foto_perfil']
                if file and allowed_file(file.filename):
                    photo_bytes = file.read(2 * 1024 * 1024 + 1)
                    if len(photo_bytes) > 2 * 1024 * 1024:
                        flash('La imagen no debe exceder 2MB', 'error')
                        return redirect(url_for('candidato_perfil'))
                    mime = file.mimetype if file.mimetype in ('image/jpeg', 'image/png', 'image/webp') else 'image/jpeg'
                    update_data['FotoPerfil'] = f"data:{mime};base64,{base64.b64encode(photo_bytes).decode()}"

            if 'cv' in request.files:
                file = request.files['cv']
                if file and file.filename.lower().endswith('.pdf'):
                    filename = secure_filename(f"cv_{candidato['CandidatoID']}.pdf")
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    update_data['CV'] = filename

            execute_query(
                """UPDATE Candidatos SET
                Nombre = ?, ApellidoPaterno = ?, ApellidoMaterno = ?,
                Telefono = ?, EstadoCivil = ?, Sexo = ?,
                FechaNacimiento = ?, Nacionalidad = ?, RFC = ?,
                Direccion = ?, Reubicacion = ?, Viajar = ?,
                LicenciaConducir = ?, ModalidadTrabajo = ?,
                PuestoActual = ?, PuestoSolicitado = ?,
                ResumenProfesional = ?,
                FotoPerfil = COALESCE(?, FotoPerfil),
                CV = COALESCE(?, CV)
                WHERE CandidatoID = ?""",
                (
                    update_data['Nombre'],
                    update_data['ApellidoPaterno'],
                    update_data['ApellidoMaterno'],
                    update_data['Telefono'],
                    update_data['EstadoCivil'],
                    update_data['Sexo'],
                    update_data['FechaNacimiento'],
                    update_data['Nacionalidad'],
                    update_data['RFC'],
                    update_data['Direccion'],
                    update_data['Reubicacion'],
                    update_data['Viajar'],
                    update_data['LicenciaConducir'],
                    update_data['ModalidadTrabajo'],
                    update_data['PuestoActual'],
                    update_data['PuestoSolicitado'],
                    update_data['ResumenProfesional'],
                    update_data.get('FotoPerfil'),
                    update_data.get('CV'),
                    candidato['CandidatoID']
                ),
                fetch=False
            )

            flash('Perfil actualizado correctamente.', 'success')
            return redirect(url_for('candidato_perfil'))

        except Exception as e:
            current_app.logger.error(f"Error al actualizar perfil: {str(e)}")
            flash('Ocurrió un error al actualizar el perfil', 'error')


    edad = calcular_edad(candidato.get('FechaNacimiento'))
    
    return render_template('candidato/perfil.html', 
                         candidato=candidato, 
                         edad=edad,
                         estados_civiles=['Soltero/a', 'Casado/a', 'Divorciado/a', 'Viudo/a', 'Unión libre'],
                         generos=['Masculino', 'Femenino', 'Otro'],
                         modalidades=['Presencial', 'Remoto', 'Híbrido'])


@app.route('/candidato/preparacion', methods=['GET', 'POST'])
@login_required
@role_required('candidato')
def candidato_preparacion():
    candidato = get_candidato_actual()
    if not candidato:
        flash('Perfil de candidato no encontrado', 'error')
        return redirect(url_for('login'))
    
    editar_index = request.args.get('editar', type=int)
    preparacion_editar = None
    
    if editar_index is not None:
        preparaciones = execute_query(
            "SELECT * FROM PreparacionAcademica WHERE CandidatoID = ?",
            (candidato['CandidatoID'],)
        )
        if 0 <= editar_index < len(preparaciones):
            preparacion_editar = preparaciones[editar_index]
    
    if request.method == 'POST':
        if 'eliminar' in request.form:
            try:
                index = int(request.form['eliminar'])
                preparaciones = execute_query(
                    "SELECT PreparacionID FROM PreparacionAcademica WHERE CandidatoID = ?",
                    (candidato['CandidatoID'],)
                )
                if 0 <= index < len(preparaciones):
                    execute_query(
                        "DELETE FROM PreparacionAcademica WHERE PreparacionID = ?",
                        (preparaciones[index]['PreparacionID'],),
                        fetch=False
                    )
                    flash('Preparación académica eliminada correctamente', 'success')
            except (ValueError, KeyError):
                flash('Error al eliminar la preparación académica', 'danger')
        else:
            required_fields = ['grado', 'estatus', 'institucion', 'pais', 'fecha_inicio']
            if not all(field in request.form and request.form[field].strip() for field in required_fields):
                flash('Complete todos los campos requeridos', 'danger')
                return redirect(url_for('candidato_preparacion'))
            
            try:
                fecha_inicio = datetime.strptime(request.form['fecha_inicio'], '%Y-%m-%d').date()
                fecha_fin = datetime.strptime(request.form['fecha_fin'], '%Y-%m-%d').date() if request.form.get('fecha_fin') else None
                
        
                if request.form['estatus'] == 'Completado' and fecha_fin:
                    duracion = (fecha_fin.year - fecha_inicio.year) - ((fecha_fin.month, fecha_fin.day) < (fecha_inicio.month, fecha_inicio.day))
                    if duracion < 3:
                        flash('Los estudios completados deben tener al menos 3 años de duración', 'danger')
                        return redirect(url_for('candidato_preparacion'))
                
 
                if fecha_fin and fecha_fin < fecha_inicio:
                    flash('La fecha de finalización no puede ser anterior a la fecha de inicio', 'danger')
                    return redirect(url_for('candidato_preparacion'))
                

                if fecha_inicio > date.today():
                    flash('La fecha de inicio no puede ser en el futuro', 'danger')
                    return redirect(url_for('candidato_preparacion'))
                

                if request.form['estatus'] == 'En progreso' and fecha_fin:
                    flash('Los estudios en progreso no deben tener fecha de finalización', 'danger')
                    return redirect(url_for('candidato_preparacion'))
                

                grados_validos = ['Primaria', 'Secundaria', 'Bachillerato', 'Licenciatura', 
                                 'Maestría', 'Doctorado', 'Técnico', 'Diplomado', 'Certificación']
                if request.form['grado'] not in grados_validos:
                    flash('Seleccione un grado académico válido', 'danger')
                    return redirect(url_for('candidato_preparacion'))
                
                nueva_preparacion = {
                    'Grado': request.form['grado'],
                    'Cedula': request.form.get('cedula', ''),
                    'Estatus': request.form['estatus'],
                    'Institucion': request.form['institucion'],
                    'Pais': request.form['pais'],
                    'FechaInicio': fecha_inicio,
                    'FechaFin': fecha_fin
                }

                if len(nueva_preparacion['Institucion'].strip()) < 5:
                    flash('El nombre de la institución debe ser válido', 'danger')
                    return redirect(url_for('candidato_preparacion'))
                
                if 'editar_index' in request.form and request.form['editar_index'].strip():
                    try:
                        index = int(request.form['editar_index'])
                        preparaciones = execute_query(
                            "SELECT PreparacionID FROM PreparacionAcademica WHERE CandidatoID = ?",
                            (candidato['CandidatoID'],)
                        )
                        if 0 <= index < len(preparaciones):
                            execute_query(
                                """UPDATE PreparacionAcademica SET
                                Grado = ?, Cedula = ?, Estatus = ?,
                                Institucion = ?, Pais = ?,
                                FechaInicio = ?, FechaFin = ?
                                WHERE PreparacionID = ?""",
                                (
                                    nueva_preparacion['Grado'],
                                    nueva_preparacion['Cedula'],
                                    nueva_preparacion['Estatus'],
                                    nueva_preparacion['Institucion'],
                                    nueva_preparacion['Pais'],
                                    nueva_preparacion['FechaInicio'],
                                    nueva_preparacion['FechaFin'],
                                    preparaciones[index]['PreparacionID']
                                ),
                                fetch=False
                            )
                            flash('Preparación académica actualizada correctamente', 'success')
                    except ValueError:
                        flash('Índice de edición inválido', 'danger')
                else:
                    execute_query(
                        """INSERT INTO PreparacionAcademica
                        (CandidatoID, Grado, Cedula, Estatus, Institucion, Pais, FechaInicio, FechaFin)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                        (
                            candidato['CandidatoID'],
                            nueva_preparacion['Grado'],
                            nueva_preparacion['Cedula'],
                            nueva_preparacion['Estatus'],
                            nueva_preparacion['Institucion'],
                            nueva_preparacion['Pais'],
                            nueva_preparacion['FechaInicio'],
                            nueva_preparacion['FechaFin']
                        ),
                        fetch=False
                    )
                    flash('Preparación académica agregada correctamente', 'success')
            except ValueError as e:
                flash(f'Error en el formato de fecha: {str(e)}', 'danger')
        
        return redirect(url_for('candidato_preparacion'))
    
    preparaciones = execute_query(
        "SELECT * FROM PreparacionAcademica WHERE CandidatoID = ? ORDER BY FechaInicio DESC",
        (candidato['CandidatoID'],)
    )
    
    return render_template('candidato/preparacion.html', 
                         candidato=candidato,
                         preparaciones=preparaciones,
                         editar_index=editar_index,
                         preparacion_editar=preparacion_editar)


@app.route('/candidato/experiencia', methods=['GET', 'POST'])
@login_required
@role_required('candidato')
def candidato_experiencia():
    candidato = get_candidato_actual()
    if not candidato:
        flash('Perfil de candidato no encontrado', 'error')
        return redirect(url_for('login'))
    

    editar_index = request.args.get('editar', type=int)
    experiencia_editar = None
    if editar_index is not None:
        experiencias = execute_query(
            "SELECT * FROM ExperienciaLaboral WHERE CandidatoID = ?",
            (candidato['CandidatoID'],)
        )
        if 0 <= editar_index < len(experiencias):
            experiencia_editar = experiencias[editar_index]
    
    if request.method == 'POST':
        try:
            if 'eliminar' in request.form:
            
                try:
                    index = int(request.form['eliminar'])
                    experiencias = execute_query(
                        "SELECT ExperienciaID FROM ExperienciaLaboral WHERE CandidatoID = ?",
                        (candidato['CandidatoID'],)
                    )
                    if 0 <= index < len(experiencias):
                        execute_query(
                            "DELETE FROM ExperienciaLaboral WHERE ExperienciaID = ?",
                            (experiencias[index]['ExperienciaID'],),
                            fetch=False
                        )
                        flash('Experiencia laboral eliminada correctamente', 'success')
                except (ValueError, KeyError):
                    flash('Error al eliminar la experiencia laboral', 'danger')
            else:

                required_fields = ['empresa', 'puesto', 'fecha_ingreso', 'funciones']
                if not all(field in request.form and request.form[field].strip() for field in required_fields):
                    flash('Por favor complete todos los campos requeridos', 'danger')
                    return redirect(url_for('candidato_experiencia'))
                
                try:
 
                    fecha_ingreso = datetime.strptime(request.form['fecha_ingreso'], '%Y-%m-%d').date()
                    fecha_salida = datetime.strptime(request.form['fecha_salida'], '%Y-%m-%d').date() if request.form.get('fecha_salida') else None
                    
                   
                    if fecha_salida and fecha_salida < fecha_ingreso:
                        flash('La fecha de salida no puede ser anterior a la fecha de ingreso', 'danger')
                        return redirect(url_for('candidato_experiencia'))

                    if fecha_salida and (fecha_salida - fecha_ingreso).days < 90:
                        flash('La experiencia laboral debe ser de al menos 3 meses', 'danger')
                        return redirect(url_for('candidato_experiencia'))

                    if fecha_ingreso > date.today():
                        flash('La fecha de ingreso no puede ser en el futuro', 'danger')
                        return redirect(url_for('candidato_experiencia'))

                    SALARIO_MINIMO_DIARIO = 207.44
                    SALARIO_MINIMO_MENSUAL = SALARIO_MINIMO_DIARIO * 30
                    
                    sueldo_inicial = float(request.form['sueldo_inicial']) if request.form.get('sueldo_inicial') else 0
                    sueldo_final = float(request.form['sueldo_final']) if request.form.get('sueldo_final') else 0
                    
                    if sueldo_inicial > 0 and sueldo_inicial < SALARIO_MINIMO_MENSUAL:
                        flash(f'El sueldo inicial no puede ser menor al salario mínimo mensual (${SALARIO_MINIMO_MENSUAL:,.2f} MXN)', 'danger')
                        return redirect(url_for('candidato_experiencia'))
                    
                    if sueldo_final > 0 and sueldo_final < sueldo_inicial:
                        flash('El sueldo final no puede ser menor al sueldo inicial', 'danger')
                        return redirect(url_for('candidato_experiencia'))
                    
                  
                    empresa = request.form['empresa'].strip()
                    if len(empresa) < 3:
                        flash('El nombre de la empresa debe tener al menos 3 caracteres', 'danger')
                        return redirect(url_for('candidato_experiencia'))
                    
        
                    puesto = request.form['puesto'].strip()
                    if len(puesto) < 3:
                        flash('El puesto debe tener al menos 3 caracteres', 'danger')
                        return redirect(url_for('candidato_experiencia'))
                    
      
                    funciones = request.form['funciones'].strip()
                    if len(funciones) < 10:
                        flash('Las funciones deben tener al menos 10 caracteres', 'danger')
                        return redirect(url_for('candidato_experiencia'))
                    

                    nueva_experiencia = {
                        'Empresa': empresa,
                        'Domicilio': request.form.get('domicilio', '').strip(),
                        'Telefono': request.form.get('telefono', '').strip(),
                        'Puesto': puesto,
                        'FechaIngreso': fecha_ingreso,
                        'FechaSalida': fecha_salida,
                        'Funciones': funciones,
                        'SueldoInicial': sueldo_inicial,
                        'SueldoFinal': sueldo_final,
                        'MotivoSeparacion': request.form.get('motivo_separacion', '').strip()
                    }
                    
                    if 'editar_index' in request.form and request.form['editar_index'].strip():
                        try:
                            index = int(request.form['editar_index'])
                            experiencias = execute_query(
                                "SELECT ExperienciaID FROM ExperienciaLaboral WHERE CandidatoID = ?",
                                (candidato['CandidatoID'],)
                            )
                            if 0 <= index < len(experiencias):
                                execute_query(
                                    """UPDATE ExperienciaLaboral SET
                                    Empresa = ?, Domicilio = ?, Telefono = ?,
                                    Puesto = ?, FechaIngreso = ?, FechaSalida = ?,
                                    Funciones = ?, SueldoInicial = ?, SueldoFinal = ?,
                                    MotivoSeparacion = ?
                                    WHERE ExperienciaID = ?""",
                                    (
                                        nueva_experiencia['Empresa'],
                                        nueva_experiencia['Domicilio'],
                                        nueva_experiencia['Telefono'],
                                        nueva_experiencia['Puesto'],
                                        nueva_experiencia['FechaIngreso'],
                                        nueva_experiencia['FechaSalida'],
                                        nueva_experiencia['Funciones'],
                                        nueva_experiencia['SueldoInicial'],
                                        nueva_experiencia['SueldoFinal'],
                                        nueva_experiencia['MotivoSeparacion'],
                                        experiencias[index]['ExperienciaID']
                                    ),
                                    fetch=False
                                )
                                flash('Experiencia laboral actualizada correctamente', 'success')
                        except ValueError:
                            flash('Índice de edición inválido', 'danger')
                    else:
                        execute_query(
                            """INSERT INTO ExperienciaLaboral
                            (CandidatoID, Empresa, Domicilio, Telefono, Puesto, 
                            FechaIngreso, FechaSalida, Funciones, SueldoInicial, SueldoFinal, MotivoSeparacion)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                            (
                                candidato['CandidatoID'],
                                nueva_experiencia['Empresa'],
                                nueva_experiencia['Domicilio'],
                                nueva_experiencia['Telefono'],
                                nueva_experiencia['Puesto'],
                                nueva_experiencia['FechaIngreso'],
                                nueva_experiencia['FechaSalida'],
                                nueva_experiencia['Funciones'],
                                nueva_experiencia['SueldoInicial'],
                                nueva_experiencia['SueldoFinal'],
                                nueva_experiencia['MotivoSeparacion']
                            ),
                            fetch=False
                        )
                        flash('Experiencia laboral agregada correctamente', 'success')
                except ValueError as e:
                    flash(f'Error en el formato de datos: {str(e)}', 'danger')
            
            return redirect(url_for('candidato_experiencia'))
        
        except Exception as e:
            current_app.logger.error(f"Error en experiencia laboral: {str(e)}")
            flash('Ocurrió un error al procesar la solicitud', 'danger')
    

    experiencias = execute_query(
        "SELECT * FROM ExperienciaLaboral WHERE CandidatoID = ? ORDER BY FechaIngreso DESC",
        (candidato['CandidatoID'],)
    )
    
    return render_template('candidato/experiencia.html',
                         candidato=candidato,
                         experiencias=experiencias,
                         editar_index=editar_index,
                         experiencia_editar=experiencia_editar)


@app.route('/candidato/referencias', methods=['GET', 'POST'])
@login_required
@role_required('candidato')
def candidato_referencias():
    candidato = get_candidato_actual()
    if not candidato:
        flash('Perfil de candidato no encontrado', 'error')
        return redirect(url_for('login'))
    

    editar_index = request.args.get('editar', type=int)
    referencia_editar = None
    if editar_index is not None:
        referencias = execute_query(
            "SELECT * FROM Referencias WHERE CandidatoID = ?",
            (candidato['CandidatoID'],)
        )
        if 0 <= editar_index < len(referencias):
            referencia_editar = referencias[editar_index]
    
    if request.method == 'POST':
        try:
            if 'eliminar' in request.form:

                try:
                    index = int(request.form['eliminar'])
                    referencias = execute_query(
                        "SELECT ReferenciaID FROM Referencias WHERE CandidatoID = ?",
                        (candidato['CandidatoID'],)
                    )
                    if 0 <= index < len(referencias):
                        execute_query(
                            "DELETE FROM Referencias WHERE ReferenciaID = ?",
                            (referencias[index]['ReferenciaID'],),
                            fetch=False
                        )
                        flash('Referencia eliminada correctamente', 'success')
                except (ValueError, KeyError):
                    flash('Error al eliminar la referencia', 'danger')
            
            else:
      
                required_fields = ['nombre', 'ocupacion', 'telefono', 'anos_conocer']
                if not all(field in request.form and request.form[field].strip() for field in required_fields):
                    flash('Complete todos los campos requeridos', 'danger')
                    return redirect(url_for('candidato_referencias'))
                
         
                nombre = request.form['nombre'].strip()
                if len(nombre) < 5 or len(nombre.split()) < 2:
                    flash('Ingrese un nombre completo válido (nombre y apellido)', 'danger')
                    return redirect(url_for('candidato_referencias'))
                
           
                ocupacion = request.form['ocupacion'].strip()
                if len(ocupacion) < 3:
                    flash('La ocupación debe tener al menos 3 caracteres', 'danger')
                    return redirect(url_for('candidato_referencias'))
                

                telefono = request.form['telefono'].strip()
                if not telefono.isdigit() or len(telefono) != 10:
                    flash('Ingrese un número de teléfono válido (10 dígitos)', 'danger')
                    return redirect(url_for('candidato_referencias'))
                
   
                try:
                    anos_conocer = int(request.form['anos_conocer'])
                    if anos_conocer < 1:
                        flash('Debe conocer a la persona al menos 1 año', 'danger')
                        return redirect(url_for('candidato_referencias'))
                    if anos_conocer > 100:
                        flash('Ingrese un valor realista de años', 'danger')
                        return redirect(url_for('candidato_referencias'))
                except ValueError:
                    flash('Años de conocer debe ser un número válido', 'danger')
                    return redirect(url_for('candidato_referencias'))
                

                empresa = request.form.get('empresa', '').strip()
                if empresa and len(empresa) < 3:
                    flash('El nombre de la empresa debe tener al menos 3 caracteres', 'danger')
                    return redirect(url_for('candidato_referencias'))
                
        
                documento = request.form.get('documento_actual', '')
                if 'documento' in request.files:
                    file = request.files['documento']
                    if file and file.filename != '':
                        if not allowed_file(file.filename):
                            flash('Solo se permiten documentos PDF (máx 5MB)', 'danger')
                            return redirect(url_for('candidato_referencias'))
                        
                   
                        file.seek(0, os.SEEK_END)
                        file_size = file.tell()
                        file.seek(0)
                        if file_size > 5 * 1024 * 1024:  
                            flash('El documento no debe exceder 5MB', 'danger')
                            return redirect(url_for('candidato_referencias'))
                        
                        filename = secure_filename(f"ref_{candidato['CandidatoID']}_{int(time.time())}.pdf")
                        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                        documento = filename
                

                nueva_referencia = {
                    'nombre': nombre,
                    'ocupacion': ocupacion,
                    'telefono': telefono,
                    'anos_conocer': anos_conocer,
                    'empresa': empresa,
                    'documento': documento
                }
                

                if 'editar_index' in request.form and request.form['editar_index'].strip():
                    try:
                        index = int(request.form['editar_index'])
                        referencias = execute_query(
                            "SELECT ReferenciaID FROM Referencias WHERE CandidatoID = ?",
                            (candidato['CandidatoID'],)
                        )
                        if 0 <= index < len(referencias):
                            execute_query(
                                """UPDATE Referencias SET
                                Nombre = ?, Ocupacion = ?, Telefono = ?,
                                AnosConocer = ?, Empresa = ?, Documento = COALESCE(?, Documento)
                                WHERE ReferenciaID = ?""",
                                (
                                    nueva_referencia['nombre'],
                                    nueva_referencia['ocupacion'],
                                    nueva_referencia['telefono'],
                                    nueva_referencia['anos_conocer'],
                                    nueva_referencia['empresa'],
                                    nueva_referencia.get('documento'),
                                    referencias[index]['ReferenciaID']
                                ),
                                fetch=False
                            )
                            flash('Referencia actualizada correctamente', 'success')
                    except ValueError:
                        flash('Error al editar la referencia', 'danger')
                else:
                    execute_query(
                        """INSERT INTO Referencias
                        (CandidatoID, Nombre, Ocupacion, Telefono, AnosConocer, Empresa, Documento)
                        VALUES (?, ?, ?, ?, ?, ?, ?)""",
                        (
                            candidato['CandidatoID'],
                            nueva_referencia['nombre'],
                            nueva_referencia['ocupacion'],
                            nueva_referencia['telefono'],
                            nueva_referencia['anos_conocer'],
                            nueva_referencia['empresa'],
                            nueva_referencia.get('documento', '')
                        ),
                        fetch=False
                    )
                    flash('Referencia agregada correctamente', 'success')
            
            return redirect(url_for('candidato_referencias'))
        
        except Exception as e:
            current_app.logger.error(f"Error en referencias: {str(e)}")
            flash('Ocurrió un error al procesar la solicitud', 'danger')
    

    referencias = execute_query(
        "SELECT * FROM Referencias WHERE CandidatoID = ?",
        (candidato['CandidatoID'],)
    )
    
    return render_template('candidato/referencias.html', 
                         candidato=candidato,
                         referencias=referencias,
                         editar_index=editar_index,
                         referencia_editar=referencia_editar)


@app.route('/candidato/habilidades', methods=['GET', 'POST'])
@login_required
@role_required('candidato')
def candidato_habilidades():
    candidato = get_candidato_actual()
    if not candidato:
        flash('Perfil de candidato no encontrado', 'error')
        return redirect(url_for('login'))
    

    todas_habilidades = execute_query("SELECT * FROM Habilidades")
    todas_competencias = execute_query("SELECT * FROM Competencias")
    

    habilidades_actuales = execute_query(
        "SELECT h.HabilidadID FROM CandidatoHabilidades ch "
        "JOIN Habilidades h ON ch.HabilidadID = h.HabilidadID "
        "WHERE ch.CandidatoID = ?",
        (candidato['CandidatoID'],)
    )
    habilidades_actuales = [h['HabilidadID'] for h in habilidades_actuales]
    
    competencias_actuales = execute_query(
        "SELECT c.CompetenciaID FROM CandidatoCompetencias cc "
        "JOIN Competencias c ON cc.CompetenciaID = c.CompetenciaID "
        "WHERE cc.CandidatoID = ?",
        (candidato['CandidatoID'],)
    )
    competencias_actuales = [c['CompetenciaID'] for c in competencias_actuales]
    
    if request.method == 'POST':
        try:

            habilidades_seleccionadas = [int(h) for h in request.form.getlist('habilidades')]
            competencias_seleccionadas = [int(c) for c in request.form.getlist('competencias')]
            
     
            execute_query(
                "DELETE FROM CandidatoHabilidades WHERE CandidatoID = ?",
                (candidato['CandidatoID'],),
                fetch=False
            )
            for habilidad_id in habilidades_seleccionadas:
                execute_query(
                    "INSERT INTO CandidatoHabilidades (CandidatoID, HabilidadID) VALUES (?, ?)",
                    (candidato['CandidatoID'], habilidad_id),
                    fetch=False
                )
            
      
            execute_query(
                "DELETE FROM CandidatoCompetencias WHERE CandidatoID = ?",
                (candidato['CandidatoID'],),
                fetch=False
            )
            for competencia_id in competencias_seleccionadas:
                execute_query(
                    "INSERT INTO CandidatoCompetencias (CandidatoID, CompetenciaID) VALUES (?, ?)",
                    (candidato['CandidatoID'], competencia_id),
                    fetch=False
                )
            
            flash('Habilidades actualizadas correctamente.', 'success')
            return redirect(url_for('candidato_habilidades'))
        
        except Exception as e:
            current_app.logger.error(f"Error al actualizar habilidades: {str(e)}")
            flash('Ocurrió un error al actualizar las habilidades', 'error')
    
    return render_template('candidato/habilidades.html', 
                         candidato=candidato,  
                         todas_habilidades=todas_habilidades,
                         todas_competencias=todas_competencias,
                         habilidades_actuales=habilidades_actuales,
                         competencias_actuales=competencias_actuales)

@app.route('/candidato/vacantes')
@login_required
@role_required('candidato')
def candidato_vacantes():
    candidato = get_candidato_actual()
    if not candidato:
        flash('Perfil de candidato no encontrado', 'error')
        return redirect(url_for('login'))

 
    query = request.args.get('q', '').strip().lower()
    modalidad = request.args.get('modalidad', '').strip().lower()
    grado_estudios = request.args.get('grado_estudios', '').strip()

    sql = """
    SELECT 
        v.VacanteID as id,
        v.Puesto,
        e.Nombre as empresa_nombre,
        v.Ubicacion,
        v.GradoEstudios as grado_estudios,
        v.Modalidad as modalidad,
        v.TipoContrato as tipo_contrato,
        FORMAT(v.FechaPublicacion, 'dd/MM/yyyy') as fecha_publicacion,
        v.PlazasDisponibles,
        v.Resumen,
        v.Salario,
        v.ExperienciaRequerida,
        v.Beneficios,
        CASE WHEN EXISTS (
            SELECT 1 FROM Postulaciones p 
            WHERE p.VacanteID = v.VacanteID 
            AND p.CandidatoID = ?
        ) THEN 1 ELSE 0 END as ya_postulado
    FROM Vacantes v
    JOIN Empresas e ON v.EmpresaID = e.EmpresaID
    WHERE v.Estatus = 'aprobada'
    AND v.PlazasDisponibles > 0
    """
    
    params = [candidato['CandidatoID']]


    if query:
        sql += """
            AND (LOWER(v.Puesto) LIKE ? OR 
                 LOWER(v.Resumen) LIKE ? OR 
                 LOWER(e.Nombre) LIKE ? OR 
                 LOWER(v.Ubicacion) LIKE ?)
        """
        params.extend([f"%{query}%"] * 4)
    
    if modalidad:
        sql += " AND LOWER(v.Modalidad) = ?"
        params.append(modalidad)
    
    if grado_estudios:
        sql += " AND v.GradoEstudios = ?"
        params.append(grado_estudios)


    sql += " ORDER BY v.FechaPublicacion DESC"

    try:
        vacantes = execute_query(sql, params)
    except Exception as e:
        print(f"Error en la consulta SQL: {e}")
        flash('Error al cargar las vacantes', 'error')
        vacantes = []
    
    return render_template('candidato/vacantes.html',
                         vacantes=vacantes,
                         search_query=query,
                         modalidad_filter=modalidad,
                         grado_filter=grado_estudios)

@app.route('/candidato/vacantes/<int:vacante_id>')
@login_required
@role_required('candidato')
def candidato_ver_vacante(vacante_id):
    candidato = get_candidato_actual()
    if not candidato:
        flash('Perfil de candidato no encontrado', 'error')
        return redirect(url_for('login'))
    

    vacante = execute_query(
        """
        SELECT 
            v.VacanteID as id,
            v.Puesto as puesto,
            v.GradoEstudios as grado_estudios,
            v.Resumen as resumen,
            v.Plazas as plazas,
            v.PlazasDisponibles,
            v.Estatus,
            v.FechaPublicacion,
            v.Salario as salario,
            v.TipoContrato as tipo_contrato,
            v.Modalidad as modalidad,
            v.Ubicacion as ubicacion,
            v.ExperienciaRequerida as experiencia_requerida,
            v.Beneficios as beneficios,
            e.Nombre as empresa_nombre,
            e.Logo as empresa_logo,
            (
                SELECT STRING_AGG(h.Nombre, ', ' ORDER BY h.Nombre)
                FROM VacanteHabilidadesRequeridas vhr
                JOIN Habilidades h ON vhr.HabilidadID = h.HabilidadID
                WHERE vhr.VacanteID = v.VacanteID
            ) as habilidades_requeridas,
            (
                SELECT STRING_AGG(h.Nombre, ', ' ORDER BY h.Nombre)
                FROM VacanteHabilidadesOpcionales vho
                JOIN Habilidades h ON vho.HabilidadID = h.HabilidadID
                WHERE vho.VacanteID = v.VacanteID
            ) as habilidades_opcionales,
            CASE WHEN EXISTS (
                SELECT 1 FROM Postulaciones p 
                WHERE p.VacanteID = v.VacanteID 
                AND p.CandidatoID = ?
            ) THEN 1 ELSE 0 END as ya_postulado
        FROM Vacantes v
        JOIN Empresas e ON v.EmpresaID = e.EmpresaID
        WHERE v.VacanteID = ? AND v.Estatus = 'aprobada'
        """,
        (candidato['CandidatoID'], vacante_id)
    )
       
    if not vacante:
        flash('Vacante no encontrada.', 'error')
        return redirect(url_for('candidato_vacantes'))
    
    vacante = vacante[0]
    
 
    vacante['fecha_publicacion'] = vacante['FechaPublicacion'].strftime('%d/%m/%Y') if vacante['FechaPublicacion'] else 'No especificada'
    

    vacante['disponibilidad'] = "Disponible" if vacante['PlazasDisponibles'] > 0 else "Agotado"
    vacante['habilidades_requeridas'] = [h.strip() for h in vacante['habilidades_requeridas'].split(',')] if vacante['habilidades_requeridas'] else []
    vacante['habilidades_opcionales'] = [h.strip() for h in vacante['habilidades_opcionales'].split(',')] if vacante['habilidades_opcionales'] else []
    vacante['empresa_logo'] = vacante['empresa_logo'] if vacante['empresa_logo'] else 'images/default-company.png'
    vacante['salario'] = vacante['salario'] if vacante['salario'] else 'Negociable'
    vacante['ubicacion'] = vacante['ubicacion'] if vacante['ubicacion'] else 'No especificada'
    vacante['beneficios'] = vacante['beneficios'] if vacante['beneficios'] else 'No especificados'
    

    vacante['responsabilidades'] = [] 
    vacante['requisitos'] = []  
    
    return render_template('candidato/ver_vacante.html', 
                         vacante=vacante,
                         ya_postulado=vacante['ya_postulado'])


@app.route('/candidato/postular/<int:vacante_id>')
@login_required
@role_required('candidato')
def candidato_postular(vacante_id):
    candidato = get_candidato_actual()
    if not candidato:
        flash('Perfil de candidato no encontrado', 'error')
        return redirect(url_for('login'))
    

    vacante = execute_query(
        "SELECT 1 FROM Vacantes WHERE VacanteID = ? AND Estatus = 'aprobada' AND PlazasDisponibles > 0",
        (vacante_id,)
    )
    if not vacante:
        flash('Vacante no disponible.', 'error')
        return redirect(url_for('candidato_vacantes'))
    

    ya_postulado = execute_query(
        "SELECT 1 FROM Postulaciones WHERE VacanteID = ? AND CandidatoID = ?",
        (vacante_id, candidato['CandidatoID'])
    )
    if ya_postulado:
        flash('Ya te has postulado a esta vacante.', 'warning')
        return redirect(url_for('candidato_ver_vacante', vacante_id=vacante_id))
    

    if not candidato.get('CV'):
        flash('Debes subir tu CV antes de postularte a una vacante.', 'warning')
        return redirect(url_for('candidato_perfil'))
    
    try:

        execute_query(
            "INSERT INTO Postulaciones (VacanteID, CandidatoID, Estatus) VALUES (?, ?, 'pendiente')",
            (vacante_id, candidato['CandidatoID']),
            fetch=False
        )
        
        flash('Postulación exitosa. La empresa revisará tu perfil.', 'success')
        return redirect(url_for('candidato_ver_vacante', vacante_id=vacante_id))
    
    except Exception as e:
        current_app.logger.error(f"Error al postular: {str(e)}")
        flash('Ocurrió un error al procesar tu postulación', 'error')
        return redirect(url_for('candidato_vacantes'))

@app.route('/candidato/postulaciones')
@login_required
@role_required('candidato')
def candidato_postulaciones():
    candidato = get_candidato_actual()
    if not candidato:
        flash('Perfil de candidato no encontrado', 'error')
        return redirect(url_for('login'))
    

    postulaciones = execute_query(
        """SELECT p.*, v.Puesto, v.Estatus as VacanteEstatus, 
        e.Nombre as EmpresaNombre, e.Logo as EmpresaLogo
        FROM Postulaciones p
        JOIN Vacantes v ON p.VacanteID = v.VacanteID
        JOIN Empresas e ON v.EmpresaID = e.EmpresaID
        WHERE p.CandidatoID = ?
        ORDER BY p.FechaPostulacion DESC""",
        (candidato['CandidatoID'],)
    )
    
    return render_template('candidato/postulaciones.html', postulaciones=postulaciones)

@app.route('/candidato/cancelar_postulacion/<int:postulacion_id>')
@login_required
@role_required('candidato')
def candidato_cancelar_postulacion(postulacion_id):
    candidato = get_candidato_actual()
    if not candidato:
        flash('Perfil de candidato no encontrado', 'error')
        return redirect(url_for('login'))
    
    try:

        postulacion = execute_query(
            "SELECT * FROM Postulaciones WHERE PostulacionID = ? AND CandidatoID = ?",
            (postulacion_id, candidato['CandidatoID'])
        )
        
        if not postulacion:
            flash('Postulación no encontrada.', 'error')
            return redirect(url_for('candidato_postulaciones'))
        

        execute_query(
            "DELETE FROM Postulaciones WHERE PostulacionID = ?",
            (postulacion_id,),
            fetch=False
        )
        
        flash('Postulación cancelada correctamente.', 'success')
        return redirect(url_for('candidato_postulaciones'))
    
    except Exception as e:
        current_app.logger.error(f"Error al cancelar postulación: {str(e)}")
        flash('Ocurrió un error al cancelar la postulación', 'error')
        return redirect(url_for('candidato_postulaciones'))




# ==================== CHATBOT INTELIGENTE MEJORADO v2 ====================
#
# Cambios respecto a la versión anterior:
#   1. Fix: "analizar otra vacante" ahora funciona (memoria de conversación en sesión).
#   2. Fix: respuesta_default ahora compara en minúsculas correctamente.
#   3. Fix: se elimina el patrón de "compatibilidad" muerto dentro de self.respuestas
#      (nunca se alcanzaba porque procesar_mensaje ya lo intercepta antes).
#   4. Fix: import random movido al principio del archivo.
#   5. Seguridad: todo dato dinámico (puesto, habilidades, nombres) se escapa con
#      html.escape antes de insertarse en las respuestas, y el frontend además
#      escapa cualquier HTML antes de aplicar el formato tipo markdown -> defensa
#      en profundidad contra XSS almacenado (p. ej. una empresa que registre un
#      puesto con <img src=x onerror=...>).
#   6. Normalización de acentos para que "qué", "cómo", etc. also disparen los
#      patrones (que están escritos sin tilde).
#   7. Precómputo de stems de las habilidades del candidato (ya no se recalculan
#      en cada iteración del bucle anidado).
#   8. Se puede pedir el análisis de una vacante específica por nombre de puesto
#      ("analiza mi compatibilidad con Analista de Datos"), y no solo la más
#      reciente.
#   9. Validación de longitud de mensaje también en el servidor (antes solo
#      existía en el cliente).
#  10. Memoria de conversación básica en sesión: recuerda las últimas vacantes
#      mostradas para poder avanzar con "analizar otra vacante".

import random
import re
import unicodedata
import html
from difflib import SequenceMatcher

import nltk
from nltk.stem import SnowballStemmer
from flask import session

# Descargar recursos de NLTK (solo primera vez)
try:
    nltk.data.find('tokenizers/punkt')
except LookupError:
    nltk.download('punkt')

MAX_MENSAJE_LEN = 500


def normalizar_texto(texto):
    """Quita acentos y pasa a minúsculas, para que los patrones sin tilde
    (ej. 'que tan compatible') también capturen 'qué tan compatible'."""
    texto = texto.lower().strip()
    texto_sin_acentos = unicodedata.normalize('NFKD', texto)
    texto_sin_acentos = ''.join(c for c in texto_sin_acentos if not unicodedata.combining(c))
    return texto_sin_acentos


def texto_seguro(valor):
    """Escapa cualquier dato que provenga de la base de datos (puesto,
    nombre de habilidad, nombre de candidato, etc.) antes de insertarlo en
    una respuesta que el frontend va a renderizar como HTML. Defensa en
    profundidad: aunque el frontend también sanitiza, esto evita que una
    empresa inyecte HTML/JS a través del nombre de una vacante o habilidad."""
    if valor is None:
        return ''
    return html.escape(str(valor), quote=True)


class ChatbotInteligente:
    def __init__(self):
        self.stemmer = SnowballStemmer('spanish')
        self.respuestas = self.cargar_respuestas()
        self.habilidades_base = {
            'python': ['programación', 'desarrollo', 'backend', 'data science', 'machine learning'],
            'java': ['programación', 'desarrollo', 'backend', 'android', 'enterprise'],
            'javascript': ['programación', 'frontend', 'web', 'react', 'angular', 'vue'],
            'sql': ['base de datos', 'datos', 'análisis', 'consultas'],
            'html': ['diseño web', 'frontend', 'maquetación'],
            'css': ['diseño web', 'frontend', 'estilos'],
            'flask': ['python', 'web', 'backend', 'api'],
            'git': ['control de versiones', 'colaboración', 'código'],
            'docker': ['contenedores', 'devops', 'despliegue'],
            'aws': ['nube', 'cloud', 'infraestructura'],
            'machine learning': ['ia', 'inteligencia artificial', 'datos', 'modelos'],
            'analisis de datos': ['datos', 'estadística', 'visualización', 'power bi'],
            'trabajo en equipo': ['colaboración', 'comunicación', 'liderazgo'],
            'comunicacion': ['presentaciones', 'escucha activa', 'negociación'],
            'liderazgo': ['gestión', 'equipos', 'toma de decisiones'],
            'resolucion de problemas': ['análisis', 'crítico', 'solución'],
        }
        self.palabras_clave_vacantes = {
            'desarrollador': ['programación', 'código', 'software', 'python', 'java'],
            'analista': ['datos', 'análisis', 'reportes', 'sql', 'excel'],
            'administrador': ['gestión', 'organización', 'liderazgo', 'planeación'],
            'diseñador': ['creativo', 'diseño', 'ux', 'ui', 'photoshop'],
            'ingeniero': ['técnico', 'solución', 'mejora', 'procesos'],
            'gerente': ['liderazgo', 'equipos', 'estrategia', 'resultados'],
            'soporte': ['atención', 'usuarios', 'técnico', 'ayuda'],
            'consultor': ['estrategia', 'mejora', 'optimización', 'recomendaciones'],
            'marketing': ['estrategias', 'digital', 'redes', 'publicidad'],
            'recursos humanos': ['reclutamiento', 'selección', 'gestión del talento'],
            'finanzas': ['presupuestos', 'contabilidad', 'inversiones', 'análisis financiero'],
        }

    def cargar_respuestas(self):
        """Base de conocimientos del chatbot.
        Nota: ya NO incluye un patrón para 'compatibilidad' aquí, porque
        procesar_mensaje intercepta esa intención antes de llegar a este
        diccionario (el patrón anterior era código muerto)."""
        return {
            # Saludos
            r'\b(holla|hola|buenas|que tal|hey|saludos)\b': [
                "¡Hola! Soy el asistente virtual de TalentUPQ. ¿En qué puedo ayudarte? 😊",
                "¡Bienvenido/a! Estoy aquí para ayudarte con el proceso de búsqueda de empleo. ¿Qué necesitas?",
                "¡Hola! ¿Cómo puedo asistirte hoy?"
            ],

            # Cómo postularse
            r'\b(postul\w*|aplicar|candidatar|como aplicar)\b': [
                "¡Con gusto te explico! Para postularte a una vacante:\n"
                "1. 🔑 Inicia sesión en tu cuenta\n"
                "2. 📝 Verifica que tu perfil esté completo (CV, experiencia y habilidades), un perfil incompleto reduce tus posibilidades\n"
                "3. 🔍 Ve a la sección 'Vacantes' y usa los filtros (modalidad, área, grado) para encontrar algo que se ajuste a ti\n"
                "4. 📖 Lee bien los requisitos antes de aplicar, así evitas postularte a algo que no encaje\n"
                "5. ✅ Haz clic en 'Postular' dentro del detalle de la vacante\n"
                "6. 📊 Da seguimiento desde 'Mis Postulaciones', ahí verás si está pendiente, aceptada o rechazada\n\n"
                "💡 Tip: si quieres saber qué tan buen candidato eres antes de postularte, puedo analizar tu compatibilidad con la vacante, solo dime 'analiza mi compatibilidad'.\n\n"
                "¿Te ayudo a revisar tu perfil o a buscar una vacante específica?"
            ],

            # Crear vacante (empresa)
            r'\b(crear vacante|publicar vacante|nueva vacante)\b': [
                "Para publicar una vacante:\n1. Ve a tu panel de empresa\n2. Haz clic en 'Nueva Vacante'\n3. Completa todos los campos (puesto, requisitos, etc.)\n4. Espera la aprobación del administrador\n\n¿Te ayudo con algún campo específico?"
            ],

            # Requisitos para postular
            r'\b(requisitos|necesito|necesario|que se necesita)\b': [
                "Los requisitos básicos son:\n✅ Tener un perfil completo\n✅ CV actualizado\n✅ Cumplir con los requisitos de la vacante\n✅ Ser estudiante o egresado UPQ\n\n¿Te gustaría saber más sobre algún requisito específico?"
            ],

            # Estado de postulación
            r'\b(estado de mi postulacion|como va mi postulacion|revisar postulacion)\b': [
                "Puedes revisar el estado de tus postulaciones en la sección 'Mis Postulaciones' de tu panel. Los estados posibles son:\n• Pendiente: La empresa aún no la revisa\n• Aceptado: ¡Felicidades! Te contactarán pronto\n• Rechazado: No te desanimes, hay más oportunidades"
            ],

            # Entrevista
            # Entrevista
            r'\b(entrevista\w*|como prepararme)\b': [
                "¡Vamos a prepararte bien! Aquí tienes consejos organizados por etapa:\n\n"
                "📌 **Antes de la entrevista**\n"
                "🎯 Investiga la empresa: qué hace, su misión y noticias recientes\n"
                "🎯 Relee la descripción del puesto y ubica 3-4 puntos donde tu experiencia calce\n"
                "🎯 Prepara 2-3 ejemplos concretos de logros usando el método STAR (Situación, Tarea, Acción, Resultado)\n"
                "🎯 Ten listas tus preguntas para el entrevistador (sobre el equipo, el día a día, retos del puesto)\n\n"
                "📌 **El día de la entrevista**\n"
                "🎯 Llega (o conéctate) 10 minutos antes\n"
                "🎯 Viste apropiado según la cultura de la empresa; ante la duda, formal es más seguro\n"
                "🎯 Si es virtual: prueba cámara, micrófono e internet antes, y cuida el fondo/iluminación\n"
                "🎯 Escucha con atención antes de responder, no hay prisa\n\n"
                "📌 **Después de la entrevista**\n"
                "🎯 Envía un correo de agradecimiento el mismo día o al siguiente\n"
                "🎯 Anota qué preguntas te costaron más, te sirve para la próxima\n\n"
                "¿Quieres que practiquemos alguna pregunta común, como '¿Cuáles son tus fortalezas y debilidades?' o '¿Por qué quieres trabajar aquí?'?"
            ],
            # Currículum/CV
            r'\b(curriculum|cv|hoja de vida)\b': [
                "Consejos para tu CV:\n📄 Manténlo de 1-2 páginas\n📄 Destaca logros, no solo tareas\n📄 Usa palabras clave de la industria\n📄 Revisa ortografía\n📄 Incluye habilidades técnicas y blandas\n\n¿Quieres que revise tu CV?"
            ],

            # Habilidades
            r'\b(habilidades|que habilidades|competencias)\b': [
                "Las habilidades más demandadas actualmente son:\n💻 Python, Java, SQL\n📊 Análisis de datos\n🤝 Trabajo en equipo\n🗣️ Comunicación efectiva\n🎯 Gestión de proyectos\n\n¿Te gustaría agregar habilidades a tu perfil?"
            ],

            # Perfil
            r'\b(completar perfil|actualizar perfil|mi perfil)\b': [
                "Para completar tu perfil:\n1. Ve a 'Mi Perfil'\n2. Completa tus datos personales\n3. Agrega tu experiencia laboral\n4. Sube tu CV y foto\n5. Añade tus habilidades\n\n¡Un perfil completo atrae más oportunidades!"
            ],

            # Vacantes disponibles
            r'\b(vacantes disponibles|que vacantes hay|buscar trabajo|oportunidades)\b': [
                "¡Claro! Puedes ver todas las vacantes disponibles en la sección 'Vacantes'. También puedes filtrar por:\n• Modalidad (presencial/remoto)\n• Tipo de contrato\n• Grado de estudios\n\n¿Te ayudo a buscar algo específico?"
            ],

            # Salario
            r'\b(salario|sueldo|cuanto pagan)\b': [
                "El salario varía según la empresa y el puesto. Puedes ver el rango salarial en cada vacante. ¿Te gustaría que te ayude a buscar vacantes dentro de tu rango esperado?"
            ],

            # Tiempo de respuesta
            r'\b(cuanto tardan|tiempo respuesta|demoran)\b': [
                "El tiempo de respuesta varía por empresa. Generalmente, las empresas responden entre 1-3 semanas. Puedes dar seguimiento desde 'Mis Postulaciones'."
            ],

            # Despedida
            r'\b(gracias|graciass|muchas gracias|ok|excelente)\b': [
                "¡De nada! ¿Necesitas ayuda con algo más? Estoy aquí para ti. 🤗",
                "¡Con gusto! Recuerda que puedes preguntarme lo que necesites sobre el proceso de búsqueda de empleo."
            ],

            # Ayuda general
            r'\b(ayuda|que puedes hacer|comandos|funciones)\b': [
                "Puedo ayudarte con:\n💬 Preguntas sobre postulación\n📝 Consejos para entrevistas\n📄 Mejora de tu CV\n🔍 Análisis de compatibilidad de vacantes\n📊 Estado de tus postulaciones\n\n¿Qué te gustaría saber?"
            ],

            # Contacto
            r'\b(contacto|soporte|quejas|problemas)\b': [
                "Puedes contactar a soporte:\n📧 Email: bolsa.trabajo@upq.edu.mx\n📞 Teléfono: (773) 108-7368\n\n¿Hay algo específico en lo que pueda ayudarte?"
            ]
        }

    # ------------------------------------------------------------------
    # Cálculo de compatibilidad
    # ------------------------------------------------------------------
    def calcular_compatibilidad(self, habilidades_candidato, experiencia_candidato, vacante_data):
        """Calcula el porcentaje de compatibilidad entre un candidato y una vacante."""
        habilidades_requeridas = vacante_data.get('habilidades_requeridas', [])
        experiencia_requerida = vacante_data.get('experiencia_requerida', 0)
        puesto_vacante = vacante_data.get('puesto', '').lower()

        # ===== Convertir experiencia_requerida a número =====
        if isinstance(experiencia_requerida, str):
            numeros = re.findall(r'\d+', experiencia_requerida)
            experiencia_requerida = int(numeros[0]) if numeros else 0
        elif not isinstance(experiencia_requerida, (int, float)):
            experiencia_requerida = 0

        # 1. Análisis de habilidades (60% del total)
        habilidades_candidato_lower = [h.lower().strip() for h in habilidades_candidato]
        habilidades_requeridas_lower = [h.lower().strip() for h in habilidades_requeridas]

        # Precómputo de stems del candidato: antes se recalculaba en cada
        # iteración del bucle anidado (uno por cada habilidad requerida).
        stems_candidato = {cand: self.stemmer.stem(cand) for cand in habilidades_candidato_lower}

        coincidencias_exactas = 0
        coincidencias_parciales = 0

        for req in habilidades_requeridas_lower:
            req_stem = self.stemmer.stem(req)
            for cand in habilidades_candidato_lower:
                cand_stem = stems_candidato[cand]
                if req == cand or req in cand or cand in req:
                    coincidencias_exactas += 1
                    break
                elif req_stem == cand_stem or SequenceMatcher(None, req, cand).ratio() > 0.7:
                    coincidencias_parciales += 0.5
                    break

        habilidades_faltantes = []
        for req in habilidades_requeridas_lower:
            encontrada = any(
                req == cand or req in cand or cand in req
                for cand in habilidades_candidato_lower
            )
            if not encontrada:
                habilidades_faltantes.append(req)

        max_habilidades = len(habilidades_requeridas_lower) if habilidades_requeridas_lower else 1
        puntaje_habilidades = min(((coincidencias_exactas + coincidencias_parciales) / max_habilidades) * 60, 60)

        # 2. Análisis de experiencia (20% del total)
        if experiencia_requerida > 0:
            if experiencia_candidato >= experiencia_requerida:
                puntaje_experiencia = 20
            else:
                puntaje_experiencia = min((experiencia_candidato / experiencia_requerida) * 20, 20)
        else:
            puntaje_experiencia = 15

        # 3. Análisis de puesto (20% del total)
        palabras_candidato = set(habilidades_candidato_lower)
        for h in habilidades_candidato_lower:
            for palabra in h.split():
                palabras_candidato.add(palabra)

        palabras_puesto = set(puesto_vacante.split())

        puntaje_puesto = 0
        for palabra in palabras_puesto:
            for clave, sinonimos in self.palabras_clave_vacantes.items():
                if palabra in clave or clave in palabra:
                    puntaje_puesto += 2
                for sinonimo in sinonimos:
                    if palabra in sinonimo or sinonimo in palabra:
                        puntaje_puesto += 1

        for hab in habilidades_candidato_lower:
            for clave, sinonimos in self.palabras_clave_vacantes.items():
                if hab in clave or clave in hab:
                    puntaje_puesto += 1
                for sinonimo in sinonimos:
                    if hab in sinonimo or sinonimo in hab:
                        puntaje_puesto += 0.5

        puntaje_puesto = min(puntaje_puesto, 20)

        total = min(puntaje_habilidades + puntaje_experiencia + puntaje_puesto, 100)

        recomendaciones = []
        if habilidades_faltantes:
            recomendaciones.append(
                f"💡 Te sugiero desarrollar estas habilidades: {', '.join(habilidades_faltantes[:3])}"
            )
        if puntaje_experiencia < 15:
            recomendaciones.append("📈 Considera obtener más experiencia en proyectos relacionados o prácticas profesionales.")
        if puntaje_puesto < 10:
            recomendaciones.append("🎯 Ajusta tu perfil para destacar habilidades relacionadas con el puesto.")

        if total >= 80:
            nivel, mensaje = "Excelente", "¡Eres un candidato muy compatible! Tienes una gran oportunidad."
        elif total >= 60:
            nivel, mensaje = "Bueno", "Tienes buena compatibilidad. Considera las recomendaciones para mejorar."
        elif total >= 40:
            nivel, mensaje = "Regular", "Hay margen de mejora. Trabaja en las habilidades sugeridas."
        else:
            nivel, mensaje = "Bajo", "Te sugerimos enfocarte en desarrollar las habilidades requeridas para esta vacante."

        return {
            'porcentaje': round(total, 1),
            'nivel': nivel,
            'mensaje': mensaje,
            'habilidades_coincidentes': [h for h in habilidades_requeridas_lower if h in habilidades_candidato_lower],
            'habilidades_faltantes': habilidades_faltantes[:5],
            'puntaje_habilidades': round(puntaje_habilidades, 1),
            'puntaje_experiencia': round(puntaje_experiencia, 1),
            'puntaje_puesto': round(puntaje_puesto, 1),
            'recomendaciones': recomendaciones[:3]
        }

    # ------------------------------------------------------------------
    # Helpers de acceso a datos
    # ------------------------------------------------------------------
    def _obtener_candidato(self, usuario_id):
        candidato = execute_query(
            "SELECT CandidatoID, Nombre, ApellidoPaterno FROM Candidatos WHERE UsuarioID = ?",
            [usuario_id]
        )
        return candidato[0] if candidato else None

    def _obtener_habilidades_candidato(self, candidato_id):
        habilidades = execute_query("""
            SELECT h.Nombre
            FROM CandidatoHabilidades ch
            JOIN Habilidades h ON ch.HabilidadID = h.HabilidadID
            WHERE ch.CandidatoID = ?
        """, [candidato_id])
        return [h['Nombre'] for h in habilidades]

    def _obtener_anios_experiencia(self, candidato_id):
        experiencia = execute_query("""
            SELECT COALESCE(
                EXTRACT(YEAR FROM AGE(CURRENT_DATE, MIN(FechaIngreso)))::INTEGER,
                0
            ) as Anios
            FROM ExperienciaLaboral
            WHERE CandidatoID = ?
        """, [candidato_id])
        return experiencia[0]['Anios'] if experiencia else 0

    def _obtener_vacantes_activas(self):
        """Trae hasta 5 vacantes aprobadas (antes solo 3), para tener margen
        al recorrerlas con 'analizar otra vacante'."""
        return execute_query("""
            SELECT
                v.VacanteID,
                v.Puesto,
                v.ExperienciaRequerida,
                (
                    SELECT STRING_AGG(h.Nombre, ', ' ORDER BY h.Nombre)
                    FROM VacanteHabilidadesRequeridas vh
                    JOIN Habilidades h ON vh.HabilidadID = h.HabilidadID
                    WHERE vh.VacanteID = v.VacanteID
                ) as HabilidadesRequeridas
            FROM Vacantes v
            WHERE v.Estatus = 'aprobada'
            ORDER BY v.FechaPublicacion DESC
            LIMIT 5
        """)

    def _formatear_resultado(self, vacante, resultado):
        """Construye el texto de respuesta escapando cualquier dato dinámico
        (puesto, habilidades) para evitar XSS almacenado si una empresa
        registró texto malicioso en esos campos."""
        puesto_seguro = texto_seguro(vacante['Puesto'])

        respuesta = f"""🔍 **Análisis de compatibilidad con la vacante: {puesto_seguro}**

📊 **Porcentaje de compatibilidad: {resultado['porcentaje']}%** ({resultado['nivel']})

📈 **Desglose:**
• Habilidades: {resultado['puntaje_habilidades']}/60
• Experiencia: {resultado['puntaje_experiencia']}/20
• Alineación con puesto: {resultado['puntaje_puesto']}/20

🎯 {texto_seguro(resultado['mensaje'])}

"""
        if resultado['habilidades_faltantes']:
            faltantes_seguras = ', '.join(texto_seguro(h) for h in resultado['habilidades_faltantes'])
            respuesta += f"\n⚠️ **Habilidades a desarrollar:**\n• {faltantes_seguras}"

        if resultado['recomendaciones']:
            recomendaciones_seguras = "\n• ".join(texto_seguro(r) for r in resultado['recomendaciones'])
            respuesta += f"\n\n💡 **Recomendaciones:**\n• {recomendaciones_seguras}"

        respuesta += "\n\n📌 ¿Te gustaría analizar otra vacante? Solo dime 'analizar otra vacante'."
        return respuesta

    # ------------------------------------------------------------------
    # Lógica de compatibilidad con memoria de conversación
    # ------------------------------------------------------------------
    def _analizar_vacante_por_indice(self, candidato, habilidades_lista, anios_experiencia, vacantes, indice):
        """Analiza la vacante en `indice` de la lista y guarda el progreso
        en sesión para que 'analizar otra vacante' avance correctamente."""
        if indice >= len(vacantes):
            return ("Ya te mostré el análisis de todas las vacantes activas disponibles por ahora. "
                    "¿Te gustaría que te ayude a buscar otras oportunidades o mejorar tu perfil? 😊")

        vacante = vacantes[indice]
        habilidades_requeridas = vacante['HabilidadesRequeridas'].split(', ') if vacante['HabilidadesRequeridas'] else []

        vacante_data = {
            'habilidades_requeridas': habilidades_requeridas,
            'experiencia_requerida': vacante['ExperienciaRequerida'],
            'puesto': vacante['Puesto']
        }

        resultado = self.calcular_compatibilidad(habilidades_lista, anios_experiencia, vacante_data)

        # Guardar contexto de conversación en sesión para soportar
        # "analizar otra vacante" en el siguiente mensaje.
        session['chatbot_vacante_ids'] = [v['VacanteID'] for v in vacantes]
        session['chatbot_indice_actual'] = indice + 1

        return self._formatear_resultado(vacante, resultado)

    def _buscar_vacante_por_nombre(self, vacantes, mensaje_normalizado):
        """Permite pedir el análisis de una vacante específica por nombre
        de puesto, ej: 'analiza mi compatibilidad con analista de datos'."""
        mejor_indice = None
        mejor_score = 0.0
        for i, vacante in enumerate(vacantes):
            puesto_normalizado = normalizar_texto(vacante['Puesto'])
            if puesto_normalizado and puesto_normalizado in mensaje_normalizado:
                return i
            score = SequenceMatcher(None, puesto_normalizado, mensaje_normalizado).ratio()
            if score > mejor_score:
                mejor_score = score
                mejor_indice = i
        return mejor_indice if mejor_score > 0.4 else None

    def _manejar_compatibilidad(self, mensaje_normalizado, usuario_id):
        if not usuario_id:
            return "No pude encontrar tu información de usuario. Asegúrate de haber iniciado sesión correctamente."

        candidato = self._obtener_candidato(usuario_id)
        if not candidato:
            return "No encontré tu perfil de candidato. Asegúrate de haber completado tu registro como candidato."

        habilidades_lista = self._obtener_habilidades_candidato(candidato['CandidatoID'])
        anios_experiencia = self._obtener_anios_experiencia(candidato['CandidatoID'])
        vacantes = self._obtener_vacantes_activas()

        if not vacantes:
            return "No encontré vacantes activas en el sistema para analizar tu compatibilidad. ¿Te gustaría explorar otras opciones?"

        # ¿El usuario pide continuar con "otra vacante"?
        pide_otra = bool(re.search(r'\b(otra vacante|siguiente vacante|otra mas|otra más)\b', mensaje_normalizado))

        if pide_otra:
            ids_previos = session.get('chatbot_vacante_ids')
            indice_actual = session.get('chatbot_indice_actual', 0)
            if ids_previos:
                # Reconstruir la lista en el mismo orden que se mostró antes,
                # usando los datos frescos de `vacantes` cuando sea posible.
                vacantes_por_id = {v['VacanteID']: v for v in vacantes}
                vacantes_ordenadas = [vacantes_por_id[i] for i in ids_previos if i in vacantes_por_id]
                return self._analizar_vacante_por_indice(
                    candidato, habilidades_lista, anios_experiencia, vacantes_ordenadas, indice_actual
                )
            # No había contexto previo: se comporta como una nueva consulta.
            return self._analizar_vacante_por_indice(candidato, habilidades_lista, anios_experiencia, vacantes, 0)

        # ¿El usuario menciona el nombre de una vacante específica?
        indice_especifico = self._buscar_vacante_por_nombre(vacantes, mensaje_normalizado)
        indice_inicial = indice_especifico if indice_especifico is not None else 0

        return self._analizar_vacante_por_indice(candidato, habilidades_lista, anios_experiencia, vacantes, indice_inicial)

    # ------------------------------------------------------------------
    # Entrada principal
    # ------------------------------------------------------------------
    def procesar_mensaje(self, mensaje, usuario_id=None):
        """Procesa el mensaje del usuario y genera respuesta."""
        if not mensaje or not mensaje.strip():
            return "Parece que enviaste un mensaje vacío. ¿En qué puedo ayudarte?"

        if len(mensaje) > MAX_MENSAJE_LEN:
            return f"⟫ Tu mensaje es demasiado largo (máximo {MAX_MENSAJE_LEN} caracteres). ¿Puedes resumirlo?"

        mensaje_normalizado = normalizar_texto(mensaje)

        # Verificar si es una consulta de compatibilidad (incluye seguimiento
        # con "otra vacante" y búsqueda por nombre de puesto).
        if re.search(r'\b(compatibilidad|que tan compatible|match|ajuste|otra vacante|siguiente vacante)\b', mensaje_normalizado):
            return self._manejar_compatibilidad(mensaje_normalizado, usuario_id)

        # Buscar coincidencias con los patrones generales
        for patron, respuestas in self.respuestas.items():
            if re.search(patron, mensaje_normalizado):
                return random.choice(respuestas)

        return self.respuesta_default(mensaje_normalizado)

    def respuesta_default(self, mensaje_normalizado):
        """Respuesta cuando no entiende la pregunta.
        Recibe el mensaje ya normalizado (minúsculas, sin acentos)."""
        temas = ["postular", "vacante", "cv", "entrevista", "habilidades", "salario"]

        for tema in temas:
            if tema in mensaje_normalizado:
                return f"Parece que preguntas sobre '{tema}'. ¿Podrías ser más específico/a? Estoy aquí para ayudarte."

        return """Lo siento, no entendí tu pregunta. Puedo ayudarte con:
• Análisis de compatibilidad de vacantes
• Proceso de postulación
• Consejos para entrevistas
• Mejora de tu CV
• Información de vacantes

Escribe "ayuda" para ver más opciones."""

    def obtener_sugerencias(self):
        """Sugerencias rápidas para el usuario"""
        return [
            "¿Cómo me postulo a una vacante?",
            "Analiza mi compatibilidad",
            "Analizar otra vacante",
            "Consejos para entrevistas",
            "¿Cómo mejorar mi CV?",
            "¿Qué vacantes hay disponibles?"
        ]


# Instancia global del chatbot
chatbot = ChatbotInteligente()


@app.route('/chatbot', methods=['GET', 'POST'])
@login_required
def chatbot_view():
    """Vista del chatbot"""
    if request.method == 'POST':
        data = request.get_json(silent=True) or {}
        mensaje = (data.get('mensaje') or '').strip()

        if not mensaje:
            return jsonify({'success': False, 'error': 'Mensaje vacío'})

        if len(mensaje) > MAX_MENSAJE_LEN:
            return jsonify({
                'success': False,
                'error': f'El mensaje excede el límite de {MAX_MENSAJE_LEN} caracteres'
            })

        respuesta = chatbot.procesar_mensaje(mensaje, current_user.id)
        return jsonify({
            'success': True,
            'respuesta': respuesta,
            'sugerencias': chatbot.obtener_sugerencias()
        })

    return render_template('chatbot.html',
                           sugerencias=chatbot.obtener_sugerencias())



@app.route('/empresa')
@login_required
@role_required('empresa')
def empresa_dashboard():
    empresa = get_empresa_actual()
    if not empresa:
        flash('Perfil de empresa no encontrado', 'error')
        return redirect(url_for('login'))

    # Vacantes de la empresa
    vacantes_empresa = execute_query(
        """SELECT v.*, 
        (SELECT COUNT(*) FROM Postulaciones p WHERE p.VacanteID = v.VacanteID) as NumPostulaciones,
        CASE v.Estatus
            WHEN 'en_revision' THEN 'badge-warning'
            WHEN 'aprobada' THEN 'badge-success'
            WHEN 'rechazada' THEN 'badge-danger'
            WHEN 'cerrada' THEN 'badge-secondary'
            ELSE 'badge-light'
        END as EstadoClase
        FROM Vacantes v
        WHERE v.EmpresaID = ?
        ORDER BY v.FechaPublicacion DESC LIMIT 3""",
        (empresa['EmpresaID'],)
    )

    # Postulaciones pendientes
    postulaciones_recientes = execute_query(
        """SELECT p.*, c.Nombre as CandidatoNombre, 
        c.ApellidoPaterno as CandidatoApellido, v.Puesto as VacantePuesto
        FROM Postulaciones p
        JOIN Candidatos c ON p.CandidatoID = c.CandidatoID
        JOIN Vacantes v ON p.VacanteID = v.VacanteID
        WHERE v.EmpresaID = ? AND p.Estatus = 'pendiente'
        ORDER BY p.FechaPostulacion DESC LIMIT 3""",
        (empresa['EmpresaID'],)
    )

    # Notificaciones
    notificaciones = execute_query(
        "SELECT COUNT(*) as Count FROM Notificaciones WHERE EmpresaID = ? AND Leida = 0",
        (empresa['EmpresaID'],)
    )
    num_notificaciones = notificaciones[0]['Count'] if notificaciones else 0

    # ========== NUEVO: Conversaciones recientes ==========
    conversaciones_recientes = []
    no_leidos_empresa = 0
    
    try:
        # Obtener las 5 conversaciones más recientes
        conversaciones_recientes = execute_query(
            """SELECT
               c.ConversacionID,
               c.VacanteID,
               c.CandidatoID,
               c.FechaInicio,
               v.Puesto as VacantePuesto,
               cand.Nombre as CandidatoNombre,
               cand.ApellidoPaterno as CandidatoApellido,
               cand.FotoPerfil as CandidatoFoto,
               (SELECT Mensaje FROM Mensajes 
                WHERE ConversacionID = c.ConversacionID 
                ORDER BY FechaEnvio DESC LIMIT 1) as UltimoMensaje,
               (SELECT FechaEnvio AT TIME ZONE 'UTC' AT TIME ZONE 'America/Mexico_City' FROM Mensajes 
                WHERE ConversacionID = c.ConversacionID 
                ORDER BY FechaEnvio DESC LIMIT 1) as UltimoMensajeFecha,
               (SELECT COUNT(*) FROM Mensajes 
                WHERE ConversacionID = c.ConversacionID 
                AND RemitenteTipo = 'candidato' 
                AND Leido = 0) as NoLeidos
               FROM Conversaciones c
               JOIN Vacantes v ON c.VacanteID = v.VacanteID
               JOIN Candidatos cand ON c.CandidatoID = cand.CandidatoID
               WHERE c.EmpresaID = ? AND c.Activa = 1
               ORDER BY UltimoMensajeFecha DESC NULLS LAST LIMIT 5""",
            (empresa['EmpresaID'],)
        )
        
        # Obtener total de mensajes no leídos para el badge
        result_no_leidos = execute_query(
            """SELECT COUNT(*) as Total FROM Mensajes m
               JOIN Conversaciones c ON m.ConversacionID = c.ConversacionID
               WHERE c.EmpresaID = ? AND m.RemitenteTipo = 'candidato' AND m.Leido = 0""",
            (empresa['EmpresaID'],)
        )
        if result_no_leidos:
            no_leidos_empresa = result_no_leidos[0]['Total']
            
    except Exception as e:
        # Si hay error (tabla no existe o algo), simplemente ignorar
        print(f"Error al obtener conversaciones: {e}")
        conversaciones_recientes = []
        no_leidos_empresa = 0

    return render_template('empresa/dashboard.html',
                         empresa=empresa,
                         vacantes_empresa=vacantes_empresa,
                         postulaciones_recientes=postulaciones_recientes,
                         num_notificaciones=num_notificaciones,
                         conversaciones_recientes=conversaciones_recientes,
                         no_leidos_empresa=no_leidos_empresa)  # Nuevas variables




@app.route('/empresa/perfil', methods=['GET', 'POST'])
@login_required
@role_required('empresa')
def empresa_perfil():
    empresa = get_empresa_actual()
    if not empresa:
        flash('Perfil de empresa no encontrado', 'error')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        try:

            logo = empresa.get('Logo', '')
            if 'logo' in request.files:
                file = request.files['logo']
                if file and allowed_file(file.filename):
                    filename = secure_filename(f"logo_{empresa['EmpresaID']}.{file.filename.rsplit('.', 1)[1].lower()}")
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    logo = filename
            

            execute_query(
                """UPDATE Empresas SET
                Nombre = ?, Direccion = ?, Telefono = ?,
                SitioWeb = ?, Descripcion = ?, Logo = ?
                WHERE EmpresaID = ?""",
                (
                    request.form['nombre'],
                    request.form['direccion'],
                    request.form['telefono'],
                    request.form['sitio_web'],
                    request.form['descripcion'],
                    logo,
                    empresa['EmpresaID']
                ),
                fetch=False
            )
            
            flash('Perfil de empresa actualizado correctamente.', 'success')
            return redirect(url_for('empresa_perfil'))
        
        except Exception as e:
            current_app.logger.error(f"Error al actualizar perfil de empresa: {str(e)}")
            flash('Ocurrió un error al actualizar el perfil', 'error')
    
    return render_template('empresa/perfil.html', empresa=empresa)

@app.route('/empresa/vacantes')
@login_required
@role_required('empresa')
def empresa_vacantes():
    empresa = get_empresa_actual()
    if not empresa:
        flash('Perfil de empresa no encontrado', 'error')
        return redirect(url_for('login'))
    

    vacantes = execute_query(
        """SELECT v.*, 
        (SELECT COUNT(*) FROM Postulaciones p WHERE p.VacanteID = v.VacanteID) as NumPostulaciones,
        CASE v.Estatus
            WHEN 'en_revision' THEN 'badge-warning'
            WHEN 'aprobada' THEN 'badge-success'
            WHEN 'rechazada' THEN 'badge-danger'
            WHEN 'cerrada' THEN 'badge-secondary'
            ELSE 'badge-light'
        END as EstadoClase
        FROM Vacantes v
        WHERE v.EmpresaID = ?
        ORDER BY v.FechaPublicacion DESC""",
        (empresa['EmpresaID'],)
    )
    
    return render_template('empresa/vacantes.html', vacantes=vacantes, empresa=empresa)

@app.route('/empresa/vacantes/nueva', methods=['GET', 'POST'])
@login_required
@role_required('empresa')
def empresa_nueva_vacante():
    empresa = get_empresa_actual()
    if not empresa:
        flash('Perfil de empresa no encontrado', 'error')
        return redirect(url_for('login'))


    todas_habilidades = execute_query("SELECT * FROM Habilidades ORDER BY Nombre")
    tipos_contrato = ['Tiempo completo', 'Medio tiempo', 'Por proyecto', 'Prácticas', 'Freelance']
    modalidades = ['Presencial', 'Remoto', 'Híbrido']
    niveles_experiencia = ['Sin experiencia', '1-3 años', '3-5 años', '5+ años']
    grados_estudios = ['Secundaria', 'Bachillerato', 'Licenciatura', 'Maestría', 'Doctorado']

    if request.method == 'POST':
        try:
    
            required_fields = {
                'puesto': 'Puesto',
                'grado_estudios': 'Grado de estudios',
                'resumen': 'Descripción del puesto',
                'plazas': 'Plazas disponibles',
                'tipo_contrato': 'Tipo de contrato',
                'modalidad': 'Modalidad de trabajo',
                'experiencia': 'Experiencia requerida'
            }
            
            for field, name in required_fields.items():
                if not request.form.get(field):
                    flash(f'El campo {name} es requerido', 'error')
                    return redirect(url_for('empresa_nueva_vacante'))

            execute_query(
                """INSERT INTO Vacantes 
                (EmpresaID, Puesto, GradoEstudios, Resumen, Plazas, PlazasDisponibles,
                Estatus, Salario, TipoContrato, Modalidad, Ubicacion, ExperienciaRequerida,
                Beneficios, FechaCierre)
                VALUES (?, ?, ?, ?, ?, ?, 'en_revision', ?, ?, ?, ?, ?, ?, ?)""",
                (
                    empresa['EmpresaID'],
                    request.form['puesto'],
                    request.form['grado_estudios'],
                    request.form['resumen'],
                    int(request.form['plazas']),
                    int(request.form['plazas']),
                    request.form.get('salario', ''),
                    request.form['tipo_contrato'],
                    request.form['modalidad'],
                    request.form.get('ubicacion', ''),
                    request.form['experiencia'],
                    request.form.get('beneficios', ''),
                    datetime.strptime(request.form['fecha_cierre'], '%Y-%m-%d').date() if request.form.get('fecha_cierre') else None
                ),
                fetch=False
            )
            

            nueva_vacante = execute_query(
                "SELECT VacanteID FROM Vacantes WHERE EmpresaID = ? ORDER BY FechaPublicacion DESC LIMIT 1",
                (empresa['EmpresaID'],)
            )
            vacante_id = nueva_vacante[0]['VacanteID']
            

            for habilidad_id in request.form.getlist('habilidades_requeridas'):
                execute_query(
                    "INSERT INTO VacanteHabilidadesRequeridas (VacanteID, HabilidadID) VALUES (?, ?)",
                    (vacante_id, int(habilidad_id)),
                    fetch=False
                )
            
            for habilidad_id in request.form.getlist('habilidades_opcionales'):
                execute_query(
                    "INSERT INTO VacanteHabilidadesOpcionales (VacanteID, HabilidadID) VALUES (?, ?)",
                    (vacante_id, int(habilidad_id)),
                    fetch=False
                )
            

            admins = execute_query("SELECT AdministradorID FROM Administradores")
            for admin in admins:
                execute_query(
                    "INSERT INTO VacantesRevision (AdministradorID, VacanteID) VALUES (?, ?)",
                    (admin['AdministradorID'], vacante_id),
                    fetch=False
                )
            
            flash('Vacante creada correctamente. Pendiente de aprobación.', 'success')
            return redirect(url_for('empresa_vacantes'))
        
        except ValueError as e:
            flash('Error en los datos proporcionados: ' + str(e), 'error')
            return redirect(url_for('empresa_nueva_vacante'))
        except Exception as e:
            current_app.logger.error(f"Error al crear vacante: {str(e)}")
            flash('Ocurrió un error al crear la vacante', 'error')
            return redirect(url_for('empresa_nueva_vacante'))
    
    return render_template('empresa/nueva_vacante.html',
                        todas_habilidades=todas_habilidades,
                        tipos_contrato=tipos_contrato,
                        modalidades=modalidades,
                        niveles_experiencia=niveles_experiencia,
                        grados_estudios=grados_estudios,
                        editar=False,
                        vacante=None)

@app.route('/empresa/vacantes/<int:vacante_id>', methods=['GET', 'POST'])
@login_required
@role_required('empresa')
def empresa_ver_vacante(vacante_id):
    empresa = get_empresa_actual()
    if not empresa:
        flash('Perfil de empresa no encontrado', 'error')
        return redirect(url_for('login'))
    

    vacante = execute_query(
        "SELECT * FROM Vacantes WHERE VacanteID = ? AND EmpresaID = ?",
        (vacante_id, empresa['EmpresaID'])
    )
    if not vacante:
        flash('Vacante no encontrada.', 'error')
        return redirect(url_for('empresa_vacantes'))
    
    vacante = vacante[0]

    habilidades_requeridas = execute_query(
        "SELECT h.* FROM VacanteHabilidadesRequeridas vh "
        "JOIN Habilidades h ON vh.HabilidadID = h.HabilidadID "
        "WHERE vh.VacanteID = ?",
        (vacante_id,)
    )
    
    habilidades_opcionales = execute_query(
        "SELECT h.* FROM VacanteHabilidadesOpcionales vh "
        "JOIN Habilidades h ON vh.HabilidadID = h.HabilidadID "
        "WHERE vh.VacanteID = ?",
        (vacante_id,)
    )
    

    postulaciones = execute_query(
        """SELECT p.*, c.*, 
        (SELECT COUNT(*) FROM CandidatoHabilidades ch 
         JOIN VacanteHabilidadesRequeridas vh ON ch.HabilidadID = vh.HabilidadID
         WHERE ch.CandidatoID = c.CandidatoID AND vh.VacanteID = ?) as HabilidadesCoincidentes
        FROM Postulaciones p
        JOIN Candidatos c ON p.CandidatoID = c.CandidatoID
        WHERE p.VacanteID = ?
        ORDER BY p.FechaPostulacion DESC""",
        (vacante_id, vacante_id)
    )
    

    num_habilidades_req = len(habilidades_requeridas)
    for post in postulaciones:
        if num_habilidades_req > 0:
            post['PorcentajeCoincidencia'] = round((post['HabilidadesCoincidentes'] / num_habilidades_req) * 100, 1)
        else:
            post['PorcentajeCoincidencia'] = 0
    
    return render_template('empresa/ver_vacante.html', 
                         vacante=vacante, 
                         postulaciones=postulaciones,
                         habilidades_requeridas=habilidades_requeridas,
                         habilidades_opcionales=habilidades_opcionales,
                         ahora=datetime.now())



@app.route('/empresa/aceptar_candidato/<int:vacante_id>/<int:candidato_id>')
@login_required
@role_required('empresa')
def empresa_aceptar_candidato(vacante_id, candidato_id):
    empresa = get_empresa_actual()
    if not empresa:
        flash('Perfil de empresa no encontrado', 'error')
        return redirect(url_for('login'))
    
    vacante = execute_query(
        "SELECT * FROM Vacantes WHERE VacanteID = ? AND EmpresaID = ?",
        (vacante_id, empresa['EmpresaID'])
    )
    if not vacante:
        flash('Vacante no encontrada.', 'error')
        return redirect(url_for('empresa_vacantes'))
    
    vacante = vacante[0]
    
    try:
        # Actualizar postulación
        execute_query(
            "UPDATE Postulaciones SET Estatus = 'aceptado' WHERE VacanteID = ? AND CandidatoID = ?",
            (vacante_id, candidato_id),
            fetch=False
        )
        
        # Actualizar plazas
        execute_query(
            "UPDATE Vacantes SET PlazasDisponibles = PlazasDisponibles - 1 WHERE VacanteID = ?",
            (vacante_id,),
            fetch=False
        )
        
        # Cerrar vacante si no hay más plazas
        if vacante['PlazasDisponibles'] <= 1:
            execute_query(
                "UPDATE Vacantes SET Estatus = 'cerrada' WHERE VacanteID = ?",
                (vacante_id,),
                fetch=False
            )
        
        # Crear conversación automáticamente
        conversacion_existente = execute_query(
            "SELECT 1 FROM Conversaciones WHERE VacanteID = ? AND CandidatoID = ?",
            (vacante_id, candidato_id)
        )
        
        if not conversacion_existente:
            execute_query(
                """INSERT INTO Conversaciones (VacanteID, CandidatoID, EmpresaID)
                   VALUES (?, ?, ?)""",
                (vacante_id, candidato_id, empresa['EmpresaID']),
                fetch=False
            )
            
            # Enviar mensaje de bienvenida automático
            conversacion = execute_query(
                "SELECT ConversacionID FROM Conversaciones WHERE VacanteID = ? AND CandidatoID = ?",
                (vacante_id, candidato_id)
            )
            
            if conversacion:
                mensaje_bienvenida = f"""¡Hola! 👋

Te escribo de {empresa['Nombre']}. Hemos revisado tu postulación para la vacante de {vacante['Puesto']} y ¡nos alegra informarte que has sido seleccionado/a para continuar con el proceso!

A través de este chat podremos coordinar los siguientes pasos:

1. 📝 Revisión de documentos
2. 📅 Programación de entrevista
3. 📋 Detalles adicionales sobre la posición

¿Qué te parece si comenzamos conversando sobre tus expectativas y disponibilidad?

Estamos muy entusiasmados con tu perfil y esperamos trabajar juntos. ✨

Saludos cordiales,
Equipo de {empresa['Nombre']}"""
                
                execute_query(
                    """INSERT INTO Mensajes (ConversacionID, RemitenteID, RemitenteTipo, Mensaje)
                       VALUES (?, ?, 'empresa', ?)""",
                    (conversacion[0]['ConversacionID'], empresa['EmpresaID'], mensaje_bienvenida),
                    fetch=False
                )
        
        # Notificación
        execute_query(
            """INSERT INTO Notificaciones 
            (EmpresaID, Mensaje, Tipo, VacanteID)
            VALUES (?, 'Tu postulación ha sido aceptada. Ya puedes chatear con la empresa.', 'postulacion', ?)""",
            (empresa['EmpresaID'], vacante_id),
            fetch=False
        )
        
        flash('Candidato aceptado correctamente. Se ha abierto un chat para seguimiento.', 'success')
        return redirect(url_for('ver_conversacion', vacante_id=vacante_id, candidato_id=candidato_id))
    
    except Exception as e:
        current_app.logger.error(f"Error al aceptar candidato: {str(e)}")
        flash('Ocurrió un error al aceptar al candidato', 'error')
        return redirect(url_for('empresa_ver_vacante', vacante_id=vacante_id))
    
@app.template_filter('nl2br')
def nl2br_filter(s):
    """Convierte saltos de línea en <br> tags"""
    if not s:
        return ''
    return s.replace('\n', '<br>\n')

@app.context_processor
def inject_global_variables():
    """Inyecta variables globales en todos los templates"""
    no_leidos = 0
    
    if 'user_id' in session:
        usuario_actual = get_usuario_actual()
        if usuario_actual:
            if usuario_actual['TipoUsuario'] == 'candidato':
                candidato = get_candidato_actual()
                if candidato:
                    try:
                        result = execute_query(
                            """SELECT COUNT(*) as Total FROM Mensajes m
                               JOIN Conversaciones c ON m.ConversacionID = c.ConversacionID
                               WHERE c.CandidatoID = ? AND m.RemitenteTipo = 'empresa' AND m.Leido = 0""",
                            (candidato['CandidatoID'],)
                        )
                        if result:
                            no_leidos = result[0]['Total']
                    except Exception:
                        pass
            elif usuario_actual['TipoUsuario'] == 'empresa':
                empresa = get_empresa_actual()
                if empresa:
                    try:
                        result = execute_query(
                            """SELECT COUNT(*) as Total FROM Mensajes m
                               JOIN Conversaciones c ON m.ConversacionID = c.ConversacionID
                               WHERE c.EmpresaID = ? AND m.RemitenteTipo = 'candidato' AND m.Leido = 0""",
                            (empresa['EmpresaID'],)
                        )
                        if result:
                            no_leidos = result[0]['Total']
                    except Exception:
                        pass
    
    return {
        'no_leidos': no_leidos,
        'current_year': datetime.now().year
    }
########


@app.route('/empresa/rechazar_candidato/<int:vacante_id>/<int:candidato_id>', methods=['GET', 'POST'])
@login_required
@role_required('empresa')
def empresa_rechazar_candidato(vacante_id, candidato_id):
    empresa = get_empresa_actual()
    if not empresa:
        flash('Perfil de empresa no encontrado', 'error')
        return redirect(url_for('login'))
    

    vacante = execute_query(
        "SELECT * FROM Vacantes WHERE VacanteID = ? AND EmpresaID = ?",
        (vacante_id, empresa['EmpresaID'])
    )
    if not vacante:
        flash('Vacante no encontrada.', 'error')
        return redirect(url_for('empresa_vacantes'))
    
    if request.method == 'POST':
        comentarios = request.form.get('comentarios', '').strip()
        if len(comentarios) < 10:
            flash('Por favor ingresa un motivo de rechazo (mínimo 10 caracteres)', 'warning')
            return redirect(url_for('empresa_rechazar_candidato', 
                                 vacante_id=vacante_id, 
                                 candidato_id=candidato_id))
        
        try:

            execute_query(
                "UPDATE Postulaciones SET Estatus = 'rechazado', Comentarios = ? WHERE VacanteID = ? AND CandidatoID = ?",
                (comentarios, vacante_id, candidato_id),
                fetch=False
            )
            
    
            execute_query(
                """INSERT INTO Notificaciones 
                (EmpresaID, Mensaje, Tipo, Comentarios, VacanteID)
                VALUES (?, 'Tu postulación ha sido rechazada', 'postulacion', ?, ?)""",
                (empresa['EmpresaID'], comentarios, vacante_id),
                fetch=False
            )
            
            flash('Candidato rechazado correctamente.', 'success')
            return redirect(url_for('empresa_ver_vacante', vacante_id=vacante_id))
        
        except Exception as e:
            current_app.logger.error(f"Error al rechazar candidato: {str(e)}")
            flash('Ocurrió un error al rechazar al candidato', 'error')
            return redirect(url_for('empresa_ver_vacante', vacante_id=vacante_id))
    
    return render_template('empresa/rechazar_candidato.html', 
                         vacante_id=vacante_id, 
                         candidato_id=candidato_id)

@app.route('/empresa/candidatos/<int:candidato_id>')
@login_required
@role_required('empresa')
def empresa_ver_candidato(candidato_id):
    empresa = get_empresa_actual()
    if not empresa:
        flash('Perfil de empresa no encontrado', 'error')
        return redirect(url_for('login'))

    candidato = execute_query(
        "SELECT * FROM Candidatos WHERE CandidatoID = ?",
        (candidato_id,)
    )
    
    if not candidato:
        flash('Candidato no encontrado', 'error')
        return redirect(url_for('empresa_vacantes'))
    
    candidato = candidato[0]
    

    nombres = [candidato['Nombre']]
    if candidato.get('ApellidoPaterno'):
        nombres.append(candidato['ApellidoPaterno'])
    if candidato.get('ApellidoMaterno'):
        nombres.append(candidato['ApellidoMaterno'])
    candidato['NombreCompleto'] = ' '.join(nombres)
    

    habilidades = execute_query(
        "SELECT h.* FROM CandidatoHabilidades ch "
        "JOIN Habilidades h ON ch.HabilidadID = h.HabilidadID "
        "WHERE ch.CandidatoID = ?",
        (candidato_id,)
    )
    

    experiencia = []
    try:
        experiencia = execute_query(
            "SELECT * FROM ExperienciaLaboral "
            "WHERE CandidatoID = ? "
            "ORDER BY CASE WHEN FechaSalida IS NULL THEN 0 ELSE 1 END, FechaSalida DESC",
            (candidato_id,)
        )
    except Exception as e:
        print(f"Error al obtener experiencia: {e}")
    

    educacion = []
    try:
        educacion = execute_query(
            "SELECT * FROM PreparacionAcademica "
            "WHERE CandidatoID = ? "
            "ORDER BY CASE WHEN FechaFin IS NULL THEN 0 ELSE 1 END, FechaFin DESC",
            (candidato_id,)
        )
    except Exception as e:
        print(f"Error al obtener educación: {e}")
    

    referencias = []
    try:
        referencias = execute_query(
            "SELECT * FROM Referencias "
            "WHERE CandidatoID = ?",
            (candidato_id,)
        )
    except Exception as e:
        print(f"Error al obtener referencias: {e}")
    

    cv_url = None
    if candidato.get('CV'):
        cv_url = url_for('static', filename='uploads/' + candidato['CV'])
    
    return render_template('empresa/ver_candidato.html',
                        candidato=candidato,
                        habilidades=habilidades,
                        experiencia=experiencia,
                        educacion=educacion,
                        referencias=referencias,
                        cv_url=cv_url)

@app.route('/empresa/notificaciones')
@login_required
@role_required('empresa')
def empresa_notificaciones():
    empresa = get_empresa_actual()
    if not empresa:
        flash('Perfil de empresa no encontrado', 'error')
        return redirect(url_for('login'))
    
   
    notificaciones = execute_query(
        "SELECT * FROM Notificaciones WHERE EmpresaID = ? ORDER BY Fecha DESC",
        (empresa['EmpresaID'],)
    )
    
  
    execute_query(
        "UPDATE Notificaciones SET Leida = 1 WHERE EmpresaID = ? AND Leida = 0",
        (empresa['EmpresaID'],),
        fetch=False
    )
    
    return render_template('empresa/notificaciones.html', notificaciones=notificaciones)

@app.route('/empresa/vacante/<int:vacante_id>/<nuevo_estado>')
@login_required
@role_required('empresa')
def empresa_cambiar_estado_vacante(vacante_id, nuevo_estado):
    empresa = get_empresa_actual()
    if not empresa:
        flash('Perfil de empresa no encontrado', 'error')
        return redirect(url_for('login'))
    
    
    vacante = execute_query(
        "SELECT * FROM Vacantes WHERE VacanteID = ? AND EmpresaID = ?",
        (vacante_id, empresa['EmpresaID'])
    )
    if not vacante:
        flash('Vacante no encontrada', 'error')
        return redirect(url_for('empresa_vacantes'))
    
    vacante = vacante[0]
    
    estados_validos = ['aprobada', 'cerrada']
    if nuevo_estado not in estados_validos:
        flash('Estado no válido', 'error')
        return redirect(url_for('empresa_ver_vacante', vacante_id=vacante_id))
    
    try:
 
        if (vacante['Estatus'] == 'cerrada' and nuevo_estado == 'aprobada') or \
           (vacante['Estatus'] == 'aprobada' and nuevo_estado == 'cerrada'):
            
            execute_query(
                "UPDATE Vacantes SET Estatus = ? WHERE VacanteID = ?",
                (nuevo_estado, vacante_id),
                fetch=False
            )
            
            flash(f'Estado de la vacante actualizado a "{nuevo_estado}"', 'success')
        else:
            flash('Transición de estado no permitida', 'error')
        
        return redirect(url_for('empresa_ver_vacante', vacante_id=vacante_id))
    
    except Exception as e:
        current_app.logger.error(f"Error al cambiar estado de vacante: {str(e)}")
        flash('Ocurrió un error al cambiar el estado', 'error')
        return redirect(url_for('empresa_ver_vacante', vacante_id=vacante_id))


def crear_admin_inicial():
    admin_email = os.getenv('INITIAL_ADMIN_EMAIL', '').strip().lower()
    admin_password = os.getenv('INITIAL_ADMIN_PASSWORD', '')
    if not admin_email or not admin_password:
        current_app.logger.info('Administrador inicial no configurado por variables de entorno.')
        return
    admin_existente = execute_query(
        "SELECT 1 FROM Usuarios WHERE LOWER(Email) = ?",
        (admin_email,)
    )
    
    if not admin_existente:
        execute_query(
            """INSERT INTO Usuarios (Email, PasswordHash, TipoUsuario, Activo)
            VALUES (?, ?, 'admin', TRUE)""",
            (admin_email, generate_password_hash(admin_password)),
            fetch=False
        )
        usuario = execute_query(
            "SELECT UsuarioID FROM Usuarios WHERE Email = ?",
            (admin_email,)
        )[0]
        execute_query(
            """INSERT INTO Administradores (AdministradorID, UsuarioID)
            VALUES ((SELECT COALESCE(MAX(AdministradorID), 0) + 1 FROM Administradores), ?)""",
            (usuario['UsuarioID'],),
            fetch=False
        )



@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():

    if 'user_id' in session and session.get('tipo') == 'admin':
        return redirect(url_for('admin_dashboard'))

    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')


        admins = execute_query(
            """SELECT UsuarioID, Email, PasswordHash, Activo FROM Usuarios
               WHERE LOWER(Email) = ? AND TipoUsuario = 'admin'""",
            ((email or '').strip().lower(),)
        )
        if admins and admins[0]['Activo'] and check_password_hash(admins[0]['PasswordHash'], password or ''):
            user_obj = User(
                id=admins[0]['UsuarioID'],
                email=admins[0]['Email'],
                tipo='admin',
                activo=admins[0]['Activo'],
            )
            login_user(user_obj, remember=True, duration=timedelta(days=30))
            session['email'] = admins[0]['Email']
            session['tipo'] = 'admin'
            session['user_id'] = admins[0]['UsuarioID']
            session.permanent = True
            flash('Bienvenido Administrador', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Credenciales administrativas incorrectas', 'error')


    return redirect(url_for('login'))


@app.post('/admin/passkeys/register/options')
@login_required
@role_required('admin')
def admin_passkey_register_options():
    rp_id, _ = webauthn_config()
    admin_id = session['user_id']
    admin_email = session.get('email', 'administrador')
    credentials = execute_query(
        "SELECT CredentialID FROM AdminPasskeys WHERE UsuarioID = ?",
        (admin_id,),
    )
    options = generate_registration_options(
        rp_id=rp_id,
        rp_name='TalentUPQ',
        user_id=str(admin_id).encode(),
        user_name=admin_email,
        user_display_name=admin_email,
        exclude_credentials=[
            PublicKeyCredentialDescriptor(id=base64url_to_bytes(row['CredentialID']))
            for row in credentials
        ],
        authenticator_selection=AuthenticatorSelectionCriteria(
            resident_key=ResidentKeyRequirement.PREFERRED,
            user_verification=UserVerificationRequirement.REQUIRED,
        ),
    )
    session['passkey_registration_challenge'] = bytes_to_base64url(options.challenge)
    return app.response_class(options_to_json(options), mimetype='application/json')


@app.post('/admin/passkeys/register/verify')
@login_required
@role_required('admin')
def admin_passkey_register_verify():
    admin_id = session['user_id']
    challenge = session.pop('passkey_registration_challenge', None)
    if not challenge:
        return jsonify({'error': 'El reto expiró. Intenta nuevamente.'}), 400
    try:
        rp_id, origin = webauthn_config()
        credential_payload = request.get_json(force=True)
        device_name = credential_payload.pop('deviceName', 'Dispositivo personal')
        verification = verify_registration_response(
            credential=credential_payload,
            expected_challenge=base64url_to_bytes(challenge),
            expected_rp_id=rp_id,
            expected_origin=origin,
            require_user_verification=True,
        )
        execute_query(
            """INSERT INTO AdminPasskeys
               (UsuarioID, CredentialID, PublicKey, SignCount, DeviceName)
               VALUES (?, ?, ?, ?, ?)""",
            (
                admin_id,
                bytes_to_base64url(verification.credential_id),
                bytes_to_base64url(verification.credential_public_key),
                verification.sign_count,
                str(device_name)[:100],
            ),
            fetch=False,
        )
        return jsonify({'ok': True, 'message': 'Huella, rostro o passkey registrada correctamente.'})
    except Exception as error:
        current_app.logger.warning('No fue posible registrar passkey: %s', error)
        return jsonify({'error': 'No fue posible registrar este dispositivo.'}), 400


@app.post('/admin/passkeys/login/options')
def admin_passkey_login_options():
    credentials = execute_query(
        """SELECT p.CredentialID FROM AdminPasskeys p
           JOIN Usuarios u ON u.UsuarioID = p.UsuarioID
           WHERE u.TipoUsuario = 'admin' AND u.Activo = TRUE"""
    )
    if not credentials:
        return jsonify({'error': 'Primero inicia con contraseña y registra este dispositivo.'}), 404
    rp_id, _ = webauthn_config()
    options = generate_authentication_options(
        rp_id=rp_id,
        allow_credentials=[
            PublicKeyCredentialDescriptor(id=base64url_to_bytes(row['CredentialID']))
            for row in credentials
        ],
        user_verification=UserVerificationRequirement.REQUIRED,
    )
    session['passkey_login_challenge'] = bytes_to_base64url(options.challenge)
    return app.response_class(options_to_json(options), mimetype='application/json')


@app.post('/admin/passkeys/login/verify')
def admin_passkey_login_verify():
    challenge = session.pop('passkey_login_challenge', None)
    payload = request.get_json(force=True)
    if not challenge:
        return jsonify({'error': 'El reto expiró. Intenta nuevamente.'}), 400
    rows = execute_query(
        """SELECT p.PasskeyID, p.UsuarioID, p.PublicKey, p.SignCount,
                  u.Email, u.Activo
           FROM AdminPasskeys p JOIN Usuarios u ON u.UsuarioID = p.UsuarioID
           WHERE p.CredentialID = ? AND u.TipoUsuario = 'admin'""",
        (payload.get('id', ''),),
    )
    if not rows or not rows[0]['Activo']:
        return jsonify({'error': 'Passkey no reconocida.'}), 401
    credential = rows[0]
    try:
        rp_id, origin = webauthn_config()
        verification = verify_authentication_response(
            credential=payload,
            expected_challenge=base64url_to_bytes(challenge),
            expected_rp_id=rp_id,
            expected_origin=origin,
            credential_public_key=base64url_to_bytes(credential['PublicKey']),
            credential_current_sign_count=credential['SignCount'],
            require_user_verification=True,
        )
        execute_query(
            "UPDATE AdminPasskeys SET SignCount = ?, LastUsedAt = CURRENT_TIMESTAMP WHERE PasskeyID = ?",
            (verification.new_sign_count, credential['PasskeyID']),
            fetch=False,
        )
        user_obj = User(credential['UsuarioID'], credential['Email'], 'admin', True)
        login_user(user_obj, remember=True, duration=timedelta(days=30))
        session.update(email=credential['Email'], tipo='admin', user_id=credential['UsuarioID'])
        session.permanent = True
        return jsonify({'ok': True, 'redirect': url_for('admin_dashboard')})
    except Exception as error:
        current_app.logger.warning('Passkey administrativa inválida: %s', error)
        return jsonify({'error': 'No se pudo validar la identidad en este dispositivo.'}), 401

# Rutas para administrador
@app.route('/admin')
@login_required
@role_required('admin')
def admin_dashboard():

    vacantes_pendientes = execute_query(
        """SELECT DISTINCT v.VacanteID as id, v.*, e.Nombre as empresa_nombre, 
           FORMAT(v.FechaPublicacion, 'dd/MM/yyyy') as fecha_formateada
           FROM Vacantes v
           JOIN Empresas e ON v.EmpresaID = e.EmpresaID
           WHERE v.Estatus = 'en_revision'
           ORDER BY v.FechaPublicacion DESC"""
    )
    
   
    if vacantes_pendientes:
        for vacante in vacantes_pendientes:
     
            habilidades_req = execute_query(
                """SELECT h.Nombre 
                   FROM VacanteHabilidadesRequeridas vh
                   JOIN Habilidades h ON vh.HabilidadID = h.HabilidadID
                   WHERE vh.VacanteID = ?""",
                (vacante['id'],)
            )
            vacante['habilidades_requeridas'] = [h['Nombre'] for h in habilidades_req]
            
    
            habilidades_opc = execute_query(
                """SELECT h.Nombre 
                   FROM VacanteHabilidadesOpcionales vh
                   JOIN Habilidades h ON vh.HabilidadID = h.HabilidadID
                   WHERE vh.VacanteID = ?""",
                (vacante['id'],)
            )
            vacante['habilidades_opcionales'] = [h['Nombre'] for h in habilidades_opc]
    

    estadisticas = {
        'total_empresas': execute_query("SELECT COUNT(*) as Count FROM Empresas")[0]['Count'],
        'total_candidatos': execute_query("SELECT COUNT(*) as Count FROM Candidatos")[0]['Count'],
        'vacantes_activas': execute_query("SELECT COUNT(*) as Count FROM Vacantes WHERE Estatus = 'aprobada'")[0]['Count'],
        'postulaciones': execute_query("SELECT COUNT(*) as Count FROM Postulaciones")[0]['Count']
    }
    
    return render_template('admin/dashboard.html', 
                         vacantes=vacantes_pendientes,
                         estadisticas=estadisticas)

@app.route('/admin/vacantes/<int:vacante_id>')
@login_required
@role_required('admin')
def admin_ver_vacante(vacante_id):

    vacante_result = execute_query(
        """SELECT 
            v.VacanteID,
            v.Puesto,
            v.GradoEstudios,
            v.Resumen,
            v.Plazas,
            v.PlazasDisponibles,
            v.Estatus,
            FORMAT(v.FechaPublicacion, 'dd/MM/yyyy') as FechaPublicacion,
            FORMAT(v.FechaAprobacion, 'dd/MM/yyyy') as FechaAprobacion,
            v.ComentariosAdmin,
            v.Salario,
            v.TipoContrato,
            v.Modalidad,
            v.Ubicacion,
            v.ExperienciaRequerida,
            v.Beneficios,
            FORMAT(v.FechaCierre, 'dd/MM/yyyy') as FechaCierre,
            e.Nombre as EmpresaNombre
           FROM Vacantes v
           JOIN Empresas e ON v.EmpresaID = e.EmpresaID
           WHERE v.VacanteID = ?""",
        (vacante_id,)
    )
    
    if not vacante_result:
        flash('Vacante no encontrada', 'error')
        return redirect(url_for('admin_dashboard'))
    

    vacante = vacante_result[0] if vacante_result else None
    

    habilidades_requeridas_result = execute_query(
        """SELECT h.Nombre 
           FROM VacanteHabilidadesRequeridas vh
           JOIN Habilidades h ON vh.HabilidadID = h.HabilidadID
           WHERE vh.VacanteID = ?""",
        (vacante_id,)
    )
    habilidades_requeridas = [h['Nombre'] for h in habilidades_requeridas_result] if habilidades_requeridas_result else []
    

    habilidades_opcionales_result = execute_query(
        """SELECT h.Nombre 
           FROM VacanteHabilidadesOpcionales vh
           JOIN Habilidades h ON vh.HabilidadID = h.HabilidadID
           WHERE vh.VacanteID = ?""",
        (vacante_id,)
    )
    habilidades_opcionales = [h['Nombre'] for h in habilidades_opcionales_result] if habilidades_opcionales_result else []
    
    return render_template('admin/ver_vacante.html',
                         vacante=vacante,
                         empresa={'nombre': vacante['EmpresaNombre']},
                         habilidades_requeridas=habilidades_requeridas,
                         habilidades_opcionales=habilidades_opcionales)

@app.route('/admin/aprobar_vacante/<int:vacante_id>', methods=['POST'])
@login_required
@role_required('admin')
def admin_aprobar_vacante(vacante_id):
    try:

        execute_query(
            """UPDATE Vacantes SET 
            Estatus = 'aprobada', 
            FechaAprobacion = GETDATE(),
            ComentariosAdmin = 'Aprobada por el administrador'
            WHERE VacanteID = ?""",
            (vacante_id,),
            fetch=False
        )
        

        vacante = execute_query(
            "SELECT EmpresaID, Puesto FROM Vacantes WHERE VacanteID = ?",
            (vacante_id,)
        )
        if vacante:
            execute_query(
                """INSERT INTO Notificaciones 
                (EmpresaID, Mensaje, Tipo, VacanteID)
                VALUES (?, 'Tu vacante ha sido aprobada', 'aprobacion', ?)""",
                (vacante[0]['EmpresaID'], vacante_id),
                fetch=False
            )
        
        flash('Vacante aprobada correctamente', 'success')
        return redirect(url_for('admin_dashboard'))
    
    except Exception as e:
        current_app.logger.error(f"Error al aprobar vacante: {str(e)}")
        flash('Error al aprobar la vacante', 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/rechazar_vacante/<int:vacante_id>', methods=['POST'])
@login_required
@role_required('admin')
def admin_rechazar_vacante(vacante_id):
    comentarios = request.form.get('comentarios', '').strip()
    if len(comentarios) < 10:
        flash('Por favor ingresa un motivo de rechazo (mínimo 10 caracteres)', 'warning')
        return redirect(url_for('admin_ver_vacante', vacante_id=vacante_id))
    
    try:

        execute_query(
            """UPDATE Vacantes SET 
            Estatus = 'rechazada', 
            ComentariosAdmin = ?
            WHERE VacanteID = ?""",
            (comentarios, vacante_id),
            fetch=False
        )
        

        vacante = execute_query(
            "SELECT EmpresaID, Puesto FROM Vacantes WHERE VacanteID = ?",
            (vacante_id,)
        )
        if vacante:
            execute_query(
                """INSERT INTO Notificaciones 
                (EmpresaID, Mensaje, Tipo, Comentarios, VacanteID)
                VALUES (?, 'Tu vacante ha sido rechazada', 'rechazo', ?, ?)""",
                (vacante[0]['EmpresaID'], comentarios, vacante_id),
                fetch=False
            )
        
        flash('Vacante rechazada correctamente', 'success')
        return redirect(url_for('admin_dashboard'))
    
    except Exception as e:
        current_app.logger.error(f"Error al rechazar vacante: {str(e)}")
        flash('Error al rechazar la vacante', 'error')
        return redirect(url_for('admin_dashboard'))


@app.route('/admin/panel-control')
@login_required
@role_required('admin')
def admin_panel_control():

    estadisticas = {
        'usuarios_count': execute_query("SELECT COUNT(*) as Count FROM Usuarios")[0]['Count'],
        'empresas_count': execute_query("SELECT COUNT(*) as Count FROM Empresas")[0]['Count'],
        'candidatos_count': execute_query("SELECT COUNT(*) as Count FROM Candidatos")[0]['Count'],
        'vacantes_count': execute_query("SELECT COUNT(*) as Count FROM Vacantes")[0]['Count'],
        'postulaciones_count': execute_query("SELECT COUNT(*) as Count FROM Postulaciones")[0]['Count'],
        'vacantes_pendientes': execute_query("SELECT COUNT(*) as Count FROM Vacantes WHERE Estatus = 'en_revision'")[0]['Count']
    }
    
    return render_template('admin/panel_control.html', **estadisticas)

@app.route('/admin/usuarios')
def admin_usuarios():

    admin_id = session.get('UsuarioID')  
    
    usuarios = execute_query(
    """SELECT 
        UsuarioID, 
        Email, 
        TipoUsuario, 
        FechaRegistro, 
        Activo
    FROM Usuarios 
    ORDER BY FechaRegistro ASC"""
)
    return render_template('admin/usuarios.html', usuarios=usuarios)




# ==================== REPORTES Y ESTADÍSTICAS ====================

@app.route('/admin/reportes')
@login_required
@role_required('admin')
def admin_reportes():
    """Panel de reportes con estadísticas y gráficas"""
    return render_template('admin/reportes.html')

@app.route('/admin/api/estadisticas')
@login_required
@role_required('admin')
def api_estadisticas():
    """API para obtener datos estadísticos con filtros"""
    try:
        # Obtener parámetros de filtro
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        tipo_usuario = request.args.get('tipo_usuario')
        estatus_vacante = request.args.get('estatus_vacante')
        empresa_id = request.args.get('empresa_id')
        
        # Construir filtros para usuarios
        usuario_filters = []
        usuario_params = []
        if fecha_inicio:
            usuario_filters.append("FechaRegistro >= ?")
            usuario_params.append(fecha_inicio)
        if fecha_fin:
            usuario_filters.append("FechaRegistro <= ?")
            usuario_params.append(fecha_fin)
        if tipo_usuario and tipo_usuario != 'todos':
            usuario_filters.append("TipoUsuario = ?")
            usuario_params.append(tipo_usuario)
        
        usuario_where = " AND ".join(usuario_filters) if usuario_filters else "1=1"
        
        # Estadísticas generales con filtros
        estadisticas = {
            'total_usuarios': execute_query(f"""
                SELECT COUNT(*) as Total FROM Usuarios WHERE {usuario_where}
            """, usuario_params)[0]['Total'],
            'total_empresas': execute_query("SELECT COUNT(*) as Total FROM Empresas")[0]['Total'],
            'total_candidatos': execute_query("SELECT COUNT(*) as Total FROM Candidatos")[0]['Total'],
        }
        
        # Filtros para vacantes
        vacante_filters = []
        vacante_params = []
        if fecha_inicio:
            vacante_filters.append("FechaPublicacion >= ?")
            vacante_params.append(fecha_inicio)
        if fecha_fin:
            vacante_filters.append("FechaPublicacion <= ?")
            vacante_params.append(fecha_fin)
        if estatus_vacante and estatus_vacante != 'todos':
            vacante_filters.append("Estatus = ?")
            vacante_params.append(estatus_vacante)
        if empresa_id and empresa_id != 'todas':
            vacante_filters.append("EmpresaID = ?")
            vacante_params.append(empresa_id)
        
        vacante_where = " AND ".join(vacante_filters) if vacante_filters else "1=1"
        
        estadisticas['total_vacantes'] = execute_query(f"""
            SELECT COUNT(*) as Total FROM Vacantes WHERE {vacante_where}
        """, vacante_params)[0]['Total']
        
        estadisticas['vacantes_aprobadas'] = execute_query(f"""
            SELECT COUNT(*) as Total FROM Vacantes 
            WHERE Estatus = 'aprobada' AND {vacante_where}
        """, vacante_params)[0]['Total']
        
        estadisticas['vacantes_pendientes'] = execute_query(f"""
            SELECT COUNT(*) as Total FROM Vacantes 
            WHERE Estatus = 'en_revision' AND {vacante_where}
        """, vacante_params)[0]['Total']
        
        estadisticas['vacantes_rechazadas'] = execute_query(f"""
            SELECT COUNT(*) as Total FROM Vacantes 
            WHERE Estatus = 'rechazada' AND {vacante_where}
        """, vacante_params)[0]['Total']
        
        estadisticas['vacantes_cerradas'] = execute_query(f"""
            SELECT COUNT(*) as Total FROM Vacantes 
            WHERE Estatus = 'cerrada' AND {vacante_where}
        """, vacante_params)[0]['Total']
        
        # Postulaciones con filtros
        postulacion_filters = []
        postulacion_params = []
        if fecha_inicio:
            postulacion_filters.append("p.FechaPostulacion >= ?")
            postulacion_params.append(fecha_inicio)
        if fecha_fin:
            postulacion_filters.append("p.FechaPostulacion <= ?")
            postulacion_params.append(fecha_fin)
        if estatus_vacante and estatus_vacante != 'todos':
            postulacion_filters.append("v.Estatus = ?")
            postulacion_params.append(estatus_vacante)
        
        postulacion_where = " AND ".join(postulacion_filters) if postulacion_filters else "1=1"
        
        estadisticas['total_postulaciones'] = execute_query(f"""
            SELECT COUNT(*) as Total 
            FROM Postulaciones p
            JOIN Vacantes v ON p.VacanteID = v.VacanteID
            WHERE {postulacion_where}
        """, postulacion_params)[0]['Total']
        
        # Registros por mes (últimos 12 meses) con filtros
        registros_mensuales = execute_query(f"""
            SELECT 
                FORMAT(FechaRegistro, 'yyyy-MM') as Mes,
                COUNT(*) as Total,
                SUM(CASE WHEN TipoUsuario = 'candidato' THEN 1 ELSE 0 END) as Candidatos,
                SUM(CASE WHEN TipoUsuario = 'empresa' THEN 1 ELSE 0 END) as Empresas
            FROM Usuarios
            WHERE FechaRegistro >= DATEADD(MONTH, -12, GETDATE())
            AND {usuario_where}
            GROUP BY FORMAT(FechaRegistro, 'yyyy-MM')
            ORDER BY Mes ASC
        """, usuario_params)
        
        # Postulaciones por mes con filtros
        postulaciones_mensuales = execute_query(f"""
            SELECT 
                FORMAT(p.FechaPostulacion, 'yyyy-MM') as Mes,
                COUNT(*) as Total,
                SUM(CASE WHEN p.Estatus = 'aceptado' THEN 1 ELSE 0 END) as Aceptadas,
                SUM(CASE WHEN p.Estatus = 'rechazado' THEN 1 ELSE 0 END) as Rechazadas,
                SUM(CASE WHEN p.Estatus = 'pendiente' THEN 1 ELSE 0 END) as Pendientes
            FROM Postulaciones p
            JOIN Vacantes v ON p.VacanteID = v.VacanteID
            WHERE p.FechaPostulacion >= DATEADD(MONTH, -12, GETDATE())
            AND {postulacion_where}
            GROUP BY FORMAT(p.FechaPostulacion, 'yyyy-MM')
            ORDER BY Mes ASC
        """, postulacion_params)
        
        # Vacantes por empresa (top 10) con filtros
        vacantes_por_empresa = execute_query(f"""
            SELECT
                e.Nombre as Empresa,
                COUNT(v.VacanteID) as Total
            FROM Empresas e
            LEFT JOIN Vacantes v ON e.EmpresaID = v.EmpresaID
            WHERE {vacante_where.replace('EmpresaID', 'e.EmpresaID')}
            GROUP BY e.Nombre
            ORDER BY Total DESC
            LIMIT 10
        """, vacante_params)
        
        # Habilidades más demandadas con filtros
        habilidades_demandadas = execute_query(f"""
            SELECT
                h.Nombre as Habilidad,
                COUNT(vh.VacanteID) as TotalVacantes
            FROM Habilidades h
            LEFT JOIN VacanteHabilidadesRequeridas vh ON h.HabilidadID = vh.HabilidadID
            LEFT JOIN Vacantes v ON vh.VacanteID = v.VacanteID
            WHERE {vacante_where.replace('Vacantes', 'v')}
            GROUP BY h.Nombre
            ORDER BY TotalVacantes DESC
            LIMIT 10
        """, vacante_params)
        
        # Estado de vacantes con filtros
        estado_vacantes = execute_query(f"""
            SELECT 
                Estatus,
                COUNT(*) as Total
            FROM Vacantes
            WHERE {vacante_where}
            GROUP BY Estatus
        """, vacante_params)
        
        # Lista de empresas para el filtro
        empresas = execute_query("""
            SELECT EmpresaID, Nombre FROM Empresas ORDER BY Nombre
        """)
        
        return jsonify({
            'success': True,
            'estadisticas': estadisticas,
            'registros_mensuales': registros_mensuales,
            'postulaciones_mensuales': postulaciones_mensuales,
            'vacantes_por_empresa': vacantes_por_empresa,
            'habilidades_demandadas': habilidades_demandadas,
            'estado_vacantes': estado_vacantes,
            'empresas': empresas
        })
        
    except Exception as e:
        current_app.logger.error(f"Error en api_estadisticas: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/api/empresas')
@login_required
@role_required('admin')
def api_empresas():
    """API para obtener lista de empresas"""
    empresas = execute_query("SELECT EmpresaID, Nombre FROM Empresas ORDER BY Nombre")
    return jsonify({'success': True, 'empresas': empresas})

@app.route('/admin/exportar_reporte/<tipo>')
@login_required
@role_required('admin')
def exportar_reporte(tipo):
    """Exportar reporte en PDF con filtros aplicados"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER
    import io
    
    # Obtener filtros de la URL
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    tipo_usuario = request.args.get('tipo_usuario')
    estatus_vacante = request.args.get('estatus_vacante')
    empresa_id = request.args.get('empresa_id')
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                           leftMargin=0.5*inch, rightMargin=0.5*inch,
                           topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    story = []
    
    # Estilos personalizados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=22,
        textColor=colors.HexColor('#1e293b'),
        alignment=TA_CENTER,
        spaceAfter=6,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#64748b'),
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    section_style = ParagraphStyle(
        'SectionStyle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2563eb'),
        spaceAfter=10,
        spaceBefore=15,
        fontName='Helvetica-Bold'
    )
    
    # Título del reporte
    nombre_tipo = {
        'usuarios': 'Usuarios',
        'vacantes': 'Vacantes',
        'postulaciones': 'Postulaciones',
        'completo': 'Completo'
    }.get(tipo, tipo.capitalize())
    
    story.append(Paragraph(f"Reporte de {nombre_tipo} - TalentUPQ", title_style))
    story.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style))
    
    # Mostrar filtros aplicados
    filtros_texto = []
    if fecha_inicio:
        filtros_texto.append(f"Desde: {fecha_inicio}")
    if fecha_fin:
        filtros_texto.append(f"Hasta: {fecha_fin}")
    if tipo_usuario and tipo_usuario != 'todos':
        filtros_texto.append(f"Tipo: {tipo_usuario}")
    if estatus_vacante and estatus_vacante != 'todos':
        filtros_texto.append(f"Estatus: {estatus_vacante}")
    if empresa_id and empresa_id != 'todas':
        empresa_nombre = execute_query(
            "SELECT Nombre FROM Empresas WHERE EmpresaID = ?", 
            [empresa_id]
        )
        if empresa_nombre:
            filtros_texto.append(f"Empresa: {empresa_nombre[0]['Nombre']}")
    
    if filtros_texto:
        story.append(Paragraph(f"Filtros aplicados: {', '.join(filtros_texto)}", subtitle_style))
    
    story.append(Spacer(1, 12))
    
    # Definir columnas según el tipo
    if tipo == 'usuarios':
        usuario_filters = []
        usuario_params = []
        if fecha_inicio:
            usuario_filters.append("FechaRegistro >= ?")
            usuario_params.append(fecha_inicio)
        if fecha_fin:
            usuario_filters.append("FechaRegistro <= ?")
            usuario_params.append(fecha_fin)
        if tipo_usuario and tipo_usuario != 'todos':
            usuario_filters.append("TipoUsuario = ?")
            usuario_params.append(tipo_usuario)
        
        usuario_where = " AND ".join(usuario_filters) if usuario_filters else "1=1"
        
        usuarios = execute_query(f"""
            SELECT Email, TipoUsuario, FechaRegistro, 
                   CASE WHEN Activo = 1 THEN 'Activo' ELSE 'Inactivo' END as Estatus
            FROM Usuarios 
            WHERE {usuario_where}
            ORDER BY FechaRegistro DESC
        """, usuario_params)
        
        data = [['Email', 'Tipo', 'Fecha Registro', 'Estatus']]
        for u in usuarios:
            data.append([
                u['Email'],
                u['TipoUsuario'],
                u['FechaRegistro'].strftime('%d/%m/%Y'),
                u['Estatus']
            ])
        
        total_activos = sum(1 for u in usuarios if u['Estatus'] == 'Activo')
        total_inactivos = len(usuarios) - total_activos
        
        story.append(Paragraph("Resumen de Usuarios", section_style))
        story.append(Paragraph(f"• Total: {len(usuarios)} usuarios", styles['Normal']))
        story.append(Paragraph(f"• Activos: {total_activos}", styles['Normal']))
        story.append(Paragraph(f"• Inactivos: {total_inactivos}", styles['Normal']))
        story.append(Spacer(1, 10))
        
    elif tipo == 'vacantes':
        vacante_filters = []
        vacante_params = []
        if fecha_inicio:
            vacante_filters.append("v.FechaPublicacion >= ?")
            vacante_params.append(fecha_inicio)
        if fecha_fin:
            vacante_filters.append("v.FechaPublicacion <= ?")
            vacante_params.append(fecha_fin)
        if estatus_vacante and estatus_vacante != 'todos':
            vacante_filters.append("v.Estatus = ?")
            vacante_params.append(estatus_vacante)
        if empresa_id and empresa_id != 'todas':
            vacante_filters.append("v.EmpresaID = ?")
            vacante_params.append(empresa_id)
        
        vacante_where = " AND ".join(vacante_filters) if vacante_filters else "1=1"
        
        vacantes = execute_query(f"""
            SELECT v.Puesto, e.Nombre as Empresa, v.Estatus, 
                   v.FechaPublicacion, v.PlazasDisponibles
            FROM Vacantes v
            JOIN Empresas e ON v.EmpresaID = e.EmpresaID
            WHERE {vacante_where}
            ORDER BY v.FechaPublicacion DESC
        """, vacante_params)
        
        data = [['Puesto', 'Empresa', 'Estatus', 'Fecha', 'Plazas']]
        for v in vacantes:
            data.append([
                v['Puesto'],
                v['Empresa'],
                v['Estatus'],
                v['FechaPublicacion'].strftime('%d/%m/%Y'),
                str(v['PlazasDisponibles'])
            ])
        
        story.append(Paragraph("Resumen de Vacantes", section_style))
        story.append(Paragraph(f"• Total: {len(vacantes)} vacantes", styles['Normal']))
        aprobadas = sum(1 for v in vacantes if v['Estatus'] == 'aprobada')
        pendientes = sum(1 for v in vacantes if v['Estatus'] == 'en_revision')
        rechazadas = sum(1 for v in vacantes if v['Estatus'] == 'rechazada')
        story.append(Paragraph(f"• Aprobadas: {aprobadas}", styles['Normal']))
        story.append(Paragraph(f"• En revisión: {pendientes}", styles['Normal']))
        story.append(Paragraph(f"• Rechazadas: {rechazadas}", styles['Normal']))
        story.append(Spacer(1, 10))
        
    elif tipo == 'postulaciones':
        postulacion_filters = []
        postulacion_params = []
        if fecha_inicio:
            postulacion_filters.append("p.FechaPostulacion >= ?")
            postulacion_params.append(fecha_inicio)
        if fecha_fin:
            postulacion_filters.append("p.FechaPostulacion <= ?")
            postulacion_params.append(fecha_fin)
        if estatus_vacante and estatus_vacante != 'todos':
            postulacion_filters.append("v.Estatus = ?")
            postulacion_params.append(estatus_vacante)
        
        postulacion_where = " AND ".join(postulacion_filters) if postulacion_filters else "1=1"
        
        postulaciones = execute_query(f"""
            SELECT v.Puesto, e.Nombre as Empresa, 
                   CONCAT_WS(' ', c.Nombre, c.ApellidoPaterno) as Candidato,
                   p.Estatus, p.FechaPostulacion
            FROM Postulaciones p
            JOIN Vacantes v ON p.VacanteID = v.VacanteID
            JOIN Empresas e ON v.EmpresaID = e.EmpresaID
            JOIN Candidatos c ON p.CandidatoID = c.CandidatoID
            WHERE {postulacion_where}
            ORDER BY p.FechaPostulacion DESC
        """, postulacion_params)
        
        data = [['Puesto', 'Empresa', 'Candidato', 'Estatus', 'Fecha']]
        for p in postulaciones:
            data.append([
                p['Puesto'],
                p['Empresa'],
                p['Candidato'],
                p['Estatus'],
                p['FechaPostulacion'].strftime('%d/%m/%Y')
            ])
        
        story.append(Paragraph("Resumen de Postulaciones", section_style))
        story.append(Paragraph(f"• Total: {len(postulaciones)} postulaciones", styles['Normal']))
        aceptadas = sum(1 for p in postulaciones if p['Estatus'] == 'aceptado')
        rechazadas = sum(1 for p in postulaciones if p['Estatus'] == 'rechazado')
        pendientes = sum(1 for p in postulaciones if p['Estatus'] == 'pendiente')
        story.append(Paragraph(f"• Aceptadas: {aceptadas}", styles['Normal']))
        story.append(Paragraph(f"• Rechazadas: {rechazadas}", styles['Normal']))
        story.append(Paragraph(f"• Pendientes: {pendientes}", styles['Normal']))
        story.append(Spacer(1, 10))
        
    else:  # completo
        stats = execute_query("""
            SELECT 
                (SELECT COUNT(*) FROM Usuarios) as TotalUsuarios,
                (SELECT COUNT(*) FROM Empresas) as TotalEmpresas,
                (SELECT COUNT(*) FROM Candidatos) as TotalCandidatos,
                (SELECT COUNT(*) FROM Vacantes) as TotalVacantes,
                (SELECT COUNT(*) FROM Postulaciones) as TotalPostulaciones
        """)[0]
        
        story.append(Paragraph("Resumen General del Sistema", section_style))
        story.append(Paragraph(f"• Total Usuarios: {stats['TotalUsuarios']}", styles['Normal']))
        story.append(Paragraph(f"• Total Empresas: {stats['TotalEmpresas']}", styles['Normal']))
        story.append(Paragraph(f"• Total Candidatos: {stats['TotalCandidatos']}", styles['Normal']))
        story.append(Paragraph(f"• Total Vacantes: {stats['TotalVacantes']}", styles['Normal']))
        story.append(Paragraph(f"• Total Postulaciones: {stats['TotalPostulaciones']}", styles['Normal']))
        story.append(Spacer(1, 10))
        
        data = [['Métrica', 'Valor']]
        data.append(['Usuarios', str(stats['TotalUsuarios'])])
        data.append(['Empresas', str(stats['TotalEmpresas'])])
        data.append(['Candidatos', str(stats['TotalCandidatos'])])
        data.append(['Vacantes', str(stats['TotalVacantes'])])
        data.append(['Postulaciones', str(stats['TotalPostulaciones'])])
    
    # Crear tabla
    if data:
        table = Table(data, colWidths=[1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(table)
    
    # Pie de página
    story.append(Spacer(1, 20))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#94a3b8'),
        alignment=TA_CENTER
    )
    story.append(Paragraph("Reporte generado por TalentUPQ - Sistema de gestión de talento", footer_style))
    
    doc.build(story)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'reporte_{tipo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf',
        mimetype='application/pdf'
    )

@app.route('/admin/exportar_excel/<tipo>')
@login_required
@role_required('admin')
def exportar_excel(tipo):
    """Exportar reporte en Excel con filtros aplicados"""
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Obtener filtros de la URL
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        tipo_usuario = request.args.get('tipo_usuario')
        estatus_vacante = request.args.get('estatus_vacante')
        empresa_id = request.args.get('empresa_id')
        
        # Crear libro de Excel
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Hoja de datos
        ws1 = wb.create_sheet("Datos")
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        cell_alignment = Alignment(horizontal="left", vertical="center")
        border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        if tipo == 'usuarios':
            usuario_filters = []
            usuario_params = []
            if fecha_inicio:
                usuario_filters.append("FechaRegistro >= ?")
                usuario_params.append(fecha_inicio)
            if fecha_fin:
                usuario_filters.append("FechaRegistro <= ?")
                usuario_params.append(fecha_fin)
            if tipo_usuario and tipo_usuario != 'todos':
                usuario_filters.append("TipoUsuario = ?")
                usuario_params.append(tipo_usuario)
            
            usuario_where = " AND ".join(usuario_filters) if usuario_filters else "1=1"
            
            usuarios = execute_query(f"""
                SELECT Email, TipoUsuario, FechaRegistro, 
                       CASE WHEN Activo = 1 THEN 'Activo' ELSE 'Inactivo' END as Estatus
                FROM Usuarios 
                WHERE {usuario_where}
                ORDER BY FechaRegistro DESC
            """, usuario_params)
            
            headers = ['Email', 'Tipo de Usuario', 'Fecha Registro', 'Estatus']
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            for row, usuario in enumerate(usuarios, 2):
                ws1.cell(row=row, column=1, value=usuario['Email']).alignment = cell_alignment
                ws1.cell(row=row, column=2, value=usuario['TipoUsuario']).alignment = cell_alignment
                ws1.cell(row=row, column=3, value=usuario['FechaRegistro'].strftime('%d/%m/%Y')).alignment = cell_alignment
                ws1.cell(row=row, column=4, value=usuario['Estatus']).alignment = cell_alignment
                for col in range(1, 5):
                    ws1.cell(row=row, column=col).border = border
            
            for col in range(1, 5):
                ws1.column_dimensions[get_column_letter(col)].width = 25
            
            # Hoja de resumen
            ws2 = wb.create_sheet("Resumen")
            ws2.cell(row=1, column=1, value="Métrica").font = Font(bold=True, size=12)
            ws2.cell(row=1, column=2, value="Valor").font = Font(bold=True, size=12)
            ws2.cell(row=2, column=1, value="Total Usuarios")
            ws2.cell(row=2, column=2, value=len(usuarios))
            activos = sum(1 for u in usuarios if u['Estatus'] == 'Activo')
            ws2.cell(row=3, column=1, value="Usuarios Activos")
            ws2.cell(row=3, column=2, value=activos)
            ws2.cell(row=4, column=1, value="Usuarios Inactivos")
            ws2.cell(row=4, column=2, value=len(usuarios) - activos)
            ws2.column_dimensions['A'].width = 25
            ws2.column_dimensions['B'].width = 15
            
        elif tipo == 'vacantes':
            vacante_filters = []
            vacante_params = []
            if fecha_inicio:
                vacante_filters.append("v.FechaPublicacion >= ?")
                vacante_params.append(fecha_inicio)
            if fecha_fin:
                vacante_filters.append("v.FechaPublicacion <= ?")
                vacante_params.append(fecha_fin)
            if estatus_vacante and estatus_vacante != 'todos':
                vacante_filters.append("v.Estatus = ?")
                vacante_params.append(estatus_vacante)
            if empresa_id and empresa_id != 'todas':
                vacante_filters.append("v.EmpresaID = ?")
                vacante_params.append(empresa_id)
            
            vacante_where = " AND ".join(vacante_filters) if vacante_filters else "1=1"
            
            vacantes = execute_query(f"""
                SELECT v.Puesto, e.Nombre as Empresa, v.Estatus, 
                       v.FechaPublicacion, v.PlazasDisponibles
                FROM Vacantes v
                JOIN Empresas e ON v.EmpresaID = e.EmpresaID
                WHERE {vacante_where}
                ORDER BY v.FechaPublicacion DESC
            """, vacante_params)
            
            headers = ['Puesto', 'Empresa', 'Estatus', 'Fecha Publicación', 'Plazas']
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            for row, vacante in enumerate(vacantes, 2):
                ws1.cell(row=row, column=1, value=vacante['Puesto']).alignment = cell_alignment
                ws1.cell(row=row, column=2, value=vacante['Empresa']).alignment = cell_alignment
                ws1.cell(row=row, column=3, value=vacante['Estatus']).alignment = cell_alignment
                ws1.cell(row=row, column=4, value=vacante['FechaPublicacion'].strftime('%d/%m/%Y')).alignment = cell_alignment
                ws1.cell(row=row, column=5, value=vacante['PlazasDisponibles']).alignment = cell_alignment
                for col in range(1, 6):
                    ws1.cell(row=row, column=col).border = border
            
            ws1.column_dimensions['A'].width = 25
            ws1.column_dimensions['B'].width = 25
            ws1.column_dimensions['C'].width = 18
            ws1.column_dimensions['D'].width = 20
            ws1.column_dimensions['E'].width = 12
            
            ws2 = wb.create_sheet("Resumen")
            ws2.cell(row=1, column=1, value="Métrica").font = Font(bold=True, size=12)
            ws2.cell(row=1, column=2, value="Valor").font = Font(bold=True, size=12)
            ws2.cell(row=2, column=1, value="Total Vacantes")
            ws2.cell(row=2, column=2, value=len(vacantes))
            aprobadas = sum(1 for v in vacantes if v['Estatus'] == 'aprobada')
            pendientes = sum(1 for v in vacantes if v['Estatus'] == 'en_revision')
            rechazadas = sum(1 for v in vacantes if v['Estatus'] == 'rechazada')
            ws2.cell(row=3, column=1, value="Aprobadas")
            ws2.cell(row=3, column=2, value=aprobadas)
            ws2.cell(row=4, column=1, value="En Revisión")
            ws2.cell(row=4, column=2, value=pendientes)
            ws2.cell(row=5, column=1, value="Rechazadas")
            ws2.cell(row=5, column=2, value=rechazadas)
            ws2.column_dimensions['A'].width = 25
            ws2.column_dimensions['B'].width = 15
            
        elif tipo == 'postulaciones':
            postulacion_filters = []
            postulacion_params = []
            if fecha_inicio:
                postulacion_filters.append("p.FechaPostulacion >= ?")
                postulacion_params.append(fecha_inicio)
            if fecha_fin:
                postulacion_filters.append("p.FechaPostulacion <= ?")
                postulacion_params.append(fecha_fin)
            if estatus_vacante and estatus_vacante != 'todos':
                postulacion_filters.append("v.Estatus = ?")
                postulacion_params.append(estatus_vacante)
            
            postulacion_where = " AND ".join(postulacion_filters) if postulacion_filters else "1=1"
            
            postulaciones = execute_query(f"""
                SELECT v.Puesto, e.Nombre as Empresa, 
                       CONCAT_WS(' ', c.Nombre, c.ApellidoPaterno) as Candidato,
                       p.Estatus, p.FechaPostulacion
                FROM Postulaciones p
                JOIN Vacantes v ON p.VacanteID = v.VacanteID
                JOIN Empresas e ON v.EmpresaID = e.EmpresaID
                JOIN Candidatos c ON p.CandidatoID = c.CandidatoID
                WHERE {postulacion_where}
                ORDER BY p.FechaPostulacion DESC
            """, postulacion_params)
            
            headers = ['Puesto', 'Empresa', 'Candidato', 'Estatus', 'Fecha Postulación']
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            for row, postulacion in enumerate(postulaciones, 2):
                ws1.cell(row=row, column=1, value=postulacion['Puesto']).alignment = cell_alignment
                ws1.cell(row=row, column=2, value=postulacion['Empresa']).alignment = cell_alignment
                ws1.cell(row=row, column=3, value=postulacion['Candidato']).alignment = cell_alignment
                ws1.cell(row=row, column=4, value=postulacion['Estatus']).alignment = cell_alignment
                ws1.cell(row=row, column=5, value=postulacion['FechaPostulacion'].strftime('%d/%m/%Y')).alignment = cell_alignment
                for col in range(1, 6):
                    ws1.cell(row=row, column=col).border = border
            
            ws1.column_dimensions['A'].width = 25
            ws1.column_dimensions['B'].width = 25
            ws1.column_dimensions['C'].width = 30
            ws1.column_dimensions['D'].width = 18
            ws1.column_dimensions['E'].width = 20
            
            ws2 = wb.create_sheet("Resumen")
            ws2.cell(row=1, column=1, value="Métrica").font = Font(bold=True, size=12)
            ws2.cell(row=1, column=2, value="Valor").font = Font(bold=True, size=12)
            ws2.cell(row=2, column=1, value="Total Postulaciones")
            ws2.cell(row=2, column=2, value=len(postulaciones))
            aceptadas = sum(1 for p in postulaciones if p['Estatus'] == 'aceptado')
            rechazadas = sum(1 for p in postulaciones if p['Estatus'] == 'rechazado')
            pendientes = sum(1 for p in postulaciones if p['Estatus'] == 'pendiente')
            ws2.cell(row=3, column=1, value="Aceptadas")
            ws2.cell(row=3, column=2, value=aceptadas)
            ws2.cell(row=4, column=1, value="Rechazadas")
            ws2.cell(row=4, column=2, value=rechazadas)
            ws2.cell(row=5, column=1, value="Pendientes")
            ws2.cell(row=5, column=2, value=pendientes)
            ws2.column_dimensions['A'].width = 25
            ws2.column_dimensions['B'].width = 15
        
        else:  # completo
            stats = execute_query("""
                SELECT 
                    (SELECT COUNT(*) FROM Usuarios) as TotalUsuarios,
                    (SELECT COUNT(*) FROM Empresas) as TotalEmpresas,
                    (SELECT COUNT(*) FROM Candidatos) as TotalCandidatos,
                    (SELECT COUNT(*) FROM Vacantes) as TotalVacantes,
                    (SELECT COUNT(*) FROM Postulaciones) as TotalPostulaciones
            """)[0]
            
            headers = ['Métrica', 'Valor']
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            metricas = [
                ('Total Usuarios', stats['TotalUsuarios']),
                ('Total Empresas', stats['TotalEmpresas']),
                ('Total Candidatos', stats['TotalCandidatos']),
                ('Total Vacantes', stats['TotalVacantes']),
                ('Total Postulaciones', stats['TotalPostulaciones'])
            ]
            
            for row, (metrica, valor) in enumerate(metricas, 2):
                ws1.cell(row=row, column=1, value=metrica).alignment = cell_alignment
                ws1.cell(row=row, column=2, value=valor).alignment = cell_alignment
                ws1.cell(row=row, column=1).border = border
                ws1.cell(row=row, column=2).border = border
            
            ws1.column_dimensions['A'].width = 30
            ws1.column_dimensions['B'].width = 20
        
        # Guardar el archivo
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'reporte_{tipo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except ImportError as e:
        current_app.logger.error(f"Error de importación: {str(e)}")
        return jsonify({
            'success': False, 
            'error': 'La librería openpyxl no está instalada. Ejecute: pip install openpyxl'
        })
    except Exception as e:
        current_app.logger.error(f"Error en exportar_excel: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})





@app.route('/admin/usuarios/crear', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def crear_usuario():
    if request.method == 'POST':
        try:
            email = request.form['email']
            password = request.form['password']
            tipo_usuario = request.form['tipo_usuario']
            
            if not email or not password:
                flash('Email y contraseña son obligatorios', 'error')
                return redirect(url_for('crear_usuario'))
            
       
            existe = execute_query(
                "SELECT 1 FROM Usuarios WHERE Email = ?",
                (email,)
            )
            if existe:
                flash('El email ya está registrado', 'error')
                return redirect(url_for('crear_usuario'))
            

            password_hash = generate_password_hash(password)
            
        
            execute_query(
                "INSERT INTO Usuarios (Email, PasswordHash, TipoUsuario) VALUES (?, ?, ?)",
                (email, password_hash, tipo_usuario),
                fetch=False
            )
            
            flash('Usuario creado exitosamente', 'success')
            return redirect(url_for('admin_usuarios'))
        
        except Exception as e:
            print(f"Error al crear usuario: {str(e)}")
            flash('Error al crear el usuario', 'error')
            return redirect(url_for('crear_usuario'))
    
    return render_template('admin/crear_usuario.html')


@app.route('/admin/usuarios/editar/<int:usuario_id>', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def editar_usuario(usuario_id):

    admin_id = session.get('UsuarioID')
    

    if usuario_id == admin_id:
        flash('No puedes editar tu propio usuario aquí', 'error')
        return redirect(url_for('admin_usuarios'))
    

    if request.method == 'POST':
        try:
            email = request.form['email']
            tipo_usuario = request.form['tipo_usuario']
            activo = request.form.get('activo', '0') == '1'
            
            if not email:
                flash('Email es obligatorio', 'error')
                return redirect(url_for('editar_usuario', usuario_id=usuario_id))
            
           
            existe = execute_query(
                "SELECT 1 FROM Usuarios WHERE Email = ? AND UsuarioID != ?",
                (email, usuario_id)
            )
            if existe:
                flash('El email ya está registrado por otro usuario', 'error')
                return redirect(url_for('editar_usuario', usuario_id=usuario_id))
            

            execute_query(
                """UPDATE Usuarios 
                   SET Email = ?, TipoUsuario = ?, Activo = ?
                   WHERE UsuarioID = ?""",
                (email, tipo_usuario, activo, usuario_id),
                fetch=False
            )
            
            flash('Usuario actualizado exitosamente', 'success')
            return redirect(url_for('admin_usuarios'))
        
        except Exception as e:
            print(f"Error al actualizar usuario: {str(e)}")
            flash('Error al actualizar el usuario', 'error')
            return redirect(url_for('editar_usuario', usuario_id=usuario_id))
    

    usuario = execute_query(
        """SELECT 
            UsuarioID, 
            Email, 
            TipoUsuario, 
            Activo,
            'Sin perfil' as TipoPerfil  
        FROM Usuarios 
        WHERE UsuarioID = ?""",
        (usuario_id,)
    )
    
    if not usuario:
        flash('Usuario no encontrado', 'error')
        return redirect(url_for('admin_usuarios'))
    
    return render_template('admin/editar_usuario.html', usuario=usuario[0])

@app.route('/admin/usuarios/eliminar/<int:usuario_id>', methods=['POST'])
@login_required
@role_required('admin')
def eliminar_usuario(usuario_id):
    
    admin_id = session.get('UsuarioID')
    

    if usuario_id == admin_id:
        flash('No puedes eliminarte a ti mismo.', 'error')
        return redirect(url_for('admin_usuarios'))
    
    try:

        usuario = execute_query(
            "SELECT 1 FROM Usuarios WHERE UsuarioID = ?",
            (usuario_id,)
        )
        
        if not usuario:
            flash('Usuario no encontrado.', 'error')
            return redirect(url_for('admin_usuarios'))
        
        execute_query(
            "DELETE FROM Usuarios WHERE UsuarioID = ?",
            (usuario_id,),
            fetch=False
        )
        
        flash('Usuario eliminado correctamente.', 'success')
        return redirect(url_for('admin_usuarios'))
    
    except Exception as e:
        print(f"Error al eliminar usuario: {str(e)}")  # Debug
        flash('Ocurrió un error al eliminar el usuario', 'error')
        return redirect(url_for('admin_usuarios'))


#CRUD EMPRESAS:

@app.route('/admin/empresas')
def admin_empresas():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM Empresas ORDER BY FechaRegistro ASC")
    empresas = cursor.fetchall()
    conn.close()
    
 
    empresas_list = []
    columns = [column[0] for column in cursor.description]
    for empresa in empresas:
        empresas_list.append(CaseInsensitiveDict(zip(columns, empresa)))
    
    return render_template('admin/empresas.html', empresas=empresas_list)

@app.route('/admin/empresas/crear', methods=['GET', 'POST'])
def crear_empresa():
    if request.method == 'POST':
        usuario_id = request.form['usuario_id']
        nombre = request.form['nombre']
        direccion = request.form.get('direccion', '')
        telefono = request.form.get('telefono', '')
        sitio_web = request.form.get('sitio_web', '')
        descripcion = request.form.get('descripcion', '')
        

        logo = None
        if 'logo' in request.files:
            file = request.files['logo']
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                logo = filename
        

        if sitio_web and not sitio_web.startswith(('http://', 'https://')):
            sitio_web = f'https://{sitio_web}'
        
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            cursor.execute("SELECT ISNULL(MAX(EmpresaID), 0) + 1 FROM Empresas")
            empresa_id = cursor.fetchone()[0]
            
            cursor.execute("""
                INSERT INTO Empresas (EmpresaID, UsuarioID, Nombre, Direccion, Telefono, SitioWeb, Descripcion, Logo)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (empresa_id, usuario_id, nombre, direccion, telefono, sitio_web, descripcion, logo))
            
            conn.commit()
            conn.close()
            
            flash('Empresa creada exitosamente', 'success')
            return redirect(url_for('admin_empresas'))
        except Exception as e:
            flash(f'Error al crear la empresa: {str(e)}', 'error')
    
    return render_template('admin/crear_empresa.html')

@app.route('/admin/empresas/editar/<int:empresa_id>', methods=['GET', 'POST'])
def editar_empresa(empresa_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if request.method == 'POST':
        usuario_id = request.form['usuario_id']
        nombre = request.form['nombre']
        direccion = request.form.get('direccion', '')
        telefono = request.form.get('telefono', '')
        sitio_web = request.form.get('sitio_web', '')
        descripcion = request.form.get('descripcion', '')
        
 
        cursor.execute("SELECT Logo FROM Empresas WHERE EmpresaID = ?", (empresa_id,))
        current_logo = cursor.fetchone()[0]
        logo = current_logo
        

        if 'logo' in request.files:
            file = request.files['logo']
            if file and allowed_file(file.filename):
         
                if current_logo and os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], current_logo)):
                    os.remove(os.path.join(app.config['UPLOAD_FOLDER'], current_logo))
                
        
                filename = secure_filename(file.filename)
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                logo = filename
        
        if sitio_web and not sitio_web.startswith(('http://', 'https://')):
            sitio_web = f'https://{sitio_web}'
        
        try:
            cursor.execute("""
                UPDATE Empresas 
                SET UsuarioID = ?, Nombre = ?, Direccion = ?, Telefono = ?, 
                    SitioWeb = ?, Descripcion = ?, Logo = ?
                WHERE EmpresaID = ?
            """, (usuario_id, nombre, direccion, telefono, sitio_web, descripcion, logo, empresa_id))
            
            conn.commit()
            conn.close()
            
            flash('Empresa actualizada exitosamente', 'success')
            return redirect(url_for('admin_empresas'))
        except Exception as e:
            flash(f'Error al actualizar la empresa: {str(e)}', 'error')
    
    cursor.execute("SELECT * FROM Empresas WHERE EmpresaID = ?", (empresa_id,))
    empresa = cursor.fetchone()
    conn.close()
    
    if not empresa:
        flash('Empresa no encontrada', 'error')
        return redirect(url_for('admin_empresas'))
    
    columns = [column[0] for column in cursor.description]
    empresa_dict = CaseInsensitiveDict(zip(columns, empresa))
    
    return render_template('admin/editar_empresa.html', empresa=empresa_dict)

@app.route('/admin/empresas/eliminar/<int:empresa_id>', methods=['POST'])
def eliminar_empresa(empresa_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        
        cursor.execute("SELECT Logo FROM Empresas WHERE EmpresaID = ?", (empresa_id,))
        logo = cursor.fetchone()[0]
        
        if logo and os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], logo)):
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], logo))
        

        cursor.execute("DELETE FROM Empresas WHERE EmpresaID = ?", (empresa_id,))
        conn.commit()
        conn.close()
        
        flash('Empresa eliminada exitosamente', 'success')
    except Exception as e:
        flash(f'Error al eliminar la empresa: {str(e)}', 'error')
    
    return redirect(url_for('admin_empresas'))


#CRUD CANDIDATOS:
@app.route('/admin/candidatos')
def admin_candidatos():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT CandidatoID, UsuarioID, Nombre, ApellidoPaterno, ApellidoMaterno, 
               Telefono, PuestoActual, PuestoSolicitado, FotoPerfil
        FROM Candidatos 
        ORDER BY Nombre, ApellidoPaterno
    """)
    candidatos = cursor.fetchall()
    conn.close()
    
    candidatos_list = []
    columns = [column[0] for column in cursor.description]
    for candidato in candidatos:
        candidatos_list.append(CaseInsensitiveDict(zip(columns, candidato)))
    
    return render_template('admin/candidatos.html', candidatos=candidatos_list)

@app.route('/admin/candidatos/crear', methods=['GET', 'POST'])
def crear_candidato():
    if request.method == 'POST':
      
        usuario_id = request.form['usuario_id']
        nombre = request.form['nombre']
        apellido_paterno = request.form['apellido_paterno']
        apellido_materno = request.form.get('apellido_materno', '')
        telefono = request.form.get('telefono', '')
        estado_civil = request.form.get('estado_civil', '')
        sexo = request.form.get('sexo', '')
        fecha_nacimiento = request.form.get('fecha_nacimiento', '')
        nacionalidad = request.form.get('nacionalidad', '')
        rfc = request.form.get('rfc', '')
        direccion = request.form.get('direccion', '')
        reubicacion = request.form.get('reubicacion') == 'on'
        viajar = request.form.get('viajar') == 'on'
        licencia = request.form.get('licencia') == 'on'
        modalidad_trabajo = request.form.get('modalidad_trabajo', '')
        puesto_actual = request.form.get('puesto_actual', '')
        puesto_solicitado = request.form.get('puesto_solicitado', '')
        resumen = request.form.get('resumen_profesional', '')


        foto_perfil = None
        cv = None
        
        if 'foto_perfil' in request.files:
            file = request.files['foto_perfil']
            if file and allowed_file(file.filename) and file.filename != '':
                filename = secure_filename(file.filename)
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                foto_perfil = filename
        
        if 'cv' in request.files:
            file = request.files['cv']
            if file and allowed_file(file.filename) and file.filename != '':
                filename = secure_filename(file.filename)
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                cv = filename

        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            

            cursor.execute("SELECT ISNULL(MAX(CandidatoID), 0) + 1 FROM Candidatos")
            candidato_id = cursor.fetchone()[0]
            
   
            fecha_nac = datetime.strptime(fecha_nacimiento, '%Y-%m-%d').date() if fecha_nacimiento else None
            
            cursor.execute("""
                INSERT INTO Candidatos (
                    CandidatoID, UsuarioID, Nombre, ApellidoPaterno, ApellidoMaterno,
                    Telefono, EstadoCivil, Sexo, FechaNacimiento, Nacionalidad, RFC,
                    Direccion, Reubicacion, Viajar, LicenciaConducir, ModalidadTrabajo,
                    PuestoActual, PuestoSolicitado, FotoPerfil, CV, ResumenProfesional
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                candidato_id, usuario_id, nombre, apellido_paterno, apellido_materno,
                telefono, estado_civil, sexo, fecha_nac, nacionalidad, rfc,
                direccion, reubicacion, viajar, licencia, modalidad_trabajo,
                puesto_actual, puesto_solicitado, foto_perfil, cv, resumen
            ))
            
            conn.commit()
            conn.close()
            
            flash('Candidato creado exitosamente', 'success')
            return redirect(url_for('admin_candidatos'))
        except Exception as e:
            flash(f'Error al crear el candidato: {str(e)}', 'error')
    
    return render_template('admin/crear_candidato.html')

@app.route('/admin/candidatos/editar/<int:candidato_id>', methods=['GET', 'POST'])
def editar_candidato(candidato_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if request.method == 'POST':

        usuario_id = request.form['usuario_id']
        nombre = request.form['nombre']
        apellido_paterno = request.form['apellido_paterno']
        apellido_materno = request.form.get('apellido_materno', '')
        telefono = request.form.get('telefono', '')
        estado_civil = request.form.get('estado_civil', '')
        sexo = request.form.get('sexo', '')
        fecha_nacimiento = request.form.get('fecha_nacimiento', '')
        nacionalidad = request.form.get('nacionalidad', '')
        rfc = request.form.get('rfc', '')
        direccion = request.form.get('direccion', '')
        reubicacion = request.form.get('reubicacion') == 'on'
        viajar = request.form.get('viajar') == 'on'
        licencia = request.form.get('licencia') == 'on'
        modalidad_trabajo = request.form.get('modalidad_trabajo', '')
        puesto_actual = request.form.get('puesto_actual', '')
        puesto_solicitado = request.form.get('puesto_solicitado', '')
        resumen = request.form.get('resumen_profesional', '')


        cursor.execute("SELECT FotoPerfil, CV FROM Candidatos WHERE CandidatoID = ?", (candidato_id,))
        current_files = cursor.fetchone()
        current_foto = current_files[0]
        current_cv = current_files[1]
        
        foto_perfil = current_foto
        cv = current_cv
        

        if 'foto_perfil' in request.files:
            file = request.files['foto_perfil']
            if file and allowed_file(file.filename) and file.filename != '':

                if current_foto and os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], current_foto)):
                    os.remove(os.path.join(app.config['UPLOAD_FOLDER'], current_foto))
                
                filename = secure_filename(file.filename)
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                foto_perfil = filename
        
        if 'cv' in request.files:
            file = request.files['cv']
            if file and allowed_file(file.filename) and file.filename != '':
    
                if current_cv and os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], current_cv)):
                    os.remove(os.path.join(app.config['UPLOAD_FOLDER'], current_cv))
                
                filename = secure_filename(file.filename)
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                cv = filename

        try:
            fecha_nac = datetime.strptime(fecha_nacimiento, '%Y-%m-%d').date() if fecha_nacimiento else None
            
            cursor.execute("""
                UPDATE Candidatos SET
                    UsuarioID = ?, Nombre = ?, ApellidoPaterno = ?, ApellidoMaterno = ?,
                    Telefono = ?, EstadoCivil = ?, Sexo = ?, FechaNacimiento = ?, Nacionalidad = ?, RFC = ?,
                    Direccion = ?, Reubicacion = ?, Viajar = ?, LicenciaConducir = ?, ModalidadTrabajo = ?,
                    PuestoActual = ?, PuestoSolicitado = ?, FotoPerfil = ?, CV = ?, ResumenProfesional = ?
                WHERE CandidatoID = ?
            """, (
                usuario_id, nombre, apellido_paterno, apellido_materno,
                telefono, estado_civil, sexo, fecha_nac, nacionalidad, rfc,
                direccion, reubicacion, viajar, licencia, modalidad_trabajo,
                puesto_actual, puesto_solicitado, foto_perfil, cv, resumen,
                candidato_id
            ))
            
            conn.commit()
            conn.close()
            
            flash('Candidato actualizado exitosamente', 'success')
            return redirect(url_for('admin_candidatos'))
        except Exception as e:
            flash(f'Error al actualizar el candidato: {str(e)}', 'error')
    

    cursor.execute("SELECT * FROM Candidatos WHERE CandidatoID = ?", (candidato_id,))
    candidato = cursor.fetchone()
    conn.close()
    
    if not candidato:
        flash('Candidato no encontrado', 'error')
        return redirect(url_for('admin_candidatos'))
    
   
    columns = [column[0] for column in cursor.description]
    candidato_dict = CaseInsensitiveDict(zip(columns, candidato))
    

    if candidato_dict['FechaNacimiento']:
        candidato_dict['FechaNacimiento'] = candidato_dict['FechaNacimiento'].strftime('%Y-%m-%d')
    
    return render_template('admin/editar_candidato.html', candidato=candidato_dict)

@app.route('/admin/candidatos/eliminar/<int:candidato_id>', methods=['POST'])
def eliminar_candidato(candidato_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:

        cursor.execute("SELECT FotoPerfil, CV FROM Candidatos WHERE CandidatoID = ?", (candidato_id,))
        files = cursor.fetchone()
        

        if files[0] and os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], files[0])):
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], files[0]))
        if files[1] and os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], files[1])):
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], files[1]))
        
       
        cursor.execute("DELETE FROM Candidatos WHERE CandidatoID = ?", (candidato_id,))
        conn.commit()
        conn.close()
        
        flash('Candidato eliminado exitosamente', 'success')
    except Exception as e:
        flash(f'Error al eliminar el candidato: {str(e)}', 'error')
    return redirect(url_for('admin_candidatos'))

#CRUD VACANTES

@app.route('/admin/vacantes')
def admin_vacantes():
    conn = get_db_connection()
    cursor = conn.cursor()
    
   
    cursor.execute("""
        SELECT v.*, e.Nombre as EmpresaNombre 
        FROM Vacantes v
        JOIN Empresas e ON v.EmpresaID = e.EmpresaID
        ORDER BY v.FechaPublicacion DESC
    """)
    vacantes = cursor.fetchall()
    

    columns = [column[0] for column in cursor.description]
    vacantes = [CaseInsensitiveDict(zip(columns, vacante)) for vacante in vacantes]
    
    conn.close()
    return render_template('admin/vacantes.html', vacantes=vacantes)

@app.route('/admin/vacantes/crear', methods=['GET', 'POST'])
def crear_vacante():
    if request.method == 'POST':

        empresa_id = request.form['empresa_id']
        puesto = request.form['puesto']
        grado_estudios = request.form['grado_estudios']
        resumen = request.form['resumen']
        plazas = request.form['plazas']
        plazas_disponibles = request.form['plazas_disponibles']
        estatus = request.form['estatus']
        salario = request.form.get('salario', '')
        tipo_contrato = request.form['tipo_contrato']
        modalidad = request.form['modalidad']
        ubicacion = request.form.get('ubicacion', '')
        experiencia_requerida = request.form['experiencia_requerida']
        beneficios = request.form.get('beneficios', '')
        fecha_cierre = request.form.get('fecha_cierre', '')
        
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            

            fecha_cierre_date = datetime.strptime(fecha_cierre, '%Y-%m-%d').date() if fecha_cierre else None
            
            cursor.execute("""
                INSERT INTO Vacantes (
                    EmpresaID, Puesto, GradoEstudios, Resumen, Plazas, PlazasDisponibles,
                    Estatus, Salario, TipoContrato, Modalidad, Ubicacion, 
                    ExperienciaRequerida, Beneficios, FechaCierre
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                empresa_id, puesto, grado_estudios, resumen, plazas, plazas_disponibles,
                estatus, salario, tipo_contrato, modalidad, ubicacion,
                experiencia_requerida, beneficios, fecha_cierre_date
            ))
            
            conn.commit()
            conn.close()
            
            flash('Vacante creada exitosamente', 'success')
            return redirect(url_for('admin_vacantes'))
        except Exception as e:
            flash(f'Error al crear la vacante: {str(e)}', 'error')
    

    return render_template('admin/crear_vacante.html')

@app.route('/admin/vacantes/editar/<int:vacante_id>', methods=['GET', 'POST'])
def editar_vacante(vacante_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if request.method == 'POST':
 
        empresa_id = request.form['empresa_id']
        puesto = request.form['puesto']
        grado_estudios = request.form['grado_estudios']
        resumen = request.form['resumen']
        plazas = request.form['plazas']
        plazas_disponibles = request.form['plazas_disponibles']
        estatus = request.form['estatus']
        salario = request.form.get('salario', '')
        tipo_contrato = request.form['tipo_contrato']
        modalidad = request.form['modalidad']
        ubicacion = request.form.get('ubicacion', '')
        experiencia_requerida = request.form['experiencia_requerida']
        beneficios = request.form.get('beneficios', '')
        fecha_cierre = request.form.get('fecha_cierre', '')
        comentarios_admin = request.form.get('comentarios_admin', '')
        
        try:
     
            fecha_cierre_date = datetime.strptime(fecha_cierre, '%Y-%m-%d').date() if fecha_cierre else None
            
            cursor.execute("""
                UPDATE Vacantes SET
                    EmpresaID = ?, Puesto = ?, GradoEstudios = ?, Resumen = ?,
                    Plazas = ?, PlazasDisponibles = ?, Estatus = ?, Salario = ?,
                    TipoContrato = ?, Modalidad = ?, Ubicacion = ?,
                    ExperienciaRequerida = ?, Beneficios = ?, FechaCierre = ?,
                    ComentariosAdmin = ?
                WHERE VacanteID = ?
            """, (
                empresa_id, puesto, grado_estudios, resumen,
                plazas, plazas_disponibles, estatus, salario,
                tipo_contrato, modalidad, ubicacion,
                experiencia_requerida, beneficios, fecha_cierre_date,
                comentarios_admin, vacante_id
            ))
            
            conn.commit()
            conn.close()
            
            flash('Vacante actualizada exitosamente', 'success')
            return redirect(url_for('admin_vacantes'))
        except Exception as e:
            flash(f'Error al actualizar la vacante: {str(e)}', 'error')

    cursor.execute("SELECT * FROM Vacantes WHERE VacanteID = ?", (vacante_id,))
    vacante = cursor.fetchone()
    
    if not vacante:
        flash('Vacante no encontrada', 'error')
        return redirect(url_for('admin_vacantes'))
    

    columns = [column[0] for column in cursor.description]
    vacante_dict = CaseInsensitiveDict(zip(columns, vacante))
    

    if vacante_dict['FechaCierre']:
        vacante_dict['FechaCierre'] = vacante_dict['FechaCierre'].strftime('%Y-%m-%d')
    if vacante_dict['FechaPublicacion']:
        vacante_dict['FechaPublicacion'] = vacante_dict['FechaPublicacion'].strftime('%Y-%m-%dT%H:%M')
    if vacante_dict['FechaAprobacion']:
        vacante_dict['FechaAprobacion'] = vacante_dict['FechaAprobacion'].strftime('%Y-%m-%dT%H:%M')
    
    conn.close()
    return render_template('admin/editar_vacante.html', vacante=vacante_dict)

@app.route('/admin/vacantes/eliminar/<int:vacante_id>', methods=['POST'])
def eliminar_vacante(vacante_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
     
        cursor.execute("UPDATE Vacantes SET Estatus = 'cerrada' WHERE VacanteID = ?", (vacante_id,))
        
        conn.commit()
        conn.close()
        
        flash('Vacante cerrada exitosamente', 'success')
    except Exception as e:
        flash(f'Error al cerrar la vacante: {str(e)}', 'error')
        if 'conn' in locals():
            conn.rollback()
    
    return redirect(url_for('admin_vacantes'))




# ==================== SISTEMA DE MENSAJERÍA ====================

@app.route('/conversacion/<int:vacante_id>/<int:candidato_id>')
@login_required
def ver_conversacion(vacante_id, candidato_id):
    """Ver conversación entre empresa y candidato"""
    usuario_actual = get_usuario_actual()
    
    # Verificar que el usuario tenga acceso a esta conversación
    if usuario_actual['TipoUsuario'] == 'empresa':
        empresa = get_empresa_actual()
        if not empresa:
            flash('Perfil de empresa no encontrado', 'error')
            return redirect(url_for('login'))
        
        # Verificar que la vacante pertenece a esta empresa
        vacante = execute_query(
            "SELECT * FROM Vacantes WHERE VacanteID = ? AND EmpresaID = ?",
            (vacante_id, empresa['EmpresaID'])
        )
        if not vacante:
            flash('No tienes acceso a esta conversación', 'error')
            return redirect(url_for('empresa_dashboard'))
    
    elif usuario_actual['TipoUsuario'] == 'candidato':
        candidato = get_candidato_actual()
        if not candidato:
            flash('Perfil de candidato no encontrado', 'error')
            return redirect(url_for('login'))
        
        # Verificar que el candidato es el correcto
        if candidato['CandidatoID'] != candidato_id:
            flash('No tienes acceso a esta conversación', 'error')
            return redirect(url_for('candidato_dashboard'))
    else:
        flash('Acceso no autorizado', 'error')
        return redirect(url_for('index'))
    
    # Obtener o crear conversación
    conversacion = execute_query(
        """SELECT * FROM Conversaciones 
           WHERE VacanteID = ? AND CandidatoID = ?""",
        (vacante_id, candidato_id)
    )
    
    if not conversacion:
        # Crear nueva conversación
        empresa_id = execute_query(
            "SELECT EmpresaID FROM Vacantes WHERE VacanteID = ?",
            (vacante_id,)
        )[0]['EmpresaID']
        
        execute_query(
            """INSERT INTO Conversaciones (VacanteID, CandidatoID, EmpresaID)
               VALUES (?, ?, ?)""",
            (vacante_id, candidato_id, empresa_id),
            fetch=False
        )
        
        conversacion = execute_query(
            """SELECT * FROM Conversaciones 
               WHERE VacanteID = ? AND CandidatoID = ?""",
            (vacante_id, candidato_id)
        )
    
    conversacion = conversacion[0]
    
    # Obtener mensajes
    mensajes = execute_query(
        """SELECT m.*,
           m.FechaEnvio AT TIME ZONE 'UTC' AT TIME ZONE 'America/Mexico_City' AS FechaEnvio,
           m.FechaLectura AT TIME ZONE 'UTC' AT TIME ZONE 'America/Mexico_City' AS FechaLectura,
           CASE 
               WHEN m.RemitenteTipo = 'candidato' THEN c.Nombre
               WHEN m.RemitenteTipo = 'empresa' THEN e.Nombre
           END as RemitenteNombre
           FROM Mensajes m
           LEFT JOIN Candidatos c ON m.RemitenteTipo = 'candidato' AND m.RemitenteID = c.CandidatoID
           LEFT JOIN Empresas e ON m.RemitenteTipo = 'empresa' AND m.RemitenteID = e.EmpresaID
           WHERE m.ConversacionID = ?
           ORDER BY m.FechaEnvio ASC""",
        (conversacion['ConversacionID'],)
    )
    
    # Marcar mensajes como leídos
    if usuario_actual['TipoUsuario'] == 'empresa':
        execute_query(
            """UPDATE Mensajes SET Leido = 1, FechaLectura = GETDATE()
               WHERE ConversacionID = ? AND RemitenteTipo = 'candidato' AND Leido = 0""",
            (conversacion['ConversacionID'],),
            fetch=False
        )
    else:
        execute_query(
            """UPDATE Mensajes SET Leido = 1, FechaLectura = GETDATE()
               WHERE ConversacionID = ? AND RemitenteTipo = 'empresa' AND Leido = 0""",
            (conversacion['ConversacionID'],),
            fetch=False
        )
    
    # Obtener información de la vacante
    vacante = execute_query(
        "SELECT v.*, e.Nombre as EmpresaNombre FROM Vacantes v "
        "JOIN Empresas e ON v.EmpresaID = e.EmpresaID "
        "WHERE v.VacanteID = ?",
        (vacante_id,)
    )[0]
    
    candidato = execute_query(
        "SELECT * FROM Candidatos WHERE CandidatoID = ?",
        (candidato_id,)
    )[0]
    
    return render_template('mensajeria/conversacion.html',
                         conversacion=conversacion,
                         mensajes=mensajes,
                         vacante=vacante,
                         candidato=candidato,
                         usuario_actual=usuario_actual)

@app.route('/enviar_mensaje', methods=['POST'])
@login_required
def enviar_mensaje():
    """Enviar un mensaje en la conversación"""
    conversacion_id = request.form.get('conversacion_id')
    mensaje = request.form.get('mensaje', '').strip()
    
    if not mensaje:
        flash('El mensaje no puede estar vacío', 'error')
        return redirect(request.referrer)
    
    if len(mensaje) > 2000:
        flash('El mensaje no puede exceder los 2000 caracteres', 'error')
        return redirect(request.referrer)
    
    usuario_actual = get_usuario_actual()
    
    # Obtener conversación
    conversacion = execute_query(
        "SELECT * FROM Conversaciones WHERE ConversacionID = ?",
        (conversacion_id,)
    )
    
    if not conversacion:
        flash('Conversación no encontrada', 'error')
        return redirect(request.referrer)
    
    conversacion = conversacion[0]
    
    # Verificar permisos
    if usuario_actual['TipoUsuario'] == 'empresa':
        empresa = get_empresa_actual()
        if not empresa or empresa['EmpresaID'] != conversacion['EmpresaID']:
            flash('No tienes permisos para enviar mensajes aquí', 'error')
            return redirect(request.referrer)
        remitente_id = empresa['EmpresaID']
        remitente_tipo = 'empresa'
    elif usuario_actual['TipoUsuario'] == 'candidato':
        candidato = get_candidato_actual()
        if not candidato or candidato['CandidatoID'] != conversacion['CandidatoID']:
            flash('No tienes permisos para enviar mensajes aquí', 'error')
            return redirect(request.referrer)
        remitente_id = candidato['CandidatoID']
        remitente_tipo = 'candidato'
    else:
        flash('Acceso no autorizado', 'error')
        return redirect(request.referrer)
    
    # Guardar mensaje
    try:
        execute_query(
            """INSERT INTO Mensajes (ConversacionID, RemitenteID, RemitenteTipo, Mensaje)
               VALUES (?, ?, ?, ?)""",
            (conversacion_id, remitente_id, remitente_tipo, mensaje),
            fetch=False
        )
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True})
        flash('Mensaje enviado correctamente', 'success')
    except Exception as e:
        current_app.logger.error(f"Error al enviar mensaje: {str(e)}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'error': 'Error al enviar el mensaje'}), 500
        flash('Error al enviar el mensaje', 'error')
    
    return redirect(request.referrer)

@app.route('/mis_conversaciones')
@login_required
def mis_conversaciones():
    """Ver todas las conversaciones del usuario"""
    usuario_actual = get_usuario_actual()
    
    if usuario_actual['TipoUsuario'] == 'empresa':
        empresa = get_empresa_actual()
        if not empresa:
            flash('Perfil de empresa no encontrado', 'error')
            return redirect(url_for('login'))
        
        conversaciones = execute_query(
            """SELECT c.*, 
               v.Puesto as VacantePuesto,
               cand.Nombre as CandidatoNombre,
               cand.ApellidoPaterno as CandidatoApellido,
               (SELECT Mensaje FROM Mensajes 
                WHERE ConversacionID = c.ConversacionID 
                ORDER BY FechaEnvio DESC LIMIT 1) as UltimoMensaje,
               (SELECT FechaEnvio AT TIME ZONE 'UTC' AT TIME ZONE 'America/Mexico_City' FROM Mensajes 
                WHERE ConversacionID = c.ConversacionID 
                ORDER BY FechaEnvio DESC LIMIT 1) as UltimoMensajeFecha,
               (SELECT COUNT(*) FROM Mensajes 
                WHERE ConversacionID = c.ConversacionID 
                AND RemitenteTipo = 'candidato' 
                AND Leido = 0) as NoLeidos
               FROM Conversaciones c
               JOIN Vacantes v ON c.VacanteID = v.VacanteID
               JOIN Candidatos cand ON c.CandidatoID = cand.CandidatoID
               WHERE c.EmpresaID = ?
               ORDER BY UltimoMensajeFecha DESC""",
            (empresa['EmpresaID'],)
        )
        
    elif usuario_actual['TipoUsuario'] == 'candidato':
        candidato = get_candidato_actual()
        if not candidato:
            flash('Perfil de candidato no encontrado', 'error')
            return redirect(url_for('login'))
        
        conversaciones = execute_query(
            """SELECT c.*, 
               v.Puesto as VacantePuesto,
               e.Nombre as EmpresaNombre,
               (SELECT Mensaje FROM Mensajes 
                WHERE ConversacionID = c.ConversacionID 
                ORDER BY FechaEnvio DESC LIMIT 1) as UltimoMensaje,
               (SELECT FechaEnvio AT TIME ZONE 'UTC' AT TIME ZONE 'America/Mexico_City' FROM Mensajes 
                WHERE ConversacionID = c.ConversacionID 
                ORDER BY FechaEnvio DESC LIMIT 1) as UltimoMensajeFecha,
               (SELECT COUNT(*) FROM Mensajes 
                WHERE ConversacionID = c.ConversacionID 
                AND RemitenteTipo = 'empresa' 
                AND Leido = 0) as NoLeidos
               FROM Conversaciones c
               JOIN Vacantes v ON c.VacanteID = v.VacanteID
               JOIN Empresas e ON c.EmpresaID = e.EmpresaID
               WHERE c.CandidatoID = ?
               ORDER BY UltimoMensajeFecha DESC""",
            (candidato['CandidatoID'],)
        )
    else:
        flash('Acceso no autorizado', 'error')
        return redirect(url_for('index'))
    
    return render_template('mensajeria/conversaciones.html', 
                         conversaciones=conversaciones,
                         tipo_usuario=usuario_actual['TipoUsuario'])

@app.route('/marcar_leidos/<int:conversacion_id>')
@login_required
def marcar_leidos(conversacion_id):
    """Marcar todos los mensajes como leídos"""
    usuario_actual = get_usuario_actual()
    
    if usuario_actual['TipoUsuario'] == 'empresa':
        remitente_tipo = 'candidato'
    elif usuario_actual['TipoUsuario'] == 'candidato':
        remitente_tipo = 'empresa'
    else:
        flash('Acceso no autorizado', 'error')
        return redirect(request.referrer)
    
    execute_query(
        """UPDATE Mensajes SET Leido = 1, FechaLectura = GETDATE()
           WHERE ConversacionID = ? AND RemitenteTipo = ? AND Leido = 0""",
        (conversacion_id, remitente_tipo),
        fetch=False
    )
    
    return redirect(request.referrer)

@app.route('/obtener_no_leidos')
@login_required
def obtener_no_leidos():
    """Obtener cantidad de mensajes no leídos (para AJAX)"""
    usuario_actual = get_usuario_actual()
    
    if usuario_actual['TipoUsuario'] == 'empresa':
        empresa = get_empresa_actual()
        if not empresa:
            return jsonify({'no_leidos': 0})
        
        no_leidos = execute_query(
            """SELECT COUNT(*) as Total FROM Mensajes m
               JOIN Conversaciones c ON m.ConversacionID = c.ConversacionID
               WHERE c.EmpresaID = ? AND m.RemitenteTipo = 'candidato' AND m.Leido = 0""",
            (empresa['EmpresaID'],)
        )
        
    elif usuario_actual['TipoUsuario'] == 'candidato':
        candidato = get_candidato_actual()
        if not candidato:
            return jsonify({'no_leidos': 0})
        
        no_leidos = execute_query(
            """SELECT COUNT(*) as Total FROM Mensajes m
               JOIN Conversaciones c ON m.ConversacionID = c.ConversacionID
               WHERE c.CandidatoID = ? AND m.RemitenteTipo = 'empresa' AND m.Leido = 0""",
            (candidato['CandidatoID'],)
        )
    else:
        return jsonify({'no_leidos': 0})
    
    return jsonify({'no_leidos': no_leidos[0]['Total'] if no_leidos else 0})




# ==================== API COMPARTIDA: WEB Y APP MÓVIL ====================
# Agregar CORS para permitir peticiones desde React
from flask_cors import CORS

# Las apps nativas no están sujetas a CORS. Esta lista protege a los clientes web.
api_origins = [
    origin.strip()
    for origin in os.getenv(
        'CORS_ORIGINS',
        'http://localhost:3000,http://127.0.0.1:3000'
    ).split(',')
    if origin.strip()
]
CORS(app, resources={r"/api/*": {"origins": api_origins}})


def api_error(message, status=400):
    return jsonify({'error': message}), status


def validate_api_fields(data, required=(), phone_fields=(), date_fields=(), max_lengths=None):
    """Validación común del servidor para todos los formularios que escriben en BD."""
    if not isinstance(data, dict):
        return 'El cuerpo de la solicitud debe ser un objeto JSON.'
    missing = [field for field in required if not str(data.get(field, '')).strip()]
    if missing:
        return f"Campos obligatorios: {', '.join(missing)}."
    for field in phone_fields:
        value = str(data.get(field, '') or '').strip()
        if value and not re.fullmatch(r'\d{10}', value):
            return f'{field} debe contener exactamente 10 dígitos.'
    for field in date_fields:
        value = data.get(field)
        if value:
            try:
                datetime.strptime(str(value)[:10], '%Y-%m-%d')
            except ValueError:
                return f'{field} debe tener formato AAAA-MM-DD.'
    for field, limit in (max_lengths or {}).items():
        if len(str(data.get(field, '') or '')) > limit:
            return f'{field} no puede exceder {limit} caracteres.'
    return None


def validate_date_order(data, start_field, end_field):
    start, end = data.get(start_field), data.get(end_field)
    if start and end and str(end)[:10] < str(start)[:10]:
        return f'{end_field} no puede ser anterior a {start_field}.'
    return None


def current_api_user():
    user_id = int(get_jwt_identity())
    users = execute_query(
        "SELECT UsuarioID, Email, TipoUsuario, Activo FROM Usuarios WHERE UsuarioID = ?",
        (user_id,)
    )
    return users[0] if users else None


def current_candidate_id():
    user = current_api_user()
    if not user or user['TipoUsuario'] != 'candidato':
        return None
    rows = execute_query(
        "SELECT CandidatoID FROM Candidatos WHERE UsuarioID = ?",
        (user['UsuarioID'],)
    )
    return rows[0]['CandidatoID'] if rows else None


@app.route('/api/v1/health', methods=['GET'])
def api_health():
    """Health check de aplicación y conectividad con PostgreSQL."""
    try:
        execute_query("SELECT 1 AS ok")
        DB_HEALTH.set(1)
        return jsonify({'status': 'ok', 'database': 'ok'})
    except Exception:
        DB_HEALTH.set(0)
        return jsonify({'status': 'error', 'database': 'unavailable'}), 503


@app.route('/metrics', methods=['GET'])
def prometheus_metrics():
    """Métricas internas; el gateway público bloquea esta ruta."""
    return current_app.response_class(generate_latest(), mimetype=CONTENT_TYPE_LATEST)


@app.route('/api/v1/chatbot', methods=['POST'])
@jwt_required()
def api_chatbot():
    """Expone en móvil el mismo motor inteligente y análisis de compatibilidad de la web."""
    data = request.get_json(silent=True) or {}
    message = str(data.get('mensaje', '')).strip()
    if not message:
        return api_error('Escribe un mensaje.')
    if len(message) > MAX_MENSAJE_LEN:
        return api_error(f'El mensaje no puede exceder {MAX_MENSAJE_LEN} caracteres.')

    user = current_api_user()
    if not user or not user['Activo']:
        return api_error('Usuario no disponible.', 401)
    return jsonify({
        'respuesta': chatbot.procesar_mensaje(message, user['UsuarioID']),
        'sugerencias': chatbot.obtener_sugerencias(),
    })


@app.route('/api/v1/auth/login', methods=['POST'])
def api_login():
    data = request.get_json(silent=True) or {}
    email = str(data.get('email', '')).strip().lower()
    password = str(data.get('password', ''))

    if not email or not password:
        return api_error('Correo y contraseña son obligatorios.')

    users = execute_query(
        """SELECT UsuarioID, Email, PasswordHash, TipoUsuario, Activo
           FROM Usuarios WHERE LOWER(Email) = ?""",
        (email,)
    )
    if not users or not users[0]['Activo'] or not check_password_hash(users[0]['PasswordHash'], password):
        return api_error('Credenciales incorrectas.', 401)

    user = users[0]
    identity = str(user['UsuarioID'])
    return jsonify({
        'access_token': create_access_token(
            identity=identity,
            additional_claims={'tipo': user['TipoUsuario']}
        ),
        'refresh_token': create_refresh_token(identity=identity),
        'usuario': {
            'UsuarioID': user['UsuarioID'],
            'Email': user['Email'],
            'TipoUsuario': user['TipoUsuario'],
        }
    })


@app.route('/api/v1/auth/refresh', methods=['POST'])
@jwt_required(refresh=True)
def api_refresh():
    user = current_api_user()
    if not user or not user['Activo']:
        return api_error('Usuario no disponible.', 401)
    return jsonify({
        'access_token': create_access_token(
            identity=str(user['UsuarioID']),
            additional_claims={'tipo': user['TipoUsuario']}
        )
    })


@app.route('/api/v1/auth/register', methods=['POST'])
def api_register():
    data = request.get_json(silent=True) or {}
    nombre = str(data.get('nombre', '')).strip()
    apellido = str(data.get('apellido', '')).strip()
    email = str(data.get('email', '')).strip().lower()
    password = str(data.get('password', ''))

    if not nombre or not apellido or not email or not password:
        return api_error('Nombre, apellido, correo y contraseña son obligatorios.')
    if not email.endswith('@upq.edu.mx'):
        return api_error('El registro de candidatos requiere un correo institucional @upq.edu.mx.')
    if not re.fullmatch(r'[^@\s]+@[^@\s]+\.[^@\s]+', email):
        return api_error('El correo electrónico no es válido.')
    if len(password) < 8 or not re.search(r'[A-Za-z]', password) or not re.search(r'\d', password):
        return api_error('La contraseña debe tener al menos 8 caracteres, una letra y un número.')

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT 1 FROM Usuarios WHERE LOWER(Email) = ?", (email,))
        if cursor.fetchone():
            return api_error('El correo ya está registrado.', 409)

        cursor.execute(
            """INSERT INTO Usuarios (Email, PasswordHash, TipoUsuario, Activo)
               VALUES (?, ?, 'candidato', TRUE)
               RETURNING UsuarioID""",
            (email, generate_password_hash(password))
        )
        user_id = cursor.fetchone()[0]
        cursor.execute(
            """INSERT INTO Candidatos
               (CandidatoID, UsuarioID, Nombre, ApellidoPaterno)
               VALUES (?, ?, ?, ?)""",
            (user_id, user_id, nombre, apellido)
        )
        conn.commit()
    except Exception as exc:
        conn.rollback()
        current_app.logger.error(f"API register error: {exc}")
        return api_error('No fue posible crear la cuenta.', 500)
    finally:
        cursor.close()
        conn.close()

    # El registro móvil comparte el mismo correo HTML de bienvenida que la web.
    # Si SMTP falla, la cuenta permanece creada y el usuario puede iniciar sesión.
    full_name = f'{nombre} {apellido}'.strip()
    welcome_email_sent = enviar_correo_bienvenida(email, full_name, 'candidato')
    if not welcome_email_sent:
        current_app.logger.warning(
            'La cuenta móvil fue creada, pero no se pudo enviar el correo de bienvenida a %s.',
            email
        )

    identity = str(user_id)
    return jsonify({
        'access_token': create_access_token(
            identity=identity,
            additional_claims={'tipo': 'candidato'}
        ),
        'refresh_token': create_refresh_token(identity=identity),
        'usuario': {
            'UsuarioID': user_id,
            'Email': email,
            'TipoUsuario': 'candidato',
        },
        'welcome_email_sent': welcome_email_sent,
        'welcome_email': email,
    }), 201


@app.route('/api/v1/auth/me', methods=['GET'])
@jwt_required()
def api_me():
    user = current_api_user()
    if not user or not user['Activo']:
        return api_error('Usuario no disponible.', 401)
    return jsonify(user)


@app.route('/api/v1/auth/password/forgot', methods=['POST'])
def api_password_forgot():
    """Envía un código temporal sin revelar si el correo está registrado."""
    data = request.get_json(silent=True) or {}
    email = str(data.get('email', '')).strip().lower()
    if not re.fullmatch(r'[^@\s]+@[^@\s]+\.[^@\s]+', email):
        return api_error('Ingresa un correo electrónico válido.')

    users = execute_query(
        "SELECT UsuarioID, Email FROM Usuarios WHERE LOWER(Email) = ?",
        (email,)
    )
    if users:
        code = generar_codigo_recuperacion()
        execute_query(
            """UPDATE Usuarios SET ResetToken = ?,
               ResetTokenExpira = CURRENT_TIMESTAMP + INTERVAL '10 minutes'
               WHERE UsuarioID = ?""",
            (encrypt_sensitive(code), users[0]['UsuarioID']),
            fetch=False
        )
        if not enviar_codigo_recuperacion(users[0]['Email'], code):
            current_app.logger.error('No fue posible enviar el código de recuperación móvil.')
            return api_error('No fue posible enviar el código. Intenta más tarde.', 503)

    return jsonify({
        'message': 'Si el correo está registrado, recibirás un código de 6 dígitos.'
    })


@app.route('/api/v1/auth/password/verify', methods=['POST'])
def api_password_verify():
    """Intercambia el código de correo por un JWT de restablecimiento de 10 minutos."""
    data = request.get_json(silent=True) or {}
    email = str(data.get('email', '')).strip().lower()
    code = str(data.get('code', '')).strip()
    if not re.fullmatch(r'\d{6}', code):
        return api_error('El código debe contener 6 dígitos.')

    users = execute_query(
        """SELECT UsuarioID, ResetToken FROM Usuarios
           WHERE LOWER(Email) = ? AND ResetTokenExpira IS NOT NULL
             AND ResetTokenExpira >= CURRENT_TIMESTAMP""",
        (email,)
    )
    if not users or not secure_equals_encrypted(users[0]['ResetToken'], code):
        return api_error('El código es incorrecto o ya expiró.', 401)

    reset_token = create_access_token(
        identity=str(users[0]['UsuarioID']),
        additional_claims={'purpose': 'password_reset'},
        expires_delta=timedelta(minutes=10)
    )
    return jsonify({'reset_token': reset_token})


@app.route('/api/v1/auth/password/reset', methods=['POST'])
@jwt_required()
def api_password_reset():
    """Cambia la contraseña únicamente con un JWT emitido para recuperación."""
    if get_jwt().get('purpose') != 'password_reset':
        return api_error('Token de recuperación no válido.', 403)

    data = request.get_json(silent=True) or {}
    password = str(data.get('password', ''))
    if (len(password) < 8 or not re.search(r'[A-Z]', password)
            or not re.search(r'[a-z]', password) or not re.search(r'\d', password)):
        return api_error(
            'La contraseña debe tener 8 caracteres, mayúscula, minúscula y número.'
        )

    execute_query(
        """UPDATE Usuarios SET PasswordHash = ?, ResetToken = NULL,
           ResetTokenExpira = NULL WHERE UsuarioID = ?""",
        (generate_password_hash(password), int(get_jwt_identity())),
        fetch=False
    )
    return jsonify({'message': 'Contraseña actualizada correctamente.'})


@app.route('/api/v1/perfil', methods=['GET', 'PUT'])
@jwt_required()
def api_perfil():
    user = current_api_user()
    if not user or user['TipoUsuario'] != 'candidato':
        return api_error('Este recurso es exclusivo para candidatos.', 403)

    profiles = execute_query(
        """SELECT c.*, u.Email
           FROM Candidatos c
           JOIN Usuarios u ON c.UsuarioID = u.UsuarioID
           WHERE c.UsuarioID = ?""",
        (user['UsuarioID'],)
    )
    if not profiles:
        return api_error('Perfil no encontrado.', 404)

    if request.method == 'GET':
        profile = dict(profiles[0])
        photo = profile.get('FotoPerfil')
        if photo and not str(photo).startswith(('data:image/', 'https://', 'http://')):
            profile['FotoPerfil'] = url_for('static', filename=f'uploads/{photo}', _external=True)
        return jsonify(profile)

    data = request.get_json(silent=True) or {}
    def pick(*keys, default=''):
        for key in keys:
            if key in data:
                return data[key]
        return default

    nombre = str(pick('nombre', 'Nombre')).strip()
    apellido = str(pick('apellido', 'ApellidoPaterno')).strip()
    email = str(pick('email', 'Email', default=user['Email'])).strip().lower()
    telefono = str(pick('telefono', 'Telefono')).strip()
    direccion = str(pick('direccion', 'Direccion')).strip()
    puesto = str(pick('puestoSolicitado', 'PuestoSolicitado')).strip()
    apellido_materno = str(pick('apellidoMaterno', 'ApellidoMaterno')).strip()
    fecha_nacimiento = pick('fechaNacimiento', 'FechaNacimiento') or None
    sexo = str(pick('sexo', 'Sexo')).strip()
    estado_civil = str(pick('estadoCivil', 'EstadoCivil')).strip()
    nacionalidad = str(pick('nacionalidad', 'Nacionalidad')).strip()
    rfc = str(pick('rfc', 'RFC')).strip().upper()
    modalidad = str(pick('modalidad', 'ModalidadTrabajo')).strip()
    puesto_actual = str(pick('puestoActual', 'PuestoActual')).strip()
    resumen = str(pick('resumen', 'ResumenProfesional')).strip()
    foto_perfil = pick('fotoPerfil', 'FotoPerfil', default=profiles[0].get('FotoPerfil'))
    if foto_perfil and not (str(foto_perfil).startswith('data:image/') or str(foto_perfil).startswith('https://')):
        return api_error('La foto de perfil no tiene un formato válido.')
    if foto_perfil and len(str(foto_perfil)) > 2_800_000:
        return api_error('La foto de perfil no puede exceder 2 MB.')
    reubicacion = bool(pick('reubicacion', 'Reubicacion', default=False))
    viajar = bool(pick('viajar', 'Viajar', default=False))
    licencia = bool(pick('licencia', 'LicenciaConducir', default=False))

    if not nombre or not apellido or not email:
        return api_error('Nombre, apellido y correo son obligatorios.')
    if not re.fullmatch(r'[^@\s]+@[^@\s]+\.[^@\s]+', email):
        return api_error('El correo electrónico no es válido.')
    if not email.endswith('@upq.edu.mx'):
        return api_error('Los candidatos deben usar su correo institucional @upq.edu.mx.')
    if telefono and not re.fullmatch(r'\d{10}', telefono):
        return api_error('El teléfono debe contener 10 dígitos.')
    if rfc and not re.fullmatch(r'[A-ZÑ&]{3,4}\d{6}[A-Z0-9]{3}', rfc):
        return api_error('El RFC debe tener 12 o 13 caracteres con formato válido.')
    if fecha_nacimiento:
        try:
            parsed_birth_date = datetime.strptime(str(fecha_nacimiento)[:10], '%Y-%m-%d').date()
            if parsed_birth_date > date.today():
                return api_error('La fecha de nacimiento no puede estar en el futuro.')
        except ValueError:
            return api_error('La fecha de nacimiento debe tener formato AAAA-MM-DD.')
    limits = {
        'Nombre': (nombre, 100), 'Apellido': (apellido, 100),
        'Dirección': (direccion, 250), 'Puesto solicitado': (puesto, 150),
        'Resumen profesional': (resumen, 3000),
    }
    for label, (value, limit) in limits.items():
        if len(value) > limit:
            return api_error(f'{label} no puede exceder {limit} caracteres.')

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "SELECT 1 FROM Usuarios WHERE LOWER(Email) = ? AND UsuarioID <> ?",
            (email, user['UsuarioID'])
        )
        if cursor.fetchone():
            return api_error('El correo ya está registrado.', 409)
        cursor.execute(
            "UPDATE Usuarios SET Email = ? WHERE UsuarioID = ?",
            (email, user['UsuarioID'])
        )
        cursor.execute(
            """UPDATE Candidatos SET Nombre = ?, ApellidoPaterno = ?,
               ApellidoMaterno = ?, Telefono = ?, Direccion = ?,
               FechaNacimiento = ?, Sexo = ?, EstadoCivil = ?, Nacionalidad = ?,
               RFC = ?, ModalidadTrabajo = ?, PuestoActual = ?, PuestoSolicitado = ?,
               ResumenProfesional = ?, Reubicacion = ?, Viajar = ?, LicenciaConducir = ?,
               FotoPerfil = ?
               WHERE UsuarioID = ?""",
            (nombre, apellido, apellido_materno or None, telefono or None,
             direccion or None, fecha_nacimiento, sexo or None, estado_civil or None,
             nacionalidad or None, rfc or None, modalidad or None,
             puesto_actual or None, puesto or None, resumen or None,
             reubicacion, viajar, licencia, foto_perfil, user['UsuarioID'])
        )
        conn.commit()
    except Exception as exc:
        conn.rollback()
        current_app.logger.error(f"API profile error: {exc}")
        return api_error('No fue posible actualizar el perfil.', 500)
    finally:
        cursor.close()
        conn.close()

    updated = execute_query(
        """SELECT c.*, u.Email FROM Candidatos c
           JOIN Usuarios u ON c.UsuarioID = u.UsuarioID
           WHERE c.UsuarioID = ?""",
        (user['UsuarioID'],)
    )[0]
    profile = dict(updated)
    photo = profile.get('FotoPerfil')
    if photo and not str(photo).startswith(('data:image/', 'https://', 'http://')):
        profile['FotoPerfil'] = url_for('static', filename=f'uploads/{photo}', _external=True)
    return jsonify(profile)


@app.route('/api/v1/experiencias', methods=['GET', 'POST'])
@jwt_required()
def api_experiencias():
    candidate_id = current_candidate_id()
    if not candidate_id:
        return api_error('Perfil de candidato no encontrado.', 403)
    if request.method == 'GET':
        return jsonify(execute_query(
            "SELECT * FROM ExperienciaLaboral WHERE CandidatoID = ? ORDER BY FechaIngreso DESC",
            (candidate_id,)
        ))
    data = request.get_json(silent=True) or {}
    required = ('Empresa', 'Puesto', 'FechaIngreso', 'Funciones')
    validation_error = validate_api_fields(
        data, required, ('Telefono',), ('FechaIngreso', 'FechaSalida'),
        {'Empresa': 100, 'Puesto': 100, 'Funciones': 2000}
    ) or validate_date_order(data, 'FechaIngreso', 'FechaSalida')
    if validation_error:
        return api_error(validation_error)
    row = execute_query(
        """INSERT INTO ExperienciaLaboral
           (CandidatoID, Empresa, Domicilio, Telefono, Puesto, FechaIngreso,
            FechaSalida, Funciones, SueldoInicial, SueldoFinal, MotivoSeparacion)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) RETURNING ExperienciaID""",
        (candidate_id, data['Empresa'], data.get('Domicilio'), data.get('Telefono'),
         data['Puesto'], data['FechaIngreso'], data.get('FechaSalida'), data['Funciones'],
         data.get('SueldoInicial'), data.get('SueldoFinal'), data.get('MotivoSeparacion'))
    )
    return jsonify({'ExperienciaID': row[0]['ExperienciaID']}), 201


@app.route('/api/v1/experiencias/<int:item_id>', methods=['PUT', 'DELETE'])
@jwt_required()
def api_experiencia_item(item_id):
    candidate_id = current_candidate_id()
    owned = execute_query(
        "SELECT 1 FROM ExperienciaLaboral WHERE ExperienciaID = ? AND CandidatoID = ?",
        (item_id, candidate_id)
    ) if candidate_id else []
    if not owned:
        return api_error('Experiencia no encontrada.', 404)
    if request.method == 'DELETE':
        execute_query("DELETE FROM ExperienciaLaboral WHERE ExperienciaID = ?", (item_id,), fetch=False)
        return jsonify({'message': 'Experiencia eliminada.'})
    data = request.get_json(silent=True) or {}
    validation_error = validate_api_fields(
        data, ('Empresa', 'Puesto', 'FechaIngreso', 'Funciones'), ('Telefono',),
        ('FechaIngreso', 'FechaSalida'), {'Empresa': 100, 'Puesto': 100, 'Funciones': 2000}
    ) or validate_date_order(data, 'FechaIngreso', 'FechaSalida')
    if validation_error:
        return api_error(validation_error)
    execute_query(
        """UPDATE ExperienciaLaboral SET Empresa=?, Domicilio=?, Telefono=?, Puesto=?,
           FechaIngreso=?, FechaSalida=?, Funciones=?, SueldoInicial=?, SueldoFinal=?,
           MotivoSeparacion=? WHERE ExperienciaID=?""",
        (data.get('Empresa'), data.get('Domicilio'), data.get('Telefono'), data.get('Puesto'),
         data.get('FechaIngreso'), data.get('FechaSalida'), data.get('Funciones'),
         data.get('SueldoInicial'), data.get('SueldoFinal'), data.get('MotivoSeparacion'), item_id),
        fetch=False
    )
    return jsonify({'message': 'Experiencia actualizada.'})


@app.route('/api/v1/preparaciones', methods=['GET', 'POST'])
@jwt_required()
def api_preparaciones():
    candidate_id = current_candidate_id()
    if not candidate_id:
        return api_error('Perfil de candidato no encontrado.', 403)
    if request.method == 'GET':
        return jsonify(execute_query(
            "SELECT * FROM PreparacionAcademica WHERE CandidatoID = ? ORDER BY FechaInicio DESC",
            (candidate_id,)
        ))
    data = request.get_json(silent=True) or {}
    required = ('Grado', 'Estatus', 'Institucion', 'Pais', 'FechaInicio')
    validation_error = validate_api_fields(
        data, required, date_fields=('FechaInicio', 'FechaFin'),
        max_lengths={'Grado': 100, 'Estatus': 30, 'Institucion': 150, 'Pais': 80}
    ) or validate_date_order(data, 'FechaInicio', 'FechaFin')
    if validation_error:
        return api_error(validation_error)
    rows = execute_query(
        """INSERT INTO PreparacionAcademica
           (CandidatoID, Grado, Cedula, Estatus, Institucion, Pais, FechaInicio, FechaFin)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?) RETURNING PreparacionID""",
        (candidate_id, data['Grado'], data.get('Cedula'), data['Estatus'],
         data['Institucion'], data['Pais'], data['FechaInicio'], data.get('FechaFin'))
    )
    return jsonify({'PreparacionID': rows[0]['PreparacionID']}), 201


@app.route('/api/v1/preparaciones/<int:item_id>', methods=['PUT', 'DELETE'])
@jwt_required()
def api_preparacion_item(item_id):
    candidate_id = current_candidate_id()
    owned = execute_query(
        "SELECT 1 FROM PreparacionAcademica WHERE PreparacionID=? AND CandidatoID=?",
        (item_id, candidate_id)
    ) if candidate_id else []
    if not owned:
        return api_error('Preparación académica no encontrada.', 404)
    if request.method == 'DELETE':
        execute_query("DELETE FROM PreparacionAcademica WHERE PreparacionID=?", (item_id,), fetch=False)
        return jsonify({'message': 'Preparación eliminada.'})
    data = request.get_json(silent=True) or {}
    validation_error = validate_api_fields(
        data, ('Grado', 'Estatus', 'Institucion', 'Pais', 'FechaInicio'),
        date_fields=('FechaInicio', 'FechaFin'),
        max_lengths={'Grado': 100, 'Estatus': 30, 'Institucion': 150, 'Pais': 80}
    ) or validate_date_order(data, 'FechaInicio', 'FechaFin')
    if validation_error:
        return api_error(validation_error)
    execute_query(
        """UPDATE PreparacionAcademica SET Grado=?, Cedula=?, Estatus=?, Institucion=?,
           Pais=?, FechaInicio=?, FechaFin=? WHERE PreparacionID=?""",
        (data.get('Grado'), data.get('Cedula'), data.get('Estatus'), data.get('Institucion'),
         data.get('Pais'), data.get('FechaInicio'), data.get('FechaFin'), item_id), fetch=False
    )
    return jsonify({'message': 'Preparación actualizada.'})


@app.route('/api/v1/referencias', methods=['GET', 'POST'])
@jwt_required()
def api_referencias():
    candidate_id = current_candidate_id()
    if not candidate_id:
        return api_error('Perfil de candidato no encontrado.', 403)
    if request.method == 'GET':
        return jsonify(execute_query("SELECT * FROM Referencias WHERE CandidatoID=?", (candidate_id,)))
    data = request.get_json(silent=True) or {}
    validation_error = validate_api_fields(
        data, ('Nombre', 'Ocupacion', 'Telefono', 'AnosConocer'), ('Telefono',),
        max_lengths={'Nombre': 150, 'Ocupacion': 100, 'Empresa': 150}
    )
    if validation_error:
        return api_error(validation_error)
    try:
        if not 0 <= int(data['AnosConocer']) <= 80:
            raise ValueError
    except (TypeError, ValueError):
        return api_error('AnosConocer debe ser un número entre 0 y 80.')
    rows = execute_query(
        """INSERT INTO Referencias
           (CandidatoID, Nombre, Ocupacion, Telefono, AnosConocer, Empresa, Documento)
           VALUES (?, ?, ?, ?, ?, ?, ?) RETURNING ReferenciaID""",
        (candidate_id, data['Nombre'], data['Ocupacion'], data['Telefono'],
         data['AnosConocer'], data.get('Empresa'), data.get('Documento'))
    )
    return jsonify({'ReferenciaID': rows[0]['ReferenciaID']}), 201


@app.route('/api/v1/referencias/<int:item_id>', methods=['PUT', 'DELETE'])
@jwt_required()
def api_referencia_item(item_id):
    candidate_id = current_candidate_id()
    owned = execute_query(
        "SELECT 1 FROM Referencias WHERE ReferenciaID=? AND CandidatoID=?",
        (item_id, candidate_id)
    ) if candidate_id else []
    if not owned:
        return api_error('Referencia no encontrada.', 404)
    if request.method == 'DELETE':
        execute_query("DELETE FROM Referencias WHERE ReferenciaID=?", (item_id,), fetch=False)
        return jsonify({'message': 'Referencia eliminada.'})
    data = request.get_json(silent=True) or {}
    validation_error = validate_api_fields(
        data, ('Nombre', 'Ocupacion', 'Telefono', 'AnosConocer'), ('Telefono',),
        max_lengths={'Nombre': 150, 'Ocupacion': 100, 'Empresa': 150}
    )
    if validation_error:
        return api_error(validation_error)
    try:
        if not 0 <= int(data['AnosConocer']) <= 80:
            raise ValueError
    except (TypeError, ValueError):
        return api_error('AnosConocer debe ser un número entre 0 y 80.')
    execute_query(
        """UPDATE Referencias SET Nombre=?, Ocupacion=?, Telefono=?, AnosConocer=?,
           Empresa=?, Documento=COALESCE(?, Documento) WHERE ReferenciaID=?""",
        (data.get('Nombre'), data.get('Ocupacion'), data.get('Telefono'), data.get('AnosConocer'),
         data.get('Empresa'), data.get('Documento'), item_id), fetch=False
    )
    return jsonify({'message': 'Referencia actualizada.'})


@app.route('/api/v1/perfil/habilidades', methods=['GET', 'PUT'])
@jwt_required()
def api_perfil_habilidades():
    candidate_id = current_candidate_id()
    if not candidate_id:
        return api_error('Perfil de candidato no encontrado.', 403)
    if request.method == 'GET':
        habilidades = execute_query("SELECT * FROM Habilidades ORDER BY Nombre")
        competencias = execute_query("SELECT * FROM Competencias ORDER BY Nombre")
        selected_h = execute_query(
            "SELECT HabilidadID FROM CandidatoHabilidades WHERE CandidatoID=?", (candidate_id,)
        )
        selected_c = execute_query(
            "SELECT CompetenciaID FROM CandidatoCompetencias WHERE CandidatoID=?", (candidate_id,)
        )
        return jsonify({
            'habilidades': habilidades,
            'competencias': competencias,
            'habilidadesActuales': [row['HabilidadID'] for row in selected_h],
            'competenciasActuales': [row['CompetenciaID'] for row in selected_c],
        })
    data = request.get_json(silent=True) or {}
    if not isinstance(data.get('habilidades', []), list) or not isinstance(data.get('competencias', []), list):
        return api_error('Habilidades y competencias deben enviarse como listas.')
    try:
        habilidad_ids = list(dict.fromkeys(int(value) for value in data.get('habilidades', [])))
        competencia_ids = list(dict.fromkeys(int(value) for value in data.get('competencias', [])))
    except (TypeError, ValueError):
        return api_error('Los identificadores de habilidades y competencias deben ser enteros.')
    if len(habilidad_ids) > 50 or len(competencia_ids) > 50:
        return api_error('No se permiten más de 50 elementos por catálogo.')
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM CandidatoHabilidades WHERE CandidatoID=?", (candidate_id,))
        for item_id in habilidad_ids:
            cursor.execute(
                "INSERT INTO CandidatoHabilidades(CandidatoID,HabilidadID) VALUES(?,?)",
                (candidate_id, int(item_id))
            )
        cursor.execute("DELETE FROM CandidatoCompetencias WHERE CandidatoID=?", (candidate_id,))
        for item_id in competencia_ids:
            cursor.execute(
                "INSERT INTO CandidatoCompetencias(CandidatoID,CompetenciaID) VALUES(?,?)",
                (candidate_id, int(item_id))
            )
        conn.commit()
    except Exception as exc:
        conn.rollback()
        current_app.logger.error(f"API skills error: {exc}")
        return api_error('No fue posible actualizar las habilidades.', 500)
    finally:
        cursor.close()
        conn.close()
    return jsonify({'message': 'Habilidades actualizadas correctamente.'})


@app.route('/api/v1/postulaciones', methods=['GET'])
@jwt_required()
def api_postulaciones():
    user = current_api_user()
    if not user or user['TipoUsuario'] != 'candidato':
        return api_error('Este recurso es exclusivo para candidatos.', 403)
    rows = execute_query(
        """SELECT p.PostulacionID, p.FechaPostulacion, p.Estatus, p.Comentarios,
                  v.VacanteID, v.Puesto, v.Modalidad, v.Ubicacion, v.Salario,
                  e.Nombre AS EmpresaNombre
           FROM Postulaciones p
           JOIN Candidatos c ON p.CandidatoID = c.CandidatoID
           JOIN Vacantes v ON p.VacanteID = v.VacanteID
           JOIN Empresas e ON v.EmpresaID = e.EmpresaID
           WHERE c.UsuarioID = ?
           ORDER BY p.FechaPostulacion DESC""",
        (user['UsuarioID'],)
    )
    return jsonify(rows)


@app.route('/api/v1/conversaciones', methods=['GET'])
@jwt_required()
def api_conversaciones():
    candidate_id = current_candidate_id()
    if not candidate_id:
        return api_error('Perfil de candidato no encontrado.', 403)
    rows = execute_query(
        """SELECT c.ConversacionID, c.VacanteID, c.CandidatoID,
                  v.Puesto AS VacantePuesto, e.Nombre AS EmpresaNombre,
                  (SELECT Mensaje FROM Mensajes m WHERE m.ConversacionID=c.ConversacionID
                   ORDER BY FechaEnvio DESC LIMIT 1) AS UltimoMensaje,
                  (SELECT FechaEnvio AT TIME ZONE 'UTC' FROM Mensajes m WHERE m.ConversacionID=c.ConversacionID
                   ORDER BY FechaEnvio DESC LIMIT 1) AS UltimoMensajeFecha,
                  (SELECT COUNT(*) FROM Mensajes m WHERE m.ConversacionID=c.ConversacionID
                   AND RemitenteTipo='empresa' AND Leido=FALSE) AS NoLeidos
           FROM Conversaciones c
           JOIN Vacantes v ON c.VacanteID=v.VacanteID
           JOIN Empresas e ON c.EmpresaID=e.EmpresaID
           WHERE c.CandidatoID=? AND c.Activa=TRUE
           ORDER BY UltimoMensajeFecha DESC NULLS LAST""",
        (candidate_id,)
    )
    return jsonify(rows)


@app.route('/api/v1/conversaciones/<int:conversation_id>', methods=['GET'])
@jwt_required()
def api_conversacion_detalle(conversation_id):
    candidate_id = current_candidate_id()
    conversations = execute_query(
        """SELECT c.ConversacionID, c.VacanteID, c.CandidatoID,
                  v.Puesto, e.Nombre AS EmpresaNombre
           FROM Conversaciones c JOIN Vacantes v ON c.VacanteID=v.VacanteID
           JOIN Empresas e ON c.EmpresaID=e.EmpresaID
           WHERE c.ConversacionID=? AND c.CandidatoID=?""",
        (conversation_id, candidate_id)
    ) if candidate_id else []
    if not conversations:
        return api_error('Conversación no encontrada.', 404)
    messages = execute_query(
        """SELECT MensajeID, ConversacionID, RemitenteID, RemitenteTipo, Mensaje,
                  FechaEnvio AT TIME ZONE 'UTC' AS FechaEnvio,
                  Leido, FechaLectura AT TIME ZONE 'UTC' AS FechaLectura
           FROM Mensajes WHERE ConversacionID=? ORDER BY FechaEnvio ASC""",
        (conversation_id,)
    )
    execute_query(
        """UPDATE Mensajes SET Leido=TRUE, FechaLectura=CURRENT_TIMESTAMP
           WHERE ConversacionID=? AND RemitenteTipo='empresa' AND Leido=FALSE""",
        (conversation_id,), fetch=False
    )
    return jsonify({'conversacion': conversations[0], 'mensajes': messages})


@app.route('/api/v1/conversaciones/<int:conversation_id>/mensajes', methods=['POST'])
@jwt_required()
def api_enviar_mensaje(conversation_id):
    candidate_id = current_candidate_id()
    if not candidate_id or not execute_query(
        "SELECT 1 FROM Conversaciones WHERE ConversacionID=? AND CandidatoID=?",
        (conversation_id, candidate_id)
    ):
        return api_error('Conversación no encontrada.', 404)
    message = str((request.get_json(silent=True) or {}).get('mensaje', '')).strip()
    if not message or len(message) > 2000:
        return api_error('El mensaje debe contener entre 1 y 2000 caracteres.')
    created = execute_query(
        """INSERT INTO Mensajes(ConversacionID,RemitenteID,RemitenteTipo,Mensaje)
           VALUES(?,?,'candidato',?)
           RETURNING MensajeID, ConversacionID, RemitenteID, RemitenteTipo, Mensaje,
                     FechaEnvio AT TIME ZONE 'UTC' AS FechaEnvio, Leido, FechaLectura""",
        (conversation_id, candidate_id, message)
    )
    return jsonify(created[0]), 201


@app.route('/api/v1/postulaciones/<int:postulacion_id>', methods=['DELETE'])
@jwt_required()
def api_cancelar_postulacion(postulacion_id):
    user = current_api_user()
    if not user or user['TipoUsuario'] != 'candidato':
        return api_error('Este recurso es exclusivo para candidatos.', 403)
    rows = execute_query(
        """SELECT p.PostulacionID, p.Estatus
           FROM Postulaciones p
           JOIN Candidatos c ON p.CandidatoID = c.CandidatoID
           WHERE p.PostulacionID = ? AND c.UsuarioID = ?""",
        (postulacion_id, user['UsuarioID'])
    )
    if not rows:
        return api_error('Postulación no encontrada.', 404)
    if rows[0]['Estatus'] != 'pendiente':
        return api_error('Sólo se pueden cancelar postulaciones pendientes.', 409)
    execute_query(
        "DELETE FROM Postulaciones WHERE PostulacionID = ?",
        (postulacion_id,),
        fetch=False
    )
    return jsonify({'message': 'Postulación cancelada correctamente.'})


@app.route('/api/v1/vacantes/<int:vacante_id>/postular', methods=['POST'])
@jwt_required()
def api_postular_vacante(vacante_id):
    user = current_api_user()
    if not user or user['TipoUsuario'] != 'candidato':
        return api_error('Este recurso es exclusivo para candidatos.', 403)

    candidates = execute_query(
        "SELECT CandidatoID FROM Candidatos WHERE UsuarioID = ?",
        (user['UsuarioID'],)
    )
    if not candidates:
        return api_error('Perfil de candidato no encontrado.', 404)
    candidate_id = candidates[0]['CandidatoID']

    vacancies = execute_query(
        """SELECT VacanteID FROM Vacantes
           WHERE VacanteID = ? AND Estatus = 'aprobada' AND PlazasDisponibles > 0""",
        (vacante_id,)
    )
    if not vacancies:
        return api_error('La vacante no está disponible.', 404)
    if execute_query(
        "SELECT 1 FROM Postulaciones WHERE VacanteID = ? AND CandidatoID = ?",
        (vacante_id, candidate_id)
    ):
        return api_error('Ya te postulaste a esta vacante.', 409)

    execute_query(
        """INSERT INTO Postulaciones (VacanteID, CandidatoID, Estatus)
           VALUES (?, ?, 'pendiente')""",
        (vacante_id, candidate_id),
        fetch=False
    )
    return jsonify({'message': 'Postulación registrada correctamente.'}), 201

@app.route('/api/v1/vacantes', methods=['GET', 'OPTIONS'])
def api_vacantes():
    """
    Obtener todas las vacantes aprobadas
    ---
    tags:
      - Vacantes
    summary: Lista todas las vacantes activas
    description: Retorna un array con todas las vacantes que están aprobadas y tienen plazas disponibles
    responses:
      200:
        description: Lista de vacantes
        schema:
          type: array
          items:
            type: object
            properties:
              VacanteID:
                type: integer
                description: ID único de la vacante
              Puesto:
                type: string
                description: Título del puesto
              EmpresaNombre:
                type: string
                description: Nombre de la empresa que publicó la vacante
              Modalidad:
                type: string
                description: Modalidad de trabajo (Presencial, Remoto, Híbrido)
              Salario:
                type: string
                description: Rango salarial ofrecido
              Ubicacion:
                type: string
                description: Ubicación del puesto
      500:
        description: Error interno del servidor
    """
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        vacantes = execute_query("""
            SELECT v.VacanteID, v.Puesto, v.Modalidad, v.Salario, v.Ubicacion,
                   v.GradoEstudios, v.TipoContrato, v.Resumen, v.FechaPublicacion,
                   v.ExperienciaRequerida, v.Beneficios, v.PlazasDisponibles,
                   e.Nombre as EmpresaNombre
            FROM Vacantes v
            JOIN Empresas e ON v.EmpresaID = e.EmpresaID
            WHERE v.Estatus = 'aprobada' AND v.PlazasDisponibles > 0
            ORDER BY v.FechaPublicacion DESC
        """)
        return jsonify(vacantes)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/v1/vacantes/<int:vacante_id>', methods=['GET', 'OPTIONS'])
def api_vacante_detalle(vacante_id):
    """
    Obtener detalle de una vacante específica
    ---
    tags:
      - Vacantes
    summary: Obtiene los detalles completos de una vacante
    description: Retorna toda la información de una vacante específica por su ID
    parameters:
      - name: vacante_id
        in: path
        type: integer
        required: true
        description: ID de la vacante a consultar
    responses:
      200:
        description: Detalle de la vacante
        schema:
          type: object
          properties:
            VacanteID:
              type: integer
            Puesto:
              type: string
            EmpresaNombre:
              type: string
            Resumen:
              type: string
            Salario:
              type: string
            Modalidad:
              type: string
            TipoContrato:
              type: string
            ExperienciaRequerida:
              type: string
      404:
        description: Vacante no encontrada
      500:
        description: Error interno del servidor
    """
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        vacante = execute_query("""
            SELECT v.*, e.Nombre as EmpresaNombre, e.Descripcion as EmpresaDescripcion,
                   e.Logo,
                   (SELECT STRING_AGG(h.Nombre, ', ' ORDER BY h.Nombre)
                    FROM VacanteHabilidadesRequeridas vh
                    JOIN Habilidades h ON vh.HabilidadID = h.HabilidadID
                    WHERE vh.VacanteID = v.VacanteID) AS HabilidadesRequeridas,
                   (SELECT STRING_AGG(h.Nombre, ', ' ORDER BY h.Nombre)
                    FROM VacanteHabilidadesOpcionales vh
                    JOIN Habilidades h ON vh.HabilidadID = h.HabilidadID
                    WHERE vh.VacanteID = v.VacanteID) AS HabilidadesOpcionales
            FROM Vacantes v
            JOIN Empresas e ON v.EmpresaID = e.EmpresaID
            WHERE v.VacanteID = ? AND v.Estatus = 'aprobada'
        """, (vacante_id,))
        
        if not vacante:
            return jsonify({'error': 'Vacante no encontrada'}), 404
        
        return jsonify(vacante[0])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/v1/estadisticas', methods=['GET', 'OPTIONS'])
def api_estadisticas_publicas():
    """
    Obtener estadísticas públicas del sistema
    ---
    tags:
      - Estadísticas
    summary: Estadísticas generales del sistema
    description: Retorna métricas como total de vacantes, empresas y distribución por modalidad
    responses:
      200:
        description: Estadísticas generales
        schema:
          type: object
          properties:
            total_vacantes:
              type: integer
              description: Número total de vacantes activas
            total_empresas:
              type: integer
              description: Número total de empresas registradas
            vacantes_por_modalidad:
              type: array
              description: Distribución de vacantes por modalidad
              items:
                type: object
                properties:
                  Modalidad:
                    type: string
                  Total:
                    type: integer
      500:
        description: Error interno del servidor
    """
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        total_vacantes = execute_query("SELECT COUNT(*) as Total FROM Vacantes WHERE Estatus = 'aprobada'")[0]['Total']
        total_empresas = execute_query("SELECT COUNT(*) as Total FROM Empresas")[0]['Total']
        
        modalidades = execute_query("""
            SELECT Modalidad, COUNT(*) as Total 
            FROM Vacantes 
            WHERE Estatus = 'aprobada'
            GROUP BY Modalidad
        """)
        
        return jsonify({
            'total_vacantes': total_vacantes,
            'total_empresas': total_empresas,
            'vacantes_por_modalidad': modalidades
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/v1/habilidades', methods=['GET', 'OPTIONS'])
def api_habilidades():
    """
    Obtener todas las habilidades disponibles
    ---
    tags:
      - Habilidades
    summary: Lista todas las habilidades registradas
    description: Retorna un array con todas las habilidades que pueden agregar los candidatos
    responses:
      200:
        description: Lista de habilidades
        schema:
          type: array
          items:
            type: object
            properties:
              HabilidadID:
                type: integer
                description: ID único de la habilidad
              Nombre:
                type: string
                description: Nombre de la habilidad
      500:
        description: Error interno del servidor
    """
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        habilidades = execute_query("SELECT HabilidadID, Nombre FROM Habilidades ORDER BY Nombre")
        return jsonify(habilidades)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/docs')
def api_docs():
    """Redirigir a la documentación Swagger"""
    return redirect('/apidocs')


@app.context_processor
def inject_current_year():
    return {'current_year': datetime.now().year}

if __name__ == '__main__':

    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    

    try:
        conn = get_db_connection()
        print("✅ Conexión exitosa a la base de datos")
        conn.close()
    except Exception as e:
        print(f"❌ Error al conectar a la base de datos: {str(e)}")
        exit(1)
    
    app.run(
        host=os.getenv('FLASK_HOST', '0.0.0.0'),
        port=int(os.getenv('PORT', '5001')),
        debug=os.getenv('FLASK_DEBUG', 'false').lower() == 'true'
    )
